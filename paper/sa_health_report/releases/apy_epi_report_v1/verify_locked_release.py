from __future__ import annotations

import argparse
import hashlib
import json
import math
import shutil
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


RELEASE_DIR = Path(__file__).resolve().parent


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run1", type=Path, required=True)
    parser.add_argument("--run2", type=Path, required=True)
    args = parser.parse_args()

    run1 = args.run1.resolve()
    run2 = args.run2.resolve()
    rows = []
    for file1 in sorted(run1.rglob("*")):
        if not file1.is_file():
            continue
        rel_path = file1.relative_to(run1)
        if rel_path.parts and rel_path.parts[0] == "replicate_level":
            continue
        if file1.name in {"environment_record.json", "run_manifest.json"}:
            continue
        if file1.suffix.lower() not in {".csv", ".xlsx", ".json"}:
            continue
        file2 = run2 / rel_path
        if not file2.exists():
            rows.append(
                {
                    "file": str(rel_path).replace("\\", "/"),
                    "comparison_type": "missing",
                    "exact_equal": False,
                    "numeric_equal": False,
                    "max_abs_difference": math.nan,
                    "details": "missing from run2",
                }
            )
            continue
        exact_equal = sha256_file(file1) == sha256_file(file2)
        numeric_equal, max_abs, details = compare_file(file1, file2)
        rows.append(
            {
                "file": str(rel_path).replace("\\", "/"),
                "comparison_type": file1.suffix.lower().lstrip("."),
                "exact_equal": exact_equal,
                "numeric_equal": numeric_equal,
                "max_abs_difference": max_abs,
                "details": details,
            }
        )

    verification = pd.DataFrame(rows)
    verification.to_csv(RELEASE_DIR / "reproducibility_verification.csv", index=False)
    summary = {
        "run1": str(run1),
        "run2": str(run2),
        "numeric_outputs_equal": bool(verification["numeric_equal"].all()) if not verification.empty else False,
        "files_compared": int(len(verification)),
        "artifact_roots": [
            artifact_manifest("run1", run1),
            artifact_manifest("run2", run2),
        ],
    }
    write_json(RELEASE_DIR / "reproducibility_verification.json", summary)

    for name in ["input_manifest.json", "environment_record.json", "run_manifest.json"]:
        src = run1 / name
        if src.exists():
            shutil.copy2(src, RELEASE_DIR / name)
    for name in ["comparison_to_existing_report.csv", "comparison_to_existing_report.xlsx"]:
        src = run1 / name
        if src.exists():
            shutil.copy2(src, RELEASE_DIR / name)
    return 0


def compare_file(file1: Path, file2: Path) -> tuple[bool, float, str]:
    suffix = file1.suffix.lower()
    if suffix == ".csv":
        left = pd.read_csv(file1)
        right = pd.read_csv(file2)
        return compare_frame(left, right)
    if suffix == ".xlsx":
        return compare_workbook(file1, file2)
    if suffix == ".json":
        left = json.loads(file1.read_text(encoding="utf-8"))
        right = json.loads(file2.read_text(encoding="utf-8"))
        scrub_manifest(left)
        scrub_manifest(right)
        return (left == right, 0.0 if left == right else math.nan, "json after timestamp/path scrub")
    return (sha256_file(file1) == sha256_file(file2), 0.0, "hash comparison")


def compare_frame(left: pd.DataFrame, right: pd.DataFrame) -> tuple[bool, float, str]:
    if left.shape != right.shape or list(left.columns) != list(right.columns):
        return False, math.nan, "shape or column mismatch"
    max_abs = 0.0
    bool_cols = [
        c for c in left.columns
        if pd.api.types.is_bool_dtype(left[c]) or pd.api.types.is_bool_dtype(right[c])
    ]
    numeric_cols = [
        c for c in left.columns
        if c not in bool_cols
        and (pd.api.types.is_numeric_dtype(left[c]) or pd.api.types.is_numeric_dtype(right[c]))
    ]
    for col in numeric_cols:
        lval = pd.to_numeric(left[col], errors="coerce")
        rval = pd.to_numeric(right[col], errors="coerce")
        diff = (lval - rval).abs()
        same = (lval.isna() & rval.isna()) | (diff == 0)
        if not bool(same.all()):
            max_abs = max(max_abs, float(diff.max(skipna=True)))
            return False, max_abs, f"numeric mismatch in {col}"
    for col in bool_cols:
        if not left[col].fillna("<NA>").equals(right[col].fillna("<NA>")):
            return False, max_abs, f"boolean mismatch in {col}"
    non_numeric = [c for c in left.columns if c not in numeric_cols and c not in bool_cols]
    if non_numeric and not left[non_numeric].equals(right[non_numeric]):
        return False, max_abs, "non-numeric mismatch"
    return True, max_abs, "all compared values equal"


def compare_workbook(file1: Path, file2: Path) -> tuple[bool, float, str]:
    wb1 = load_workbook(file1, data_only=False, read_only=True)
    wb2 = load_workbook(file2, data_only=False, read_only=True)
    try:
        if wb1.sheetnames != wb2.sheetnames:
            return False, math.nan, "sheet mismatch"
        for sheet in wb1.sheetnames:
            rows1 = list(wb1[sheet].iter_rows(values_only=True))
            rows2 = list(wb2[sheet].iter_rows(values_only=True))
            if rows1 != rows2:
                return False, math.nan, f"cell mismatch in {sheet}"
    finally:
        wb1.close()
        wb2.close()
    return True, 0.0, "all workbook cell values/formulas equal"


def artifact_manifest(label: str, root: Path) -> dict[str, Any]:
    files = []
    for path in sorted(root.rglob("*")):
        if path.is_file():
            files.append(
                {
                    "path": str(path.relative_to(root)).replace("\\", "/"),
                    "sha256": sha256_file(path),
                    "bytes": path.stat().st_size,
                }
            )
    return {"label": label, "root": str(root), "files": files}


def scrub_manifest(value: Any) -> None:
    if isinstance(value, dict):
        for key in list(value):
            if key in {"created_at_utc", "recorded_at_utc", "run_date_utc"}:
                value[key] = "<timestamp>"
            elif key in {"output_files"}:
                continue
            else:
                scrub_manifest(value[key])
    elif isinstance(value, list):
        for item in value:
            scrub_manifest(item)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


if __name__ == "__main__":
    raise SystemExit(main())
