from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import platform
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from importlib import metadata
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


RELEASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = RELEASE_DIR.parents[3]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from engine.apy.calibration import calibrate_from_config, disease_multiplier_from_flags
from engine.apy.cohort import infection_probability, preventable_active_risk
from engine.apy.config import build_default_config
from engine.apy.data import load_parameters_from_config
from engine.apy.runner import MODEL_VERSION, run_scenario_with_do_nothing
from engine.apy.simulation import simulate_one_cohort_from_config


N = 1500
N_REPS = 2000
SEED = 1
SCREEN_WINDOW = 2
FOLLOW_HORIZON = 20
TEST_TYPE = "IGRA"
REGIMEN = "3HP"
DEFAULT_LTBI_PREVALENCE = 47 / 624
DEFAULT_ACTIVE_TB_2Y_PREVALENCE = 10 / 770

MAIN_STRATEGIES = ["random", "ltbi", "prevent"]
CURE_AUDIT_STRATEGIES = ["ltbi", "cure"]
STRATEGY_LABELS = {
    "random": "Random screening",
    "ltbi": "LTBI-targeted screening",
    "prevent": "Active TB prevention-targeted screening",
    "cure": "Cure-targeted screening",
}
SCREEN_COUNT_GRID = [
    0,
    50,
    100,
    150,
    200,
    250,
    300,
    350,
    400,
    450,
    500,
    600,
    700,
    800,
    900,
    1000,
    1100,
    1200,
    1300,
    1400,
    1500,
]
PREVALENCE_SCENARIOS = {
    "base_default_LTBI": None,
    "low_LTBI_1pct": 0.01,
}
PREVALENCE_UPDATE_SCREEN_COUNTS = [0, 50, 100, 150, 200, 300, 450, 600, 1000, 1500]
PREVALENCE_UPDATE_CHECKPOINTS = [50, 100, 150, 200, 300]
PREVALENCE_UPDATE_SCENARIOS = [
    ("LTBI_0_5pct", 0.005, "0.5% LTBI"),
    ("LTBI_1pct", 0.01, "1% LTBI"),
    ("LTBI_2pct", 0.02, "2% LTBI"),
    ("LTBI_3pct", 0.03, "3% LTBI"),
    ("LTBI_5pct", 0.05, "5% LTBI"),
    ("LTBI_7_53pct", DEFAULT_LTBI_PREVALENCE, "7.53% LTBI"),
    ("LTBI_10pct", 0.10, "10% LTBI"),
]
CHECKPOINTS = [50, 100, 150, 200, 300]
STOPPING_THRESHOLDS = [0, 1, 2, 3, 4, 5, 10]
CHART_BLOCKS = [
    ("NNS to prevent one active TB case", "NNS_prevent_one_active_TB"),
    ("Number of active TB cases prevented", "nPreventedActiveTB_median"),
    ("False-positive tests", "nFalsePositiveTests_median"),
    ("True-positive tests", "nTruePositiveTests_median"),
    ("False-positive share among test positives", "falsePositiveShareAmongTestPositive"),
    ("False-positive treated", "nFalsePositiveTreated_median"),
    ("False-positive treated share among TPT starts", "falsePositiveTreatedShareAmongCoursesStarted"),
    ("TPT starts", "nTotalCoursesStarted_median"),
    ("TPT completions", "nTotalCoursesCompleted_median"),
]
HISTORICAL_COMPARISON_ABS_TOLERANCE = 1e-12


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--compare-existing", action="store_true")
    args = parser.parse_args()

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    if any(output_dir.iterdir()):
        raise SystemExit(f"Output directory is not empty: {output_dir}")

    write_frozen_scenario_expansions(output_dir)
    input_manifest = build_input_manifest()
    write_json(output_dir / "input_manifest.json", input_manifest)
    write_json(output_dir / "environment_record.json", build_environment_record())

    curve_df, raw_by_scenario = run_curve_scenarios(output_dir)
    targeting_profiles = build_targeting_order_profiles()
    stopping_df = build_stopping_rule_candidates(curve_df, raw_by_scenario)
    cure_audit_df = run_cure_audit(output_dir)
    example_cohort = build_example_cohort_ordering()
    decile_summary = build_rank_decile_summary(example_cohort)
    prevalence_grid_df = run_prevalence_update_grid(output_dir)
    expected_early_yield_df = prevalence_grid_df[
        prevalence_grid_df["screen_count_requested"].isin(PREVALENCE_UPDATE_CHECKPOINTS)
    ].copy()

    curve_df.to_csv(output_dir / "locked_curve_data.csv", index=False)
    to_tidy_long(curve_df).to_csv(output_dir / "locked_tidy_long.csv", index=False)
    write_chart_tables(output_dir, curve_df)
    targeting_profiles.to_csv(output_dir / "locked_targeting_order.csv", index=False)
    stopping_df.to_csv(output_dir / "locked_stopping_rule_candidates.csv", index=False)
    cure_audit_df.to_csv(output_dir / "locked_cure_audit.csv", index=False)
    example_cohort.to_csv(output_dir / "locked_example_cohort.csv", index=False)
    decile_summary.to_csv(output_dir / "locked_rank_decile_summary.csv", index=False)
    prevalence_grid_df.to_csv(output_dir / "locked_prevalence_scenario_grid.csv", index=False)
    expected_early_yield_df.to_csv(output_dir / "locked_expected_early_yield.csv", index=False)
    write_locked_workbook(
        output_dir / "locked_chart_data_workbook.xlsx",
        curve_df,
        cure_audit_df,
        targeting_profiles,
        example_cohort,
        decile_summary,
        stopping_df,
        prevalence_grid_df,
        expected_early_yield_df,
    )

    if args.compare_existing:
        write_existing_report_comparison(output_dir, curve_df, prevalence_grid_df)

    run_manifest = build_run_manifest(output_dir, input_manifest)
    write_json(output_dir / "run_manifest.json", run_manifest)
    return 0


def scenario_config(strategy: str, screen_count: int, ltbi_prevalence: float | None) -> dict[str, Any]:
    config = build_default_config()
    config.update(
        {
            "scenarioLabel": f"Locked APY epi v1 - {strategy} - {screen_count}",
            "N": N,
            "nReps": N_REPS,
            "seed": SEED,
            "screenWindow": SCREEN_WINDOW,
            "followHorizon": FOLLOW_HORIZON,
            "screenCoverage": screen_count / N,
            "screeningStrategy": strategy,
            "testType": TEST_TYPE,
            "regimen": REGIMEN,
            "pStartTPT": 0.85,
            "regimenPComplete": 0.80,
            "regimenADRstop": 0.05,
            "regimenEffFull": 0.85,
        }
    )
    if ltbi_prevalence is not None:
        config["ltbiPrevalence"] = float(ltbi_prevalence)
        config["activeTBPrevalence"] = (
            DEFAULT_ACTIVE_TB_2Y_PREVALENCE * float(ltbi_prevalence) / DEFAULT_LTBI_PREVALENCE
        )
    return config


def run_curve_scenarios(output_dir: Path) -> tuple[pd.DataFrame, dict[tuple[str, str, int], pd.DataFrame]]:
    rows: list[dict[str, Any]] = []
    raw_by_scenario: dict[tuple[str, str, int], pd.DataFrame] = {}
    raw_dir = output_dir / "replicate_level" / "curve"
    raw_dir.mkdir(parents=True, exist_ok=True)
    for prevalence_name, ltbi_prevalence in PREVALENCE_SCENARIOS.items():
        for strategy in MAIN_STRATEGIES:
            for screen_count in SCREEN_COUNT_GRID:
                config = scenario_config(strategy, screen_count, ltbi_prevalence)
                model_out = run_scenario_with_do_nothing(config)
                raw = model_out["results"]["raw"].copy()
                raw.insert(0, "prevalence_scenario", prevalence_name)
                raw.insert(1, "screen_count_requested", screen_count)
                raw.insert(2, "screenCoverage_requested", screen_count / N)
                raw_by_scenario[(prevalence_name, strategy, screen_count)] = raw
                raw.to_csv(raw_dir / f"{prevalence_name}__{strategy}__n_screened_{screen_count}.csv", index=False)
                rows.append(build_curve_row(prevalence_name, strategy, screen_count, config, model_out))
    return pd.DataFrame(rows), raw_by_scenario


def run_cure_audit(output_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    raw_dir = output_dir / "replicate_level" / "cure_audit"
    raw_dir.mkdir(parents=True, exist_ok=True)
    for strategy in CURE_AUDIT_STRATEGIES:
        for screen_count in SCREEN_COUNT_GRID:
            model_out = run_scenario_with_do_nothing(scenario_config(strategy, screen_count, None))
            raw = model_out["results"]["raw"].copy()
            raw.insert(0, "prevalence_scenario", "base_default_LTBI")
            raw.insert(1, "screen_count_requested", screen_count)
            raw.to_csv(raw_dir / f"base_default_LTBI__{strategy}__n_screened_{screen_count}.csv", index=False)
            rows.append(
                {
                    "screeningStrategy": strategy,
                    "screen_count_requested": screen_count,
                    "nScreened_median": summary_value(model_out, "nScreened"),
                    "nTestPositiveNonActive_median": summary_value(model_out, "nTestPositiveNonActive"),
                    "nCuredInfection_median": summary_value(model_out, "nCuredInfection"),
                    "nPreventedActiveTB_median": summary_value(model_out, "nPreventedActiveTB"),
                    "nFalsePositiveTests_median": summary_value(model_out, "nFalsePositiveTests"),
                }
            )
    audit = pd.DataFrame(rows)
    wide = audit.pivot(index="screen_count_requested", columns="screeningStrategy")
    diffs = []
    for metric in [
        "nTestPositiveNonActive_median",
        "nCuredInfection_median",
        "nPreventedActiveTB_median",
        "nFalsePositiveTests_median",
    ]:
        diff = (wide[(metric, "cure")] - wide[(metric, "ltbi")]).abs()
        diffs.append({"metric": metric, "max_abs_difference": float(diff.max()), "mean_abs_difference": float(diff.mean())})
    return pd.DataFrame(diffs)


def run_prevalence_update_grid(output_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    raw_dir = output_dir / "replicate_level" / "prevalence_update"
    raw_dir.mkdir(parents=True, exist_ok=True)
    for scenario_id, ltbi_prevalence, prevalence_label in PREVALENCE_UPDATE_SCENARIOS:
        for strategy in MAIN_STRATEGIES:
            for screen_count in PREVALENCE_UPDATE_SCREEN_COUNTS:
                config = scenario_config(strategy, screen_count, ltbi_prevalence)
                model_out = run_scenario_with_do_nothing(config)
                raw = model_out["results"]["raw"].copy()
                raw.insert(0, "prevalence_scenario", scenario_id)
                raw.insert(1, "screen_count_requested", screen_count)
                raw.to_csv(raw_dir / f"{scenario_id}__{strategy}__n_screened_{screen_count}.csv", index=False)
                row = build_curve_row(scenario_id, strategy, screen_count, config, model_out)
                rows.append(
                    {
                        "prevalence_scenario": scenario_id,
                        "ltbi_prevalence": ltbi_prevalence,
                        "prevalence_label": prevalence_label,
                        "screeningStrategy": strategy,
                        "strategy_label": STRATEGY_LABELS[strategy],
                        "screen_count_requested": screen_count,
                        "screenCoverage": screen_count / N,
                        "expected_test_positive_nonactive": row["nTestPositiveNonActive_median"],
                        "expected_true_positive_tests": row["nTruePositiveTests_median"],
                        "expected_false_positive_tests": row["nFalsePositiveTests_median"],
                        "expected_PPV": row["PPV_trueAmongTestPositive"],
                        "expected_false_positive_share": row["falsePositiveShareAmongTestPositive"],
                        "expected_active_TB_prevented": row["nPreventedActiveTB_median"],
                        "expected_NNS_prevent_TB": row["NNS_prevent_one_active_TB"],
                    }
                )
    return pd.DataFrame(rows)


def build_curve_row(prevalence_name: str, strategy: str, screen_count: int, config: dict[str, Any], model_out: dict[str, Any]) -> dict[str, Any]:
    dynamic = model_out["bundle"]["technical"]["dynamicComparison"]
    med = lambda name: summary_value(model_out, name)
    low = lambda name: summary_value(model_out, name, "Low95")
    high = lambda name: summary_value(model_out, name, "High95")
    n_test_positive = med("nTestPositiveNonActive")
    n_false_positive = med("nFalsePositiveTests")
    n_started = med("nTotalCoursesStarted")
    n_false_positive_treated = med("nFalsePositiveTreated")
    n_true_positive = n_test_positive - n_false_positive
    n_true_positive_treated = n_started - n_false_positive_treated
    n_prevented = med("nPreventedActiveTB")
    return {
        "prevalence_scenario": prevalence_name,
        "screeningStrategy": strategy,
        "strategy_label": STRATEGY_LABELS[strategy],
        "screen_count_requested": screen_count,
        "screenCoverage": config["screenCoverage"],
        "nScreened_median": med("nScreened"),
        "nScreened_low95": low("nScreened"),
        "nScreened_high95": high("nScreened"),
        "nTestPositiveNonActive_median": n_test_positive,
        "nFalsePositiveTests_median": n_false_positive,
        "nFalsePositiveTreated_median": n_false_positive_treated,
        "nTotalCoursesStarted_median": n_started,
        "nTotalCoursesCompleted_median": med("nTotalCoursesCompleted"),
        "nCuredInfection_median": med("nCuredInfection"),
        "nPreventedActiveTB_median": n_prevented,
        "nActiveBy20y_DoNothing_median": dynamic.get("cumulative_baseline_active_tb_cases"),
        "nActiveBy20y_AfterStrategy_median": dynamic.get("cumulative_intervention_active_tb_cases"),
        "relativeReduction20y_median": dynamic.get("relative_reduction_cumulative_active_tb_cases"),
        "nIGRApos_median": med("nIGRApos"),
        "nTestPositive_median": med("nTestPositive"),
        "nScreenedBCG_median": med("nScreenedBCG"),
        "nFalsePositiveTestsBCG_median": med("nFalsePositiveTestsBCG"),
        "nFalsePositiveTestsNoBCG_median": med("nFalsePositiveTestsNoBCG"),
        "nTruePositiveTests_median": n_true_positive,
        "nTruePositiveTreated_median": n_true_positive_treated,
        "falsePositiveShareAmongTestPositive": safe_div(n_false_positive, n_false_positive + n_true_positive),
        "PPV_trueAmongTestPositive": safe_div(n_true_positive, n_false_positive + n_true_positive),
        "falsePositiveTreatedShareAmongCoursesStarted": safe_div(n_false_positive_treated, n_started),
        "falsePositiveTestsPerTruePositiveTest": safe_div(n_false_positive, n_true_positive),
        "falsePositiveTreatedPerTruePositiveTreated": safe_div(n_false_positive_treated, n_true_positive_treated),
        "NNS_prevent_one_active_TB": safe_div(med("nScreened"), n_prevented),
        "NNT_started_prevent_one_active_TB": safe_div(n_started, n_prevented),
        "NNT_completed_prevent_one_active_TB": safe_div(med("nTotalCoursesCompleted"), n_prevented),
        "notes": "Low LTBI scenario uses ltbiPrevalence=0.01 sensitivity." if prevalence_name == "low_LTBI_1pct" else "Default APY LTBI calibration.",
    }


def summary_value(model_out: dict[str, Any], metric: str, column: str = "Median") -> float:
    summary = model_out["results"]["summary"]
    matched = summary.loc[summary["Metric"] == metric, column]
    return math.nan if matched.empty else float(matched.iloc[0])


def safe_div(numerator: float, denominator: float) -> float:
    if denominator is None or pd.isna(denominator) or float(denominator) == 0:
        return math.nan
    return float(numerator) / float(denominator)


def build_targeting_order_profiles() -> pd.DataFrame:
    config = scenario_config("prevent", 0, None)
    calibration = calibrate_from_config(config)
    pars = calibration["parameters"]
    rows: list[dict[str, Any]] = []
    for age in pars["exactAgeValues"]:
        for mj in [False, True]:
            for contact in [False, True]:
                for renal in [False, True]:
                    for diabetes in [False, True]:
                        for smoking in [False, True]:
                            for cld in [False, True]:
                                for alcohol in [False, True]:
                                    p_ltbi = float(infection_probability(age, pars, calibration["ageInfLogLambda"], calibration["ageInfGamma"], mj, contact, renal))
                                    disease_mult = disease_multiplier_from_flags(pars, mj, contact, renal, diabetes, smoking, cld, alcohol)
                                    prevention_score = p_ltbi * float(preventable_active_risk(disease_mult, calibration["lambdaEarly"], calibration["lambdaLate"], SCREEN_WINDOW, FOLLOW_HORIZON))
                                    rows.append(
                                        {
                                            "age": age,
                                            "age_band": age_band_from_age(age),
                                            "marijuana_use": mj,
                                            "close_contact": contact,
                                            "renal_disease": renal,
                                            "diabetes": diabetes,
                                            "smoking": smoking,
                                            "chronic_lung_disease": cld,
                                            "alcohol_drugs": alcohol,
                                            "pLTBI": p_ltbi,
                                            "LTBI risk score": p_ltbi,
                                            "disease multiplier": disease_mult,
                                            "prevention score": prevention_score,
                                        }
                                    )
    profiles = pd.DataFrame(rows)
    profiles["LTBI targeting priority"] = profiles["LTBI risk score"].rank(method="first", ascending=False)
    profiles["prevention targeting priority"] = profiles["prevention score"].rank(method="first", ascending=False)
    ltbi_top = profiles.nsmallest(100, "LTBI targeting priority").copy()
    ltbi_top["targeting_list"] = "top_100_LTBI_targeted"
    prevent_top = profiles.nsmallest(100, "prevention targeting priority").copy()
    prevent_top["targeting_list"] = "top_100_prevention_targeted"
    out = pd.concat([ltbi_top, prevent_top], ignore_index=True)
    out.insert(0, "rank", out.groupby("targeting_list").cumcount() + 1)
    out["notes"] = "Model-derived risk profile ordering. This is not a list of real people."
    return out


def build_example_cohort_ordering() -> pd.DataFrame:
    config = scenario_config("prevent", N, None)
    cohort = simulate_one_cohort_from_config(config, n=N, seed=SEED)["cohort"].copy()
    cohort = cohort.rename(columns={"id": "example_id", "pInfection": "pLTBI"})
    cohort["age band"] = cohort["ageYears"].map(age_band_from_age)
    cohort["LTBI rank"] = cohort["ltbiRiskScore"].rank(method="first", ascending=False)
    cohort["prevention rank"] = cohort["preventTargetScore"].rank(method="first", ascending=False)
    return cohort[
        [
            "example_id",
            "ageYears",
            "age band",
            "MJ",
            "contact",
            "renal",
            "diabetes",
            "smoking",
            "chronicLungDisease",
            "alcoholDrugs",
            "pLTBI",
            "ltbiRiskScore",
            "diseaseMultiplier",
            "preventTargetScore",
            "LTBI rank",
            "prevention rank",
        ]
    ].sort_values("prevention rank")


def build_rank_decile_summary(example: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for rank_col in ["LTBI rank", "prevention rank"]:
        ordered = example.sort_values(rank_col).copy()
        ordered["decile"] = pd.qcut(ordered[rank_col], 10, labels=False) + 1
        for decile, sub in ordered.groupby("decile"):
            rows.append(
                {
                    "ranking": rank_col,
                    "decile": int(decile),
                    "n": int(len(sub)),
                    "mean_pLTBI": float(sub["pLTBI"].mean()),
                    "mean_ltbiRiskScore": float(sub["ltbiRiskScore"].mean()),
                    "mean_diseaseMultiplier": float(sub["diseaseMultiplier"].mean()),
                    "mean_preventTargetScore": float(sub["preventTargetScore"].mean()),
                    "renal_prevalence": float(sub["renal"].mean()),
                    "contact_prevalence": float(sub["contact"].mean()),
                    "marijuana_prevalence": float(sub["MJ"].mean()),
                    "diabetes_prevalence": float(sub["diabetes"].mean()),
                    "smoking_prevalence": float(sub["smoking"].mean()),
                    "cld_prevalence": float(sub["chronicLungDisease"].mean()),
                    "alcohol_drugs_prevalence": float(sub["alcoholDrugs"].mean()),
                    "min_rank": float(sub[rank_col].min()),
                    "max_rank": float(sub[rank_col].max()),
                }
            )
    return pd.DataFrame(rows)


def build_stopping_rule_candidates(curve_df: pd.DataFrame, raw_by_scenario: dict[tuple[str, str, int], pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for prevalence_name in PREVALENCE_SCENARIOS:
        for strategy in MAIN_STRATEGIES:
            for checkpoint in CHECKPOINTS:
                current = select_curve(curve_df, prevalence_name, strategy, checkpoint)
                to_450 = select_curve(curve_df, prevalence_name, strategy, 450)
                to_1500 = select_curve(curve_df, prevalence_name, strategy, 1500)
                rows.append(
                    {
                        "row_type": "checkpoint_summary",
                        "prevalence_scenario": prevalence_name,
                        "strategy": strategy,
                        "checkpoint_screened": checkpoint,
                        "min_test_positive_nonactive_to_continue": None,
                        "median_test_positive_nonactive": current["nTestPositiveNonActive_median"],
                        "median_true_positive_tests": current["nTruePositiveTests_median"],
                        "median_false_positive_tests": current["nFalsePositiveTests_median"],
                        "PPV": current["PPV_trueAmongTestPositive"],
                        "false_positive_share": current["falsePositiveShareAmongTestPositive"],
                        "projected_TB_cases_prevented_if_continue_to_450": to_450["nPreventedActiveTB_median"],
                        "projected_TB_cases_prevented_if_continue_to_1500": to_1500["nPreventedActiveTB_median"],
                        "NNS_prevent_TB_at_checkpoint": current["NNS_prevent_one_active_TB"],
                        "NNS_prevent_TB_if_continue_to_450": to_450["NNS_prevent_one_active_TB"],
                        "NNS_prevent_TB_if_continue_to_1500": to_1500["NNS_prevent_one_active_TB"],
                        "probability_yield_below_threshold": None,
                        "median_yield_above_threshold": None,
                        "expected_consequences_of_stopping": "Stopping here would forgo the projected additional direct person-level TB prevention shown in the 450 and 1500 columns.",
                        "notes": "Stopping rules should be review triggers, not automatic stop decisions.",
                    }
                )
            raw = raw_by_scenario[(prevalence_name, strategy, 100)]
            current_100 = select_curve(curve_df, prevalence_name, strategy, 100)
            to_450 = select_curve(curve_df, prevalence_name, strategy, 450)
            to_1500 = select_curve(curve_df, prevalence_name, strategy, 1500)
            for threshold in STOPPING_THRESHOLDS:
                yield_values = raw["nTestPositiveNonActive"].astype(float)
                above = yield_values[yield_values >= threshold]
                rows.append(
                    {
                        "row_type": "threshold_at_100_screened",
                        "prevalence_scenario": prevalence_name,
                        "strategy": strategy,
                        "checkpoint_screened": 100,
                        "min_test_positive_nonactive_to_continue": threshold,
                        "median_test_positive_nonactive": current_100["nTestPositiveNonActive_median"],
                        "median_true_positive_tests": current_100["nTruePositiveTests_median"],
                        "median_false_positive_tests": current_100["nFalsePositiveTests_median"],
                        "PPV": current_100["PPV_trueAmongTestPositive"],
                        "false_positive_share": current_100["falsePositiveShareAmongTestPositive"],
                        "projected_TB_cases_prevented_if_continue_to_450": to_450["nPreventedActiveTB_median"],
                        "projected_TB_cases_prevented_if_continue_to_1500": to_1500["nPreventedActiveTB_median"],
                        "NNS_prevent_TB_at_checkpoint": current_100["NNS_prevent_one_active_TB"],
                        "NNS_prevent_TB_if_continue_to_450": to_450["NNS_prevent_one_active_TB"],
                        "NNS_prevent_TB_if_continue_to_1500": to_1500["NNS_prevent_one_active_TB"],
                        "probability_yield_below_threshold": float((yield_values < threshold).mean()),
                        "median_yield_above_threshold": math.nan if above.empty else float(above.median()),
                        "expected_consequences_of_stopping": "Use as a review trigger only; projected benefits are from independently simulated continuation scenarios.",
                        "notes": "Probability uses replicate-level model output for the 100-screened scenario; prospective validation still requires operational data.",
                    }
                )
    return pd.DataFrame(rows)


def select_curve(curve_df: pd.DataFrame, prevalence_name: str, strategy: str, screen_count: int) -> dict[str, Any]:
    row = curve_df[
        (curve_df["prevalence_scenario"] == prevalence_name)
        & (curve_df["screeningStrategy"] == strategy)
        & (curve_df["screen_count_requested"] == screen_count)
    ]
    if row.empty:
        raise ValueError((prevalence_name, strategy, screen_count))
    return row.iloc[0].to_dict()


def to_tidy_long(curve_df: pd.DataFrame) -> pd.DataFrame:
    value_cols = [c for c in curve_df.columns if c not in {"prevalence_scenario", "screeningStrategy", "strategy_label", "screen_count_requested", "screenCoverage", "notes"}]
    return curve_df.melt(
        id_vars=["prevalence_scenario", "screeningStrategy", "strategy_label", "screen_count_requested", "screenCoverage"],
        value_vars=value_cols,
        var_name="outcome",
        value_name="value",
    )


def write_chart_tables(output_dir: Path, curve_df: pd.DataFrame) -> None:
    rows = []
    for prevalence_name in PREVALENCE_SCENARIOS:
        for label, metric in CHART_BLOCKS:
            for screen_count in SCREEN_COUNT_GRID:
                for strategy in MAIN_STRATEGIES:
                    row = select_curve(curve_df, prevalence_name, strategy, screen_count)
                    rows.append(
                        {
                            "prevalence_scenario": prevalence_name,
                            "chart_block": label,
                            "metric": metric,
                            "screen_count_requested": screen_count,
                            "screeningStrategy": strategy,
                            "strategy_label": STRATEGY_LABELS[strategy],
                            "value": row[metric],
                        }
                    )
    pd.DataFrame(rows).to_csv(output_dir / "locked_chart_data_long.csv", index=False)


def write_locked_workbook(path: Path, curve_df: pd.DataFrame, cure_audit_df: pd.DataFrame, targeting_profiles: pd.DataFrame, example_cohort: pd.DataFrame, decile_summary: pd.DataFrame, stopping_df: pd.DataFrame, prevalence_grid_df: pd.DataFrame, expected_early_yield_df: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_readme_sheet(wb)
    write_run_settings_sheet(wb, cure_audit_df)
    write_df(wb, "Curve_Data_All", curve_df)
    write_chart_data_sheet(wb, "Chart_Data_Base_LTBI", curve_df, "base_default_LTBI")
    write_chart_data_sheet(wb, "Chart_Data_Low_LTBI_1pct", curve_df, "low_LTBI_1pct")
    write_df(wb, "Targeting_Order_Profile", targeting_profiles)
    write_df(wb, "Targeting_Order_Example_Cohort", example_cohort)
    write_df(wb, "Rank_Decile_Summary", decile_summary)
    write_df(wb, "Stopping_Rule_Candidates", stopping_df)
    write_df(wb, "Prevalence_Scenario_Grid", prevalence_grid_df)
    write_df(wb, "Expected_Early_Yield", expected_early_yield_df)
    write_definitions_sheet(wb)
    wb.save(path)


def write_readme_sheet(wb: Workbook) -> None:
    rows = [
        ("Purpose", "Frozen APY epidemiology report v1 chart-data workbook."),
        ("Scope", "Epidemiological outputs only; no costs or health-economic formulas."),
        ("Backend", "Python APY backend."),
        ("Planning axis", "Actual number screened, from 0 to 1500."),
        ("Use guardrail", "Supports planning and sequencing, not denying screening or treatment."),
    ]
    write_key_value_sheet(wb, "README", rows)


def write_run_settings_sheet(wb: Workbook, cure_audit_df: pd.DataFrame) -> None:
    rows = [
        ("N", N),
        ("nReps", N_REPS),
        ("seed", SEED),
        ("screenWindow", SCREEN_WINDOW),
        ("followHorizon", FOLLOW_HORIZON),
        ("testType", TEST_TYPE),
        ("regimen", REGIMEN),
        ("screen_count_grid", ", ".join(str(x) for x in SCREEN_COUNT_GRID)),
        ("prevalence_scenarios", ", ".join(PREVALENCE_SCENARIOS.keys())),
    ]
    ws = write_key_value_sheet(wb, "Run_Settings", rows)
    start = len(rows) + 4
    ws.cell(start, 1, "Cure vs LTBI audit").font = Font(bold=True)
    write_df_to_sheet(ws, cure_audit_df, start + 1)


def write_chart_data_sheet(wb: Workbook, title: str, curve_df: pd.DataFrame, prevalence_name: str) -> None:
    ws = wb.create_sheet(title)
    current_row = 1
    for label, metric in CHART_BLOCKS:
        ws.cell(current_row, 1, label).font = Font(bold=True)
        headers = ["number_screened"] + [STRATEGY_LABELS[s] for s in MAIN_STRATEGIES]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(current_row + 1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for offset, screen_count in enumerate(SCREEN_COUNT_GRID, start=1):
            ws.cell(current_row + 1 + offset, 1, screen_count)
            for col, strategy in enumerate(MAIN_STRATEGIES, start=2):
                row = select_curve(curve_df, prevalence_name, strategy, screen_count)
                ws.cell(current_row + 1 + offset, col, clean_cell(row[metric]))
        current_row += len(SCREEN_COUNT_GRID) + 4
    set_widths(ws)


def write_definitions_sheet(wb: Workbook) -> None:
    rows = [
        ("number screened", "Actual count selected for screening in the model scenario."),
        ("test positive non-active", "Positive test among people not active TB at screening; includes true and false positives."),
        ("false-positive test", "Uninfected simulated person with a positive test."),
        ("true-positive test", "nTestPositiveNonActive minus nFalsePositiveTests."),
        ("PPV", "True-positive tests divided by all non-active test positives."),
        ("NNS to prevent one active TB case", "nScreened divided by nPreventedActiveTB."),
        ("direct person-level active TB cases prevented", "Cases prevented among screened/treated people only."),
        ("downstream transmission", "Not included in these APY outputs."),
    ]
    write_key_value_sheet(wb, "Definitions", rows, headers=("Term", "Definition"))


def write_df(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title)
    write_df_to_sheet(ws, df, 1)


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int) -> None:
    for col, header in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, clean_cell(value))
    ws.freeze_panes = f"A{start_row + 1}"
    set_widths(ws)


def write_key_value_sheet(wb: Workbook, title: str, rows: list[tuple[Any, Any]], headers: tuple[str, str] = ("Item", "Value")):
    ws = wb.create_sheet(title)
    ws.cell(1, 1, headers[0]).font = Font(bold=True)
    ws.cell(1, 2, headers[1]).font = Font(bold=True)
    for row_idx, (key, value) in enumerate(rows, start=2):
        ws.cell(row_idx, 1, key)
        ws.cell(row_idx, 2, value)
    set_widths(ws)
    return ws


def write_existing_report_comparison(output_dir: Path, curve_df: pd.DataFrame, prevalence_grid_df: pd.DataFrame) -> None:
    existing_dir = REPO_ROOT / "outputs" / "apy_cho_epi_outputs"
    rows = []
    old_curve_path = existing_dir / "apy_cho_epi_curve_data.csv"
    if old_curve_path.exists():
        old_curve = pd.read_csv(old_curve_path)
        rows.extend(compare_frame("apy_cho_epi_curve_data.csv", old_curve, curve_df, ["prevalence_scenario", "screeningStrategy", "screen_count_requested"]))
    old_prev_path = existing_dir / "apy_prevalence_scenario_grid.csv"
    if old_prev_path.exists():
        old_prev = pd.read_csv(old_prev_path)
        rows.extend(compare_frame("apy_prevalence_scenario_grid.csv", old_prev, prevalence_grid_df, ["prevalence_scenario", "screeningStrategy", "screen_count_requested"]))
    comparison = pd.DataFrame(rows)
    comparison.to_csv(output_dir / "comparison_to_existing_report.csv", index=False)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    write_df_to_sheet(ws, comparison, 1)
    wb.save(output_dir / "comparison_to_existing_report.xlsx")


def compare_frame(source: str, old_df: pd.DataFrame, new_df: pd.DataFrame, key_cols: list[str]) -> list[dict[str, Any]]:
    rows = []
    numeric_cols = [
        c for c in old_df.columns
        if c in new_df.columns and c not in key_cols and pd.api.types.is_numeric_dtype(old_df[c])
    ]
    old_long = old_df[key_cols + numeric_cols].melt(id_vars=key_cols, var_name="outcome", value_name="old_value")
    new_long = new_df[key_cols + numeric_cols].melt(id_vars=key_cols, var_name="outcome", value_name="locked_run_value")
    merged = old_long.merge(new_long, on=key_cols + ["outcome"], how="outer")
    for row in merged.itertuples(index=False):
        old_value = getattr(row, "old_value")
        new_value = getattr(row, "locked_run_value")
        old_missing = pd.isna(old_value)
        new_missing = pd.isna(new_value)
        difference_class = "exact"
        if old_missing and new_missing:
            diff = rel = 0.0
            passed = True
        elif old_missing or new_missing:
            diff = rel = math.nan
            passed = False
            difference_class = "missing_value"
        else:
            diff = float(new_value) - float(old_value)
            rel = math.nan if float(old_value) == 0 else diff / float(old_value)
            passed = abs(diff) <= HISTORICAL_COMPARISON_ABS_TOLERANCE
            if diff != 0.0 and passed:
                difference_class = "formatting_or_float_precision_only"
            elif not passed:
                difference_class = "numerical_difference"
        row_dict = row._asdict()
        row_dict["source_file"] = source
        row_dict["absolute_difference"] = diff
        row_dict["relative_difference"] = rel
        row_dict["comparison_abs_tolerance"] = HISTORICAL_COMPARISON_ABS_TOLERANCE
        row_dict["difference_class"] = difference_class
        row_dict["pass_fail"] = "pass" if passed else "fail"
        rows.append(row_dict)
    return rows


def build_input_manifest() -> dict[str, Any]:
    cfg = scenario_config("prevent", 450, None)
    pars = load_parameters_from_config(cfg)
    cal_base = calibrate_from_config(cfg)
    cal_low = calibrate_from_config(scenario_config("prevent", 450, 0.01))
    input_files = [
        REPO_ROOT / "abm" / "default_data.csv",
        REPO_ROOT / "abm" / "default_age_distribution.csv",
        RELEASE_DIR / "scenarios" / "base_default_ltbi.json",
        RELEASE_DIR / "scenarios" / "low_ltbi_1pct.json",
        RELEASE_DIR / "scenarios" / "prevalence_ltbi_0_5pct.json",
        RELEASE_DIR / "scenarios" / "prevalence_ltbi_2pct.json",
        RELEASE_DIR / "scenarios" / "prevalence_ltbi_3pct.json",
        RELEASE_DIR / "scenarios" / "prevalence_ltbi_5pct.json",
        RELEASE_DIR / "scenarios" / "prevalence_ltbi_7_53pct.json",
        RELEASE_DIR / "scenarios" / "prevalence_ltbi_10pct.json",
    ]
    return {
        "input_files": [{"path": rel(path), "sha256": sha256_file(path)} for path in input_files],
        "resolved_parameters": {
            "N": N,
            "nReps": N_REPS,
            "seed": SEED,
            "screenWindow": SCREEN_WINDOW,
            "followHorizon": FOLLOW_HORIZON,
            "testType": TEST_TYPE,
            "testSensitivity": 0.95,
            "testSpecificity": 0.98,
            "regimen": REGIMEN,
            "pStartTPT": 0.85,
            "regimenPComplete": 0.80,
            "regimenADRstop": 0.05,
            "regimenEffFull": 0.85,
            "partialShortCourseMode": "threshold80",
            "partialDoseFractionADR": 0.30,
            "partialDoseFractionOther": 0.60,
            "targetAgeOR": 7.54,
            "earlyLateRatio": 5,
            "ltbiPrevalence_base": DEFAULT_LTBI_PREVALENCE,
            "activeTBPrevalence_base_2y": DEFAULT_ACTIVE_TB_2Y_PREVALENCE,
            "ltbiPrevalence_low": 0.01,
            "activeTBPrevalence_low_2y": DEFAULT_ACTIVE_TB_2Y_PREVALENCE * 0.01 / DEFAULT_LTBI_PREVALENCE,
            "risk_factor_prevalence_by_age": {
                "MJ": pars["mjPrevByAge"],
                "contact": pars["contactPrevByAge"],
                "renal": pars["renalPrevByAge"],
                "diabetes": pars["diabetesPrevByAge"],
                "smoking": pars["smokingPrevByAge"],
                "cld": pars["cldPrevByAge"],
                "alcohol": pars["alcoholPrevByAge"],
            },
            "infection_OR": pars["infOR"],
            "disease_OR": pars["disOR"],
            "totalFemalePrev": pars["totalFemalePrev"],
            "totalBCGPrev": pars["totalBCGPrev"],
            "base_calibration": calibration_manifest(cal_base),
            "low_1pct_calibration": calibration_manifest(cal_low),
            "strategies": MAIN_STRATEGIES,
            "screen_count_grid": SCREEN_COUNT_GRID,
            "prevalence_update_screen_count_grid": PREVALENCE_UPDATE_SCREEN_COUNTS,
            "prevalence_update_scenarios": [
                {"scenario_id": sid, "ltbiPrevalence": ltbi, "label": label}
                for sid, ltbi, label in PREVALENCE_UPDATE_SCENARIOS
            ],
        },
    }


def build_environment_record() -> dict[str, Any]:
    packages = {}
    for package in ["numpy", "pandas", "scipy", "openpyxl", "matplotlib", "streamlit", "altair"]:
        try:
            packages[package] = metadata.version(package)
        except metadata.PackageNotFoundError:
            packages[package] = "not installed"
    matlab_version = shutil.which("matlab")
    return {
        "recorded_at_utc": now_iso(),
        "operating_system": platform.platform(),
        "python_version": sys.version,
        "python_executable": sys.executable,
        "matlab_executable": matlab_version,
        "matlab_version": matlab_version_string() if matlab_version else "not used; executable not found on PATH",
        "packages": packages,
        "rng": {
            "implementation": "numpy.random.default_rng",
            "bit_generator": "PCG64",
            "parent_seed": SEED,
            "replicate_seed_draw": "default_rng(seed).integers(0, np.iinfo(np.uint32).max, size=nReps)",
        },
    }


def build_run_manifest(output_dir: Path, input_manifest: dict[str, Any]) -> dict[str, Any]:
    output_files = [
        path for path in sorted(output_dir.rglob("*"))
        if path.is_file() and path.name != "run_manifest.json"
    ]
    parent_rng = np.random.default_rng(SEED)
    seed_schedule = parent_rng.integers(0, np.iinfo(np.uint32).max, size=N_REPS)
    return {
        "created_at_utc": now_iso(),
        "git_commit": git_value(["git", "rev-parse", "HEAD"]),
        "git_branch": git_value(["git", "branch", "--show-current"]),
        "git_status_short": git_value(["git", "status", "--short"]),
        "backend": "python",
        "model_version": MODEL_VERSION,
        "scenario_files": [
            {"path": item["path"], "sha256": item["sha256"]}
            for item in input_manifest["input_files"]
            if item["path"].startswith("paper/sa_health_report/releases/apy_epi_report_v1/scenarios/")
        ],
        "input_file_checksums": input_manifest["input_files"],
        "cohort_size": N,
        "replicate_count": N_REPS,
        "seed": SEED,
        "seed_schedule_sha256": sha256_bytes(",".join(map(str, seed_schedule.tolist())).encode("ascii")),
        "seed_schedule_first_10": seed_schedule[:10].astype(int).tolist(),
        "rng": "numpy.random.default_rng using PCG64",
        "summary_statistic_definitions": {
            "Median": "pandas Series.median across replicate-level raw output",
            "Low95": "empirical quantile p=0.025 with pos = 1 + (n - 1) * p",
            "High95": "empirical quantile p=0.975 with pos = 1 + (n - 1) * p",
            "PPV_and_false_positive_ratios": "calculated from median component values in curve rows",
            "relativeReduction20y": "median of replicate-level prevented active TB divided by baseline active TB by 20 years",
        },
        "output_files": [{"path": str(path.relative_to(output_dir)), "sha256": sha256_file(path), "bytes": path.stat().st_size} for path in output_files],
    }


def write_frozen_scenario_expansions(output_dir: Path) -> None:
    scenario_dir = output_dir / "expanded_scenarios"
    scenario_dir.mkdir(parents=True, exist_ok=True)
    for prevalence_name, ltbi_prevalence in PREVALENCE_SCENARIOS.items():
        for strategy in MAIN_STRATEGIES:
            for screen_count in SCREEN_COUNT_GRID:
                write_json(scenario_dir / f"{prevalence_name}__{strategy}__n_screened_{screen_count}.json", scenario_config(strategy, screen_count, ltbi_prevalence))
    for scenario_id, ltbi_prevalence, _label in PREVALENCE_UPDATE_SCENARIOS:
        for strategy in MAIN_STRATEGIES:
            for screen_count in PREVALENCE_UPDATE_SCREEN_COUNTS:
                write_json(scenario_dir / f"{scenario_id}__{strategy}__n_screened_{screen_count}.json", scenario_config(strategy, screen_count, ltbi_prevalence))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=json_default) + "\n", encoding="utf-8")


def write_comparison_summary(release_dir: Path, run1: Path, run2: Path) -> None:
    rows = []
    for file1 in sorted(run1.rglob("*")):
        if not file1.is_file():
            continue
        rel_path = file1.relative_to(run1)
        if rel_path.parts and rel_path.parts[0] == "replicate_level":
            continue
        if file1.suffix.lower() not in {".csv", ".xlsx"}:
            continue
        file2 = run2 / rel_path
        if file2.exists():
            equal, max_abs = compare_numeric_file(file1, file2)
            rows.append({"file": str(rel_path), "numeric_equal": equal, "max_abs_difference": max_abs})
        else:
            rows.append({"file": str(rel_path), "numeric_equal": False, "max_abs_difference": math.nan})
    df = pd.DataFrame(rows)
    df.to_csv(release_dir / "reproducibility_verification.csv", index=False)


def compare_numeric_file(file1: Path, file2: Path) -> tuple[bool, float]:
    if file1.suffix.lower() == ".csv":
        left = pd.read_csv(file1)
        right = pd.read_csv(file2)
        return compare_numeric_frames(left, right)
    if file1.suffix.lower() == ".xlsx":
        max_abs = 0.0
        equal = True
        wb1 = load_workbook(file1, data_only=False, read_only=True)
        wb2 = load_workbook(file2, data_only=False, read_only=True)
        for sheet in sorted(set(wb1.sheetnames) & set(wb2.sheetnames)):
            rows1 = list(wb1[sheet].iter_rows(values_only=True))
            rows2 = list(wb2[sheet].iter_rows(values_only=True))
            if rows1 != rows2:
                equal = False
        wb1.close()
        wb2.close()
        return equal, max_abs
    return sha256_file(file1) == sha256_file(file2), 0.0


def compare_numeric_frames(left: pd.DataFrame, right: pd.DataFrame) -> tuple[bool, float]:
    if left.shape != right.shape or list(left.columns) != list(right.columns):
        return False, math.nan
    numeric_cols = [c for c in left.columns if pd.api.types.is_numeric_dtype(left[c]) or pd.api.types.is_numeric_dtype(right[c])]
    max_abs = 0.0
    for col in numeric_cols:
        l = pd.to_numeric(left[col], errors="coerce")
        r = pd.to_numeric(right[col], errors="coerce")
        diff = (l - r).abs()
        if not ((l.isna() & r.isna()) | (diff == 0)).all():
            max_abs = max(max_abs, float(diff.max(skipna=True)))
            return False, max_abs
    non_numeric_cols = [c for c in left.columns if c not in numeric_cols]
    if non_numeric_cols and not left[non_numeric_cols].equals(right[non_numeric_cols]):
        return False, max_abs
    return True, max_abs


def calibration_manifest(calibration: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in calibration.items()
        if key != "parameters"
    }


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value) or math.isinf(float(value)):
            return None
    except (TypeError, ValueError):
        return value
    return value


def set_widths(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 42)


def age_band_from_age(age: float) -> str:
    age = float(age)
    if age < 5:
        return "0-4"
    if age < 15:
        return "5-14"
    if age < 25:
        return "15-24"
    if age < 35:
        return "25-34"
    if age < 45:
        return "35-44"
    if age < 55:
        return "45-54"
    if age < 65:
        return "55-64"
    if age < 75:
        return "65-74"
    if age < 85:
        return "75-84"
    return "85+"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def git_value(args: list[str]) -> str:
    try:
        return subprocess.check_output(args, cwd=REPO_ROOT, text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return ""


def matlab_version_string() -> str:
    try:
        out = subprocess.check_output(["matlab", "-batch", "disp(version)"], text=True, stderr=subprocess.STDOUT, timeout=60)
        return out.strip()
    except Exception as exc:
        return f"not used; version unavailable: {exc}"


def rel(path: Path) -> str:
    return str(path.resolve().relative_to(REPO_ROOT)).replace("\\", "/")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def json_default(value: Any) -> Any:
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, Path):
        return str(value)
    if pd.isna(value):
        return None
    raise TypeError(f"Cannot serialize {type(value)!r}")


if __name__ == "__main__":
    raise SystemExit(main())
