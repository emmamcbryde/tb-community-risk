from __future__ import annotations

from copy import deepcopy
from io import BytesIO, StringIO
import csv
import hashlib
import json
from typing import Any

from openpyxl import Workbook

from engine.apy.costing import TARGET_CURRENCY, TARGET_PRICE_YEAR, normalise_cost_table
from engine.apy.economics import ensure_authoritative_cost_items, sync_legacy_cost_fields_from_cost_items
from engine.apy.evidence import (
    EXPECTED_COST_BASIS,
    REVIEWED_EXCLUSION_STATUS,
    REVIEWED_NUMERICAL_STATUSES,
    assess_apy_reference_readiness,
    load_apy_evidence_registry,
    validate_cost_item_evidence,
)


EDITABLE_ASSUMPTION_COLUMNS = [
    "assumptionId",
    "category",
    "description",
    "currentValue",
    "unit",
    "originalCurrency",
    "originalPriceYear",
    "targetCurrency",
    "targetPriceYear",
    "costBasis",
    "sourceCitation",
    "sourceLocation",
    "reviewStatus",
    "reviewStatusLabel",
    "provisional",
    "inclusionStatus",
    "inclusionStatusLabel",
    "bundledIntoAssumptionId",
    "doubleCountingGroup",
    "notes",
    "unresolvedReason",
    "conversionStatus",
    "inflationIndexId",
    "sourceYearIndexValue",
    "targetYearIndexValue",
    "inflationFactor",
    "convertedTargetYearCost",
    "conversionWarnings",
    "validationMessage",
]

ADVANCED_EVIDENCE_COLUMNS = [
    "assumptionId",
    "category",
    "sourceLocation",
    "targetCurrency",
    "targetPriceYear",
    "sourceYearIndexValue",
    "targetYearIndexValue",
    "bundledIntoAssumptionId",
    "doubleCountingGroup",
    "unresolvedReason",
]

DERIVED_CONVERSION_COLUMNS = [
    "conversionStatus",
    "inflationIndexId",
    "sourceYearIndexValue",
    "targetYearIndexValue",
    "inflationFactor",
    "convertedTargetYearCost",
    "conversionWarnings",
]

STANDARD_EDITOR_COLUMNS = [
    "description",
    "currentValue",
    "unit",
    "originalCurrency",
    "originalPriceYear",
    "costBasis",
    "sourceCitation",
    "reviewStatusLabel",
    "provisional",
    "inclusionStatusLabel",
    "notes",
    "validationMessage",
]

CONVERSION_AUDIT_COLUMNS = [
    "assumptionId",
    "description",
    "conversionStatus",
    "inflationIndexId",
    "inflationFactor",
    "convertedTargetYearCost",
    "conversionWarnings",
]

CURRENT_COST_ASSUMPTION_GROUPS = {
    "selected": "Current selected test and regimen",
    "program": "Programme and delivery costs",
    "active_tb": "Active-TB care",
}

TEST_COST_IDS = {
    "IGRA": "cost.test_igra",
    "TST": "cost.test_tst",
}

REGIMEN_COST_IDS = {
    "3HP": "cost.regimen_3hp",
    "4R": "cost.regimen_4r",
    "3HR": "cost.regimen_3hr",
    "6H": "cost.regimen_6h",
    "9H": "cost.regimen_9h",
}

DALY_ASSUMPTION_IDS = {
    "daly.active_tb_disability_weight",
    "daly.active_tb_duration",
    "daly.tb_case_fatality_risk",
    "daly.yll_per_tb_death",
    "daly.tpt_health_loss",
    "daly.adr_health_loss",
    "daly.post_tb_sequelae",
}
THRESHOLD_ASSUMPTION_ID = "threshold.gdp_per_capita"

STATUS_LABEL_TO_CODE = {
    "Unresolved": "unresolved",
    "Reviewed numerical assumption": "configured_reviewed",
    "Reviewed model-derived assumption": "model_derived_reviewed",
    "Reviewed exclusion": "reviewed_exclusion",
    "Unreviewed repository input": "unreviewed_repository_input",
    "Legacy placeholder": "legacy_placeholder",
    "Migrated legacy unverified": "migrated_legacy_unverified",
}
STATUS_CODE_TO_LABEL = {value: key for key, value in STATUS_LABEL_TO_CODE.items()}

INCLUSION_LABEL_TO_CODE = {
    "Included": "included",
    "Excluded": "excluded",
    "Bundled into another item": "bundled",
}
INCLUSION_CODE_TO_LABEL = {value: key for key, value in INCLUSION_LABEL_TO_CODE.items()}

DALY_ROW_TO_CONFIG = {
    "daly.active_tb_disability_weight": ("activeTBDisabilityWeight", "numeric"),
    "daly.active_tb_duration": ("activeTBDurationYears", "numeric"),
    "daly.tb_case_fatality_risk": ("tbCaseFatalityRisk", "numeric"),
    "daly.yll_per_tb_death": ("yllPerTBDeath", "numeric"),
    "daly.tpt_health_loss": ("dalyLossPerTPTStarted", "optional_tpt"),
    "daly.adr_health_loss": ("dalyLossPerADRStop", "optional_adr"),
    "daly.post_tb_sequelae": ("postTBSequelae", "post_tb"),
}


def editable_assumption_rows(
    registry: list[dict[str, Any]] | None = None,
    economics_config: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    if registry is None and isinstance(economics_config, dict):
        registry = economics_config.get("assumptionEvidenceRegistry")
    rows = deepcopy(registry) if registry is not None else load_apy_evidence_registry()
    cost_basis_by_id = _cost_basis_lookup(economics_config or {})
    conversion_by_id = _conversion_lookup(economics_config or {})
    out = []
    for row in rows:
        item = {column: _clean_value(row.get(column)) for column in EDITABLE_ASSUMPTION_COLUMNS}
        item["provisional"] = _bool(row.get("provisional"))
        item["reviewStatus"] = normalise_status_value(item.get("reviewStatus"))
        item["reviewStatusLabel"] = status_label(item["reviewStatus"])
        item["inclusionStatus"] = normalise_inclusion_value(item.get("inclusionStatus"))
        item["inclusionStatusLabel"] = inclusion_label(item["inclusionStatus"])
        item["costBasis"] = row.get("costBasis") or cost_basis_by_id.get(row.get("assumptionId"), "")
        item.update(conversion_by_id.get(row.get("assumptionId"), {}))
        item["validationMessage"] = row.get("validationMessage", "")
        out.append(item)
    return out


def status_label(value: Any) -> str:
    code = normalise_status_value(value)
    return STATUS_CODE_TO_LABEL.get(code, code or "Unresolved")


def inclusion_label(value: Any) -> str:
    code = normalise_inclusion_value(value)
    return INCLUSION_CODE_TO_LABEL.get(code, code or "Included")


def normalise_status_value(value: Any) -> str:
    text = str(value or "").strip()
    return STATUS_LABEL_TO_CODE.get(text, text or "unresolved")


def normalise_inclusion_value(value: Any) -> str:
    text = str(value or "").strip()
    return INCLUSION_LABEL_TO_CODE.get(text, text or "included")


def rows_from_display_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        item = dict(row)
        if item.get("reviewStatusLabel"):
            item["reviewStatus"] = normalise_status_value(item["reviewStatusLabel"])
        else:
            item["reviewStatus"] = normalise_status_value(item.get("reviewStatus"))
            item["reviewStatusLabel"] = status_label(item["reviewStatus"])
        if item.get("inclusionStatusLabel"):
            item["inclusionStatus"] = normalise_inclusion_value(item["inclusionStatusLabel"])
        else:
            item["inclusionStatus"] = normalise_inclusion_value(item.get("inclusionStatus"))
            item["inclusionStatusLabel"] = inclusion_label(item["inclusionStatus"])
        out.append(item)
    return out


def workspace_source_hash(economics_config: dict[str, Any] | None) -> str:
    source = deepcopy(economics_config or {})
    for transient in (
        "assumptionEvidenceRegistry",
        "assumptionEvidenceValidation",
    ):
        source.pop(transient, None)
    payload = json.dumps(source, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def rows_hash(rows: list[dict[str, Any]]) -> str:
    comparable = []
    ignored = {"validationMessage", *DERIVED_CONVERSION_COLUMNS}
    for row in rows_from_display_rows(rows):
        comparable.append({key: value for key, value in row.items() if key not in ignored})
    payload = json.dumps(comparable, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def new_workspace_state(
    economics_config: dict[str, Any] | None,
    registry: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if registry is None and isinstance(economics_config, dict):
        registry = economics_config.get("assumptionEvidenceRegistry")
    rows = editable_assumption_rows(registry, economics_config or {})
    return {
        "rows": rows,
        "sourceHash": workspace_source_hash(economics_config or {}),
        "baselineRowsHash": rows_hash(rows),
        "hasUnsavedEdits": False,
        "validated": False,
        "applied": False,
        "validation": None,
        "presetConflict": False,
        "pendingSourceHash": "",
    }


def reconcile_workspace_state(
    state: dict[str, Any] | None,
    economics_config: dict[str, Any] | None,
    *,
    action: str | None = None,
    registry: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    current_hash = workspace_source_hash(economics_config or {})
    if not state or not isinstance(state.get("rows"), list):
        return new_workspace_state(economics_config, registry)
    out = deepcopy(state)
    out["hasUnsavedEdits"] = rows_hash(out.get("rows") or []) != out.get("baselineRowsHash")

    if action in {"reset", "discard"}:
        return new_workspace_state(economics_config, registry)
    if action == "keep":
        out["sourceHash"] = current_hash
        out["pendingSourceHash"] = ""
        out["presetConflict"] = False
        out["validated"] = False
        out["validation"] = None
        return out

    if out.get("sourceHash") != current_hash:
        if out.get("hasUnsavedEdits"):
            out["presetConflict"] = True
            out["pendingSourceHash"] = current_hash
            out["validated"] = False
            out["validation"] = None
        else:
            return new_workspace_state(economics_config, registry)
    return out


def update_workspace_rows(state: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
    incoming_rows = rows_from_display_rows(rows)
    incoming_hash = rows_hash(incoming_rows)
    current_hash = rows_hash(state.get("rows") or [])
    out = deepcopy(state)
    out["rows"] = incoming_rows
    out["hasUnsavedEdits"] = incoming_hash != out.get("baselineRowsHash")
    if incoming_hash != current_hash:
        out["validated"] = False
        out["validation"] = None
        out["applied"] = False
    return out


def mark_workspace_validated(state: dict[str, Any], validation: dict[str, Any]) -> dict[str, Any]:
    out = deepcopy(state)
    out["rows"] = validation["rows"]
    out["validation"] = validation
    out["validated"] = True
    return out


def mark_workspace_applied(state: dict[str, Any], economics_config: dict[str, Any]) -> dict[str, Any]:
    out = deepcopy(state)
    out["sourceHash"] = workspace_source_hash(economics_config)
    out["baselineRowsHash"] = rows_hash(out.get("rows") or [])
    out["hasUnsavedEdits"] = False
    out["applied"] = True
    out["presetConflict"] = False
    out["pendingSourceHash"] = ""
    return out


def validate_editable_assumptions(
    rows: list[dict[str, Any]],
    economics_config: dict[str, Any] | None = None,
    *,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    working_rows = [_normalise_edit_row(row) for row in rows]
    econ = _economics_config_with_rows(economics_config or {}, working_rows)
    readiness = assess_apy_reference_readiness(config or {}, econ, working_rows)
    row_messages: dict[str, list[str]] = {}
    metadata = econ.get("metadata") or {}
    analysis_target_currency = str(metadata.get("targetCurrency") or TARGET_CURRENCY)
    analysis_target_year = str(metadata.get("targetPriceYear") or TARGET_PRICE_YEAR)

    cost_items = normalise_cost_table(econ.get("costItems") or [])
    item_by_assumption = {f"cost.{item.get('costItemId')}": item for item in cost_items}
    for row in working_rows:
        assumption_id = row.get("assumptionId", "")
        messages: list[str] = []
        category = row.get("category")
        if category == "cost":
            value = _number_or_none(row.get("currentValue"))
            if value is not None and value < 0:
                messages.append("Cost value must be non-negative.")
            if row.get("inclusionStatus") == "bundled":
                messages.extend(_validate_bundled_row(row, working_rows))
            elif row.get("inclusionStatus") == "excluded":
                messages.extend(_validate_exclusion_row(row))
            else:
                messages.extend(
                    validate_cost_item_evidence(
                        item_by_assumption.get(assumption_id, {}),
                        row,
                        target_currency=analysis_target_currency,
                        target_price_year=analysis_target_year,
                    )
                )
        elif category == "daly":
            messages.extend(_validate_daly_row(row))
        elif category == "threshold":
            messages.extend(_validate_threshold_row(row, econ))
        elif category == "epidemiology":
            if not _row_reviewed(row):
                messages.append("Epidemiology assumption is not reviewed and remains a blocker.")
        if row.get("provisional"):
            messages.append("Assumption remains provisional.")
        if messages:
            row_messages[assumption_id] = messages
        row["validationMessage"] = "; ".join(messages)

    cost_row_messages = {
        assumption_id: messages
        for assumption_id, messages in row_messages.items()
        if assumption_id.startswith("cost.")
    }
    daly_row_messages = {
        assumption_id: messages
        for assumption_id, messages in row_messages.items()
        if assumption_id.startswith("daly.")
    }
    threshold_row_messages = {
        assumption_id: messages
        for assumption_id, messages in row_messages.items()
        if assumption_id.startswith("threshold.")
    }
    cost_ready = bool(readiness["costReady"] and not cost_row_messages)
    daly_category_ready = bool(readiness["dalyReady"] and not daly_row_messages)
    threshold_ready = bool(readiness["thresholdReady"] and not threshold_row_messages)
    cost_consequence_ready = bool(readiness["epidemiologyReady"] and cost_ready)
    daly_ready = bool(cost_consequence_ready and daly_category_ready)
    icer_ready = bool(daly_ready)
    nmb_ready = bool(icer_ready and threshold_ready)
    blockers = sorted(set(readiness["unresolvedAssumptionIds"] + list(row_messages)))
    return {
        "isValidForApplication": not any(_fatal_application_messages(messages) for messages in row_messages.values()),
        "rows": working_rows,
        "rowMessages": row_messages,
        "readiness": readiness,
        "summary": {
            "costConsequenceReady": cost_consequence_ready,
            "dalyReady": daly_ready,
            "icerReady": icer_ready,
            "nmbReady": nmb_ready,
            "epidemiologyReady": readiness["epidemiologyReady"],
            "costReady": cost_ready,
            "dalyCategoryReady": daly_category_ready,
            "thresholdReady": threshold_ready,
            "overallClinicianReady": readiness["overallClinicianReady"],
            "remainingBlockers": blockers,
        },
    }


def apply_assumptions_to_economics_config(
    economics_config: dict[str, Any],
    rows: list[dict[str, Any]],
    *,
    validate_first: bool = True,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    working_rows = [_normalise_edit_row(row) for row in rows]
    report = None
    if validate_first:
        report = validate_editable_assumptions(working_rows, economics_config, config=config or {})
        if not report["isValidForApplication"]:
            raise ValueError("Edited assumptions contain validation errors and were not applied.")
        working_rows = report["rows"]

    econ = _economics_config_with_rows(economics_config or {}, working_rows)
    econ["assumptionEvidenceRegistry"] = deepcopy(working_rows)
    if report is None:
        report = validate_editable_assumptions(working_rows, economics_config, config=config or {})
    econ["assumptionEvidenceValidation"] = report["summary"]
    return sync_legacy_cost_fields_from_cost_items(econ)


def _economics_config_with_rows(economics_config: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
    econ = deepcopy(economics_config or {})
    econ["costItems"] = _apply_cost_rows(econ, rows)
    _apply_threshold_rows(econ, rows)
    _apply_daly_rows(econ, rows)
    return econ


def assumptions_csv(rows: list[dict[str, Any]]) -> str:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=EDITABLE_ASSUMPTION_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _serialise_cell(row.get(column)) for column in EDITABLE_ASSUMPTION_COLUMNS})
    return output.getvalue()


def assumptions_workbook(rows: list[dict[str, Any]], validation: dict[str, Any] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Edited_assumptions"
    ws.append(EDITABLE_ASSUMPTION_COLUMNS)
    for row in rows:
        ws.append([_serialise_cell(row.get(column)) for column in EDITABLE_ASSUMPTION_COLUMNS])
    status = wb.create_sheet("Validation_status")
    summary = (validation or {}).get("summary", {})
    status.append(["Field", "Value"])
    for key, value in summary.items():
        status.append([key, _serialise_cell(value)])
    blockers = wb.create_sheet("Blockers")
    blockers.append(["assumptionId", "message"])
    for assumption_id, messages in (validation or {}).get("rowMessages", {}).items():
        blockers.append([assumption_id, "; ".join(messages)])
    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()


def parse_assumptions_csv(payload: bytes | str) -> list[dict[str, Any]]:
    text = payload.decode("utf-8-sig") if isinstance(payload, bytes) else payload
    reader = csv.DictReader(StringIO(text))
    rows = []
    for row in reader:
        rows.append(_normalise_edit_row(row))
    return rows


def group_rows(rows: list[dict[str, Any]], group: str) -> list[dict[str, Any]]:
    if group == "Costs":
        return [row for row in rows if row.get("category") == "cost"]
    if group == "DALYs":
        return [row for row in rows if row.get("category") == "daly"]
    if group == "Threshold":
        return [row for row in rows if row.get("category") == "threshold"]
    if group == "Epidemiology blockers":
        return [row for row in rows if row.get("category") == "epidemiology" and not _row_reviewed(row)]
    return rows


def ordered_editor_rows(rows: list[dict[str, Any]], *, advanced: bool = False) -> list[dict[str, Any]]:
    columns = list(STANDARD_EDITOR_COLUMNS)
    if advanced:
        columns.extend(column for column in EDITABLE_ASSUMPTION_COLUMNS if column not in columns)
    else:
        columns.append("assumptionId")
        columns.extend(
            column
            for column in EDITABLE_ASSUMPTION_COLUMNS
            if column not in columns and column not in ADVANCED_EVIDENCE_COLUMNS and column not in DERIVED_CONVERSION_COLUMNS
        )
    return [{column: row.get(column, "") for column in columns} for row in rows]


def conversion_audit_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {column: row.get(column, "") for column in CONVERSION_AUDIT_COLUMNS}
        for row in rows
        if row.get("category") == "cost"
    ]


def fatal_validation_rows(validation: dict[str, Any] | None) -> list[dict[str, Any]]:
    out = []
    for assumption_id, messages in (validation or {}).get("rowMessages", {}).items():
        fatal = [message for message in messages if _fatal_application_messages([message])]
        if fatal:
            out.append(
                {
                    "assumptionId": assumption_id,
                    "messages": "; ".join(fatal),
                }
            )
    return out


def assess_current_analysis_economic_readiness(
    config: dict[str, Any] | None,
    economics_config: dict[str, Any] | None,
    event_ledger: dict[str, Any] | None = None,
    evidence_registry: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Assess economic readiness for the selected analysis pathway only.

    Full-library readiness remains available separately. This function filters
    cost blockers to the currently selected diagnostic test, regimen and
    always-applicable programme/disease cost components.
    """
    rows = editable_assumption_rows(evidence_registry, economics_config or {})
    validation = validate_editable_assumptions(rows, economics_config or {}, config=config or {})
    row_by_id = {row.get("assumptionId"): row for row in validation["rows"]}
    messages_by_id = validation.get("rowMessages") or {}
    current_ids = _current_analysis_assumption_ids(config or {}, event_ledger or {})
    alternative_ids = _alternative_strategy_assumption_ids(current_ids)

    blockers = []
    for assumption_id in sorted(current_ids):
        row = row_by_id.get(assumption_id)
        if not row:
            continue
        messages = messages_by_id.get(assumption_id) or []
        if messages:
            blockers.append(_blocker_record(row, messages, _current_blocker_group(assumption_id), True))

    alternative_blockers = []
    for assumption_id in sorted(alternative_ids):
        row = row_by_id.get(assumption_id)
        if not row:
            continue
        messages = messages_by_id.get(assumption_id) or []
        if messages:
            alternative_blockers.append(
                _blocker_record(row, messages, "Alternative strategies", False)
            )

    cost_ids = {assumption_id for assumption_id in current_ids if assumption_id.startswith("cost.")}
    daly_ids = {assumption_id for assumption_id in current_ids if assumption_id.startswith("daly.")}
    threshold_messages = messages_by_id.get(THRESHOLD_ASSUMPTION_ID) or []
    current_cost_ready = not any(messages_by_id.get(assumption_id) for assumption_id in cost_ids)
    current_daly_ready = not any(messages_by_id.get(assumption_id) for assumption_id in daly_ids)
    current_icer_ready = bool(current_cost_ready and current_daly_ready)
    current_nmb_ready = bool(current_icer_ready and not threshold_messages)
    full_readiness = assess_apy_reference_readiness(
        config or {},
        economics_config or {},
        evidence_registry or load_apy_evidence_registry(),
    )

    grouped: dict[str, list[dict[str, Any]]] = {}
    for blocker in blockers + alternative_blockers:
        grouped.setdefault(blocker["group"], []).append(blocker)

    return {
        "currentAnalysisCostReady": current_cost_ready,
        "currentAnalysisDALYReady": current_daly_ready,
        "currentAnalysisICERReady": current_icer_ready,
        "currentAnalysisNMBReady": current_nmb_ready,
        "fullStrategyLibraryReady": bool(full_readiness.get("costReady") and full_readiness.get("dalyReady")),
        "overallReferenceEvidenceReady": bool(full_readiness.get("overallClinicianReady")),
        "currentApplicableAssumptionIds": sorted(current_ids),
        "alternativeApplicableAssumptionIds": sorted(alternative_ids),
        "currentBlockers": blockers,
        "alternativeStrategyBlockers": alternative_blockers,
        "blockerGroups": grouped,
        "validation": validation,
        "fullReadiness": full_readiness,
    }


def _apply_cost_rows(econ: dict[str, Any], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items = ensure_authoritative_cost_items(econ.get("costItems") or [])
    by_id = {item["costItemId"]: deepcopy(item) for item in items}
    for row in rows:
        assumption_id = row.get("assumptionId", "")
        if row.get("category") != "cost" or not assumption_id.startswith("cost."):
            continue
        item_id = assumption_id.split(".", 1)[1]
        item = by_id.get(item_id)
        if not item:
            continue
        value = _number_or_none(row.get("currentValue"))
        if value is None and row.get("inclusionStatus") in {"excluded", "bundled"}:
            value = 0.0
        item["originalCost"] = value
        item["originalCurrency"] = row.get("originalCurrency", "")
        item["originalPriceYear"] = row.get("originalPriceYear", "")
        item["targetCurrency"] = row.get("targetCurrency") or TARGET_CURRENCY
        item["targetPriceYear"] = row.get("targetPriceYear") or TARGET_PRICE_YEAR
        item["sourceCitation"] = row.get("sourceCitation", "")
        item["pageTableItem"] = row.get("sourceLocation", "")
        item["sourceYearStatus"] = "explicit" if row.get("originalPriceYear") else "unknown"
        item["notes"] = row.get("notes", "")
        item.setdefault("resourceUse", {})
        if row.get("costBasis"):
            item["resourceUse"]["costBasis"] = row.get("costBasis")
        item["conversionStatus"] = "not_converted"
        item["conversionApplied"] = False
        item["costRecordType"] = "source"
        item["warnings"] = []
    return list(by_id.values())


def _current_analysis_assumption_ids(config: dict[str, Any], event_ledger: dict[str, Any]) -> set[str]:
    test = _selected_test(config, event_ledger)
    regimen = _selected_regimen(config, event_ledger)
    ids = {
        TEST_COST_IDS.get(test, TEST_COST_IDS["IGRA"]),
        REGIMEN_COST_IDS.get(regimen, REGIMEN_COST_IDS["3HP"]),
        "cost.active_tb_disease",
        "cost.false_positive_incremental",
        "cost.program_setup",
        "cost.program_running",
        "cost.tpt_adr_management",
        *DALY_ASSUMPTION_IDS,
        THRESHOLD_ASSUMPTION_ID,
    }
    return {assumption_id for assumption_id in ids if assumption_id}


def _alternative_strategy_assumption_ids(current_ids: set[str]) -> set[str]:
    alternatives = set(TEST_COST_IDS.values()) | set(REGIMEN_COST_IDS.values())
    return alternatives - current_ids


def _selected_test(config: dict[str, Any], event_ledger: dict[str, Any]) -> str:
    for value in (
        config.get("test"),
        config.get("screeningTest"),
        config.get("testType"),
        (config.get("screening") or {}).get("test"),
        (event_ledger.get("metadata") or {}).get("test"),
    ):
        text = str(value or "").strip().upper()
        if text in TEST_COST_IDS:
            return text
    return "IGRA"


def _selected_regimen(config: dict[str, Any], event_ledger: dict[str, Any]) -> str:
    for value in (
        config.get("regimen"),
        config.get("preventiveRegimen"),
        config.get("regimenKey"),
        (config.get("treatment") or {}).get("regimen"),
        (config.get("intervention") or {}).get("regimen"),
        (event_ledger.get("metadata") or {}).get("regimen"),
    ):
        text = str(value or "").strip().upper().removeprefix("X")
        if text in REGIMEN_COST_IDS:
            return text
    return "3HP"


def _current_blocker_group(assumption_id: str) -> str:
    if assumption_id in {TEST_COST_IDS["IGRA"], TEST_COST_IDS["TST"]} or assumption_id in set(REGIMEN_COST_IDS.values()):
        return "Current selected test and regimen"
    if assumption_id in {"cost.program_setup", "cost.program_running", "cost.false_positive_incremental", "cost.tpt_adr_management"}:
        return "Programme and delivery costs"
    if assumption_id == "cost.active_tb_disease":
        return "Active-TB care"
    if assumption_id.startswith("daly."):
        return "DALY assumptions"
    if assumption_id == THRESHOLD_ASSUMPTION_ID:
        return "Threshold for NMB"
    return "Current selected test and regimen"


def _blocker_record(
    row: dict[str, Any],
    messages: list[str],
    group: str,
    current_analysis: bool,
) -> dict[str, Any]:
    return {
        "assumptionId": row.get("assumptionId", ""),
        "description": row.get("description", ""),
        "group": group,
        "whyBlocking": "; ".join(messages),
        "missingFields": ", ".join(_missing_fields_from_messages(messages)),
        "editableRowId": row.get("assumptionId", ""),
        "currentAnalysis": current_analysis,
    }


def _missing_fields_from_messages(messages: list[str]) -> list[str]:
    fields: set[str] = set()
    for message in messages:
        lower = message.lower()
        if "source" in lower:
            fields.add("source citation")
        if "price year" in lower:
            fields.add("original price year")
        if "currency" in lower:
            fields.add("currency")
        if "cost basis" in lower or "basis" in lower:
            fields.add("cost basis")
        if "value" in lower or "cost is missing" in lower:
            fields.add("current value")
        if "reviewed" in lower or "status" in lower:
            fields.add("review status")
        if "provisional" in lower:
            fields.add("provisional flag")
        if "rationale" in lower or "notes" in lower:
            fields.add("notes")
        if "bundled" in lower:
            fields.add("bundling details")
    return sorted(fields) or ["see validation message"]


def _apply_threshold_rows(econ: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    for row in rows:
        if row.get("assumptionId") != "threshold.gdp_per_capita":
            continue
        threshold = econ.setdefault("threshold", {})
        threshold["value"] = _number_or_none(row.get("currentValue"))
        threshold["currency"] = row.get("targetCurrency") or row.get("originalCurrency") or TARGET_CURRENCY
        threshold["referenceYear"] = row.get("targetPriceYear") or row.get("originalPriceYear") or None
        threshold["source"] = row.get("sourceCitation", "")
        threshold["status"] = row.get("reviewStatus", "")
        threshold["notes"] = row.get("notes", "")


def _apply_daly_rows(econ: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    daly = econ.setdefault("dalyAssumptions", {})
    for row in rows:
        assumption_id = row.get("assumptionId", "")
        if assumption_id not in DALY_ROW_TO_CONFIG:
            continue
        field, kind = DALY_ROW_TO_CONFIG[assumption_id]
        if kind == "numeric":
            daly[field] = _assumption_record(row)
        elif kind == "optional_tpt":
            if row.get("inclusionStatus") == "excluded":
                daly["includeTPTHealthLoss"] = False
                daly["tptHealthLossExclusionStatus"] = row.get("reviewStatus", "")
                daly["tptHealthLossExclusionRationale"] = row.get("notes", "")
            else:
                daly["includeTPTHealthLoss"] = True
                daly[field] = _assumption_record(row)
        elif kind == "optional_adr":
            if row.get("inclusionStatus") == "excluded":
                daly["includeADRHealthLoss"] = False
                daly["adrHealthLossExclusionStatus"] = row.get("reviewStatus", "")
                daly["adrHealthLossExclusionRationale"] = row.get("notes", "")
            else:
                daly["includeADRHealthLoss"] = True
                daly[field] = _assumption_record(row)
        elif kind == "post_tb":
            daly["includePostTBSequelae"] = row.get("inclusionStatus") == "included"
            daly["postTBSequelaeStatus"] = row.get("reviewStatus", "")
            daly["postTBSequelaeExclusionRationale"] = row.get("notes", "")


def _assumption_record(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "value": _number_or_none(row.get("currentValue")),
        "source": row.get("sourceCitation", ""),
        "status": row.get("reviewStatus", ""),
        "notes": row.get("notes", ""),
        "provisional": row.get("provisional", True),
        "unit": row.get("unit", ""),
    }


def _validate_daly_row(row: dict[str, Any]) -> list[str]:
    if row.get("inclusionStatus") == "excluded":
        return _validate_exclusion_row(row)
    messages = []
    if row.get("reviewStatus") not in REVIEWED_NUMERICAL_STATUSES:
        messages.append("DALY numerical assumption requires reviewed numerical status.")
    if _number_or_none(row.get("currentValue")) is None:
        messages.append("DALY numerical value is missing or non-numeric.")
    if not str(row.get("unit") or "").strip():
        messages.append("DALY assumption unit is missing.")
    if not str(row.get("sourceCitation") or "").strip():
        messages.append("DALY source citation is missing.")
    value = _number_or_none(row.get("currentValue"))
    if value is not None and value < 0:
        messages.append("DALY value must be non-negative.")
    if row.get("assumptionId") in {"daly.active_tb_disability_weight", "daly.tb_case_fatality_risk"} and value is not None and value > 1:
        messages.append("DALY probability or disability-weight value must be in [0, 1].")
    return messages


def _validate_threshold_row(row: dict[str, Any], econ: dict[str, Any]) -> list[str]:
    messages = []
    value = _number_or_none(row.get("currentValue"))
    if value is None or value <= 0:
        messages.append("Threshold value must be a positive number.")
    if row.get("reviewStatus") not in REVIEWED_NUMERICAL_STATUSES:
        messages.append("Threshold requires reviewed numerical status.")
    if not str(row.get("sourceCitation") or "").strip():
        messages.append("Threshold source citation is missing.")
    metadata = econ.get("metadata") or {}
    if row.get("targetCurrency") and row.get("targetCurrency") != metadata.get("targetCurrency", TARGET_CURRENCY):
        messages.append("Threshold currency must match the analysis target currency.")
    if row.get("targetPriceYear") and row.get("targetPriceYear") != metadata.get("targetPriceYear", TARGET_PRICE_YEAR):
        messages.append("Threshold reference year must match the analysis target price year.")
    return messages


def _validate_exclusion_row(row: dict[str, Any]) -> list[str]:
    messages = []
    if row.get("reviewStatus") != REVIEWED_EXCLUSION_STATUS:
        messages.append("Exclusion requires reviewed_exclusion status.")
    if not str(row.get("notes") or "").strip():
        messages.append("Reviewed exclusion requires rationale and scope in notes.")
    return messages


def _validate_bundled_row(row: dict[str, Any], rows: list[dict[str, Any]]) -> list[str]:
    messages = []
    destination = row.get("bundledIntoAssumptionId", "")
    ids = {item.get("assumptionId") for item in rows}
    if not destination or destination not in ids:
        messages.append("Bundled row requires an existing bundled-into assumption id.")
    if row.get("reviewStatus") not in {REVIEWED_EXCLUSION_STATUS, *REVIEWED_NUMERICAL_STATUSES}:
        messages.append("Bundling requires reviewed status.")
    if row.get("currentValue") not in (None, ""):
        messages.append("Bundled row must not also be separately costed.")
    return messages


def _fatal_application_messages(messages: list[str]) -> bool:
    fatal_fragments = [
        "must be in [0, 1]",
        "must be non-negative",
        "must match the analysis",
        "incompatible with the event mapping",
        "Bundled row requires an existing",
        "must not also be separately costed",
    ]
    return any(
        any(fragment in message for fragment in fatal_fragments)
        for message in messages
    )


def _normalise_edit_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {column: _clean_value(row.get(column)) for column in EDITABLE_ASSUMPTION_COLUMNS}
    out["provisional"] = _bool(row.get("provisional"))
    out["reviewStatus"] = normalise_status_value(out.get("reviewStatusLabel") or out.get("reviewStatus"))
    out["reviewStatusLabel"] = status_label(out["reviewStatus"])
    out["inclusionStatus"] = normalise_inclusion_value(out.get("inclusionStatusLabel") or out.get("inclusionStatus"))
    out["inclusionStatusLabel"] = inclusion_label(out["inclusionStatus"])
    return out


def _cost_basis_lookup(economics_config: dict[str, Any]) -> dict[str, str]:
    items = ensure_authoritative_cost_items(economics_config.get("costItems") or [])
    return {
        f"cost.{item.get('costItemId')}": str((item.get("resourceUse") or {}).get("costBasis") or "")
        for item in items
    }


def _conversion_lookup(economics_config: dict[str, Any]) -> dict[str, dict[str, Any]]:
    out = {}
    for item in normalise_cost_table(ensure_authoritative_cost_items(economics_config.get("costItems") or [])):
        warnings = item.get("warnings") or []
        out[f"cost.{item.get('costItemId')}"] = {
            "conversionStatus": item.get("conversionStatus", ""),
            "inflationIndexId": item.get("inflationIndexId", ""),
            "sourceYearIndexValue": item.get("sourceYearIndexValue", ""),
            "targetYearIndexValue": item.get("targetYearIndexValue", ""),
            "inflationFactor": item.get("inflationFactor", ""),
            "convertedTargetYearCost": item.get("convertedTargetYearCost", ""),
            "conversionWarnings": "; ".join(str(warning) for warning in warnings),
        }
    return out


def _row_reviewed(row: dict[str, Any]) -> bool:
    if row.get("provisional"):
        return False
    if row.get("inclusionStatus") == "excluded":
        return row.get("reviewStatus") == REVIEWED_EXCLUSION_STATUS and bool(str(row.get("notes") or "").strip())
    return row.get("reviewStatus") in REVIEWED_NUMERICAL_STATUSES and bool(str(row.get("sourceCitation") or "").strip())


def _number_or_none(value: Any) -> float | None:
    if value in (None, "", []):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes"}


def _clean_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _serialise_cell(value: Any) -> Any:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, dict, tuple)):
        return str(value)
    return "" if value is None else value
