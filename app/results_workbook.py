from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import subprocess
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.epidemiology_inputs import (
    AGE_GROUP_LABELS,
    ADVANCED_RISK_FACTORS,
    PRINCIPAL_RISK_FACTORS,
    RISK_FACTOR_LABELS,
    prevalence_source_label,
)
from engine.apy.costing import normalise_cost_table, unresolved_cost_warnings
from engine.apy.scenario import DIRECT_EFFECTS_SCOPE_STATEMENT, scenario_from_config


PROGRESSION_MULTIPLIERS = {
    "contact": "from default_data.csv unless overridden",
    "MJ": "from default_data.csv unless overridden",
    "renal": 3.6,
    "diabetes": "from default_data.csv unless overridden",
    "smoking": "from default_data.csv unless overridden",
    "cld": "from default_data.csv unless overridden",
    "alcohol": "from default_data.csv unless overridden",
    "female": "",
    "BCG": "",
}


def build_results_workbook(
    *,
    config: dict[str, Any],
    bundle: dict[str, Any],
    backend_status: dict[str, Any] | None = None,
    economics_results: dict[str, Any] | None = None,
    economics_config: dict[str, Any] | None = None,
    results_stale: bool = False,
    dirty_economics: bool = False,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _write_readme(wb, config, bundle, backend_status, results_stale)
    _write_scenario_inputs(wb, config, backend_status)
    _write_scenario_contract(wb, config)
    _write_risk_factor_inputs(wb, config)
    _write_rows_sheet(
        wb,
        "Headline_results",
        _humanise_metric_rows(bundle.get("headline", {}).get("keyMetricsRows") or []),
    )
    _write_rows_sheet(
        wb,
        "Summary_results",
        _humanise_metric_rows(bundle.get("headline", {}).get("summaryRows") or []),
    )
    _write_natural_history(wb, bundle)
    _write_event_ledger(wb, bundle)
    _write_technical_metadata(wb, config, bundle, backend_status, results_stale)
    _write_economics(wb, economics_results, economics_config, dirty_economics, results_stale)
    _autosize_all(wb)
    output = BytesIO()
    wb.save(output)
    payload = output.getvalue()
    load_workbook(BytesIO(payload), read_only=True).close()
    return payload


def _write_readme(
    wb: Workbook,
    config: dict[str, Any],
    bundle: dict[str, Any],
    backend_status: dict[str, Any] | None,
    results_stale: bool,
) -> None:
    metadata = bundle.get("metadata", {})
    experimental = bool((backend_status or {}).get("experimental") or metadata.get("backend") == "python")
    rows = [
        ("Scenario name", config.get("scenarioLabel", metadata.get("scenarioLabel", ""))),
        ("Purpose", "Structured APY epidemiological results workbook for public-health planning."),
        ("Interpretation guardrail", "Targeting supports sequencing and planning, not exclusion or denial of screening or treatment."),
        ("Transmission", DIRECT_EFFECTS_SCOPE_STATEMENT),
        ("Backend warning", "Python APY backend is experimental; MATLAB remains the reference backend." if experimental else "MATLAB reference backend."),
        ("Stale results", "Yes - rerun before using for decisions." if results_stale else "No."),
    ]
    _write_key_value_sheet(wb, "Read_me", rows)


def _write_scenario_inputs(
    wb: Workbook,
    config: dict[str, Any],
    backend_status: dict[str, Any] | None,
) -> None:
    rows = [
        ("backend", (backend_status or {}).get("name", "")),
        ("scenario label", config.get("scenarioLabel")),
        ("cohort size", config.get("N")),
        ("replicates", config.get("nReps")),
        ("seed", config.get("seed")),
        ("screening coverage", config.get("screenCoverage"), "percentage"),
        ("screening window", config.get("screenWindow"), "years"),
        ("screening window years", config.get("screeningWindowYears"), "years"),
        ("early progression period", config.get("earlyProgressionPeriodYears"), "years"),
        ("active-TB calibration horizon", config.get("activeTBCalibrationHorizonYears"), "years"),
        ("follow-up horizon", config.get("followHorizon"), "years"),
        ("follow-up horizon years", config.get("followUpHorizonYears"), "years"),
        ("screening strategy", config.get("screeningStrategy")),
        ("test", config.get("testType")),
        ("test sensitivity", config.get("testSensitivity"), "probability"),
        ("test specificity", config.get("testSpecificity"), "probability"),
        ("TST sensitivity", config.get("tstSensitivity"), "probability"),
        ("TST specificity with BCG", config.get("tstSpecificityBCG"), "probability"),
        ("TST specificity without BCG", config.get("tstSpecificityNoBCG"), "probability"),
        ("regimen", config.get("regimen")),
        ("LTBI prevalence source", prevalence_source_label(config.get("ltbiPrevalence"))),
        ("LTBI prevalence input", config.get("ltbiPrevalence"), "percentage"),
        ("active-TB prevalence source", prevalence_source_label(config.get("activeTBPrevalence"))),
        ("active-TB prevalence input", config.get("activeTBPrevalence"), "percentage"),
        ("TPT start probability", config.get("pStartTPT"), "percentage"),
        ("regimen completion probability", config.get("regimenPComplete"), "percentage"),
        ("ADR stop probability", config.get("regimenADRstop"), "percentage"),
        ("full-course efficacy", config.get("regimenEffFull"), "percentage"),
    ]
    _write_rows_sheet(
        wb,
        "Scenario_inputs",
        [
            {"Input": item, "Value": value, "Units / format": unit or ""}
            for item, value, *rest in rows
            for unit in [rest[0] if rest else ""]
        ],
    )


def _write_scenario_contract(wb: Workbook, config: dict[str, Any]) -> None:
    scenario = scenario_from_config(config)
    rows = [
        {"Field": "contractVersion", "Value": scenario.get("contractVersion")},
        {"Field": "scenarioName", "Value": scenario.get("scenarioName")},
        {"Field": "scenarioVersion", "Value": scenario.get("scenarioVersion")},
        {"Field": "populationPresetId", "Value": scenario.get("populationPresetId")},
        {"Field": "populationSize", "Value": scenario.get("populationSize")},
        {"Field": "ageProfileSource", "Value": scenario.get("ageProfileSource")},
        {"Field": "ltbiPrevalenceAssumptions", "Value": scenario.get("ltbiPrevalenceAssumptions")},
        {"Field": "riskFactorAssumptions", "Value": scenario.get("riskFactorAssumptions")},
        {"Field": "targetingCriteria", "Value": scenario.get("targetingCriteria")},
        {"Field": "eligible", "Value": scenario.get("eligible")},
        {"Field": "screened", "Value": scenario.get("screened")},
        {"Field": "screeningWindowYears", "Value": scenario.get("screeningWindowYears")},
        {"Field": "earlyProgressionPeriodYears", "Value": scenario.get("earlyProgressionPeriodYears")},
        {"Field": "activeTBCalibrationHorizonYears", "Value": scenario.get("activeTBCalibrationHorizonYears")},
        {"Field": "followUpHorizonYears", "Value": scenario.get("followUpHorizonYears")},
        {"Field": "comparator", "Value": scenario.get("comparator")},
        {"Field": "intervention", "Value": scenario.get("intervention")},
        {"Field": "sourcesAndNotes", "Value": scenario.get("sourcesAndNotes")},
        {"Field": "scopeStatement", "Value": scenario.get("scopeStatement")},
    ]
    _write_rows_sheet(wb, "Scenario_contract", rows)


def _write_risk_factor_inputs(wb: Workbook, config: dict[str, Any]) -> None:
    risk_prev = config.get("riskPrev") or {}
    rows = []
    for key in PRINCIPAL_RISK_FACTORS + ADVANCED_RISK_FACTORS:
        value = risk_prev.get(key)
        source = "default_data.csv" if value is None or value == [] else "custom"
        values = _risk_values(value)
        rows.append(
            {
                "Risk factor": RISK_FACTOR_LABELS[key],
                "Config field": f"riskPrev.{key}",
                "Source": source,
                "Prevalence 0-4": values[0],
                "Prevalence 5-14": values[1],
                "Prevalence >=15": values[2],
                "Progression multiplier": PROGRESSION_MULTIPLIERS.get(key, ""),
            }
        )
    _write_rows_sheet(wb, "Risk_factor_inputs", rows)


def _write_natural_history(wb: Workbook, bundle: dict[str, Any]) -> None:
    dynamic = bundle.get("technical", {}).get("dynamicComparison", {})
    if isinstance(dynamic, dict):
        rows = [
            {"Field": key, "Value": value}
            for key, value in dynamic.items()
            if key != "metricRows"
        ]
        metric_rows = dynamic.get("metricRows") or []
        rows.extend({"Field": f"metric.{row.get('Metric')}", "Value": row.get("Median")} for row in metric_rows)
    else:
        rows = [{"Field": "Natural history", "Value": "Not available"}]
    _write_rows_sheet(wb, "Natural_history", rows)


def _write_event_ledger(wb: Workbook, bundle: dict[str, Any]) -> None:
    ledger = bundle.get("technical", {}).get("eventLedger") or {}
    downloads = bundle.get("downloads", {})
    _write_rows_sheet(
        wb,
        "Event_ledger_totals",
        _rows_from_table(downloads.get("eventLedgerTotals") or ledger.get("replicateTotals")),
    )
    _write_rows_sheet(
        wb,
        "Event_ledger_annual",
        _rows_from_table(downloads.get("eventLedgerAnnual") or ledger.get("annualEvents")),
    )
    _write_rows_sheet(
        wb,
        "Event_definitions",
        _rows_from_table(downloads.get("eventDefinitions") or ledger.get("definitions")),
    )
    validation = ledger.get("validation") if isinstance(ledger, dict) else {}
    validation = validation or {}
    rows = [{"Field": "isValid", "Value": validation.get("isValid")}]
    for issue in validation.get("errors", []) or []:
        rows.append({"Field": "error", "Value": issue})
    for issue in validation.get("warnings", []) or []:
        rows.append({"Field": "warning", "Value": issue})
    _write_rows_sheet(wb, "Event_ledger_validation", rows)


def _write_technical_metadata(
    wb: Workbook,
    config: dict[str, Any],
    bundle: dict[str, Any],
    backend_status: dict[str, Any] | None,
    results_stale: bool,
) -> None:
    metadata = bundle.get("metadata", {})
    technical = bundle.get("technical", {})
    calibration = technical.get("calibration", {})
    rows = [
        ("backend", metadata.get("backend") or (backend_status or {}).get("name")),
        ("model version", metadata.get("modelVersion")),
        ("contract version", metadata.get("contractVersion")),
        ("Git branch", _git_value(["git", "branch", "--show-current"])),
        ("Git commit", _git_value(["git", "rev-parse", "--short", "HEAD"])),
        ("run date/time", datetime.now(timezone.utc).isoformat()),
        ("results stale at export", results_stale),
        ("csvFile", config.get("csvFile")),
        ("ageDistributionFile", config.get("ageDistributionFile") or "default_age_distribution.csv if available next to default_data.csv"),
        ("Python limitation", "Python APY backend is experimental and does not include the progression attributable-risk add-on."),
    ]
    for key, value in calibration.items():
        rows.append((f"calibration.{key}", value))
    _write_key_value_sheet(wb, "Technical_metadata", rows)


def _write_economics(
    wb: Workbook,
    economics_results: dict[str, Any] | None,
    economics_config: dict[str, Any] | None,
    dirty_economics: bool,
    results_stale: bool,
) -> None:
    ws = wb.create_sheet("Economics")
    if not economics_results:
        _write_rows_to_existing_sheet(
            ws,
            [{"Status": "Economics not run", "Value": "", "Notes": "No zero values have been substituted for missing economics outputs."}],
        )
        _write_economics_assumptions(wb, economics_results, economics_config)
        return
    rows = economics_results.get("summaryRows") or []
    if not rows:
        rows = [{"Status": "Economics results present", "Value": "", "Notes": "No summary rows returned."}]
    _write_rows_to_existing_sheet(ws, rows)
    start = ws.max_row + 2
    ws.cell(start, 1, "Economics stale").font = Font(bold=True)
    ws.cell(start, 2, bool(dirty_economics or results_stale))
    ws.cell(start + 1, 1, "Economics config supplied").font = Font(bold=True)
    ws.cell(start + 1, 2, bool(economics_config))
    _write_economics_assumptions(wb, economics_results, economics_config)


def _write_economics_assumptions(
    wb: Workbook,
    economics_results: dict[str, Any] | None,
    economics_config: dict[str, Any] | None,
) -> None:
    source = economics_results or economics_config or {}
    metadata = source.get("metadata", {}) if isinstance(source, dict) else {}
    rows = [
        {"Field": "perspective", "Value": metadata.get("perspective")},
        {"Field": "targetCurrency", "Value": metadata.get("targetCurrency", metadata.get("currencyCode"))},
        {"Field": "targetPriceYear", "Value": metadata.get("targetPriceYear", metadata.get("priceYear"))},
        {"Field": "scopeStatement", "Value": metadata.get("scopeStatement", DIRECT_EFFECTS_SCOPE_STATEMENT)},
        {"Field": "costNormalisation", "Value": source.get("costNormalisation", {}) if isinstance(source, dict) else {}},
        {"Field": "discounting", "Value": source.get("discounting", {}) if isinstance(source, dict) else {}},
        {"Field": "healthOutcome", "Value": source.get("healthOutcome", {}) if isinstance(source, dict) else {}},
        {"Field": "threshold", "Value": source.get("threshold", {}) if isinstance(source, dict) else {}},
    ]
    _write_rows_sheet(wb, "Economics_assumptions", rows)

    cost_items = []
    if economics_results and isinstance(source, dict):
        cost_items = source.get("costItems") or source.get("costItemsTable") or []
    if not cost_items and isinstance(economics_config, dict):
        cost_items = normalise_cost_table(economics_config.get("costItems") or [])
    _write_rows_sheet(wb, "Cost_normalisation", cost_items)

    warnings = unresolved_cost_warnings(cost_items)
    if isinstance(source, dict):
        threshold = source.get("threshold", {})
        if isinstance(threshold, dict) and threshold.get("value") in (None, "", []):
            warnings.append("threshold.value: unresolved GDP-per-capita benchmark value")
    _write_rows_sheet(
        wb,
        "Unresolved_assumptions",
        [{"Warning": warning} for warning in warnings] or [{"Warning": ""}],
    )


def _humanise_metric_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        metric = row.get("Metric", row.get("metric", ""))
        out.append(
            {
                "Metric": metric,
                "Label": _metric_label(str(metric)),
                "Median": row.get("Median"),
                "Low95": row.get("Low95"),
                "High95": row.get("High95"),
                "Units": _metric_units(str(metric)),
            }
        )
    return out


def _metric_label(metric: str) -> str:
    labels = {
        "nScreened": "People screened",
        "nTestPositiveNonActive": "Test-positive, non-active TB",
        "nFalsePositiveTreated": "False-positive people starting TPT",
        "nTotalCoursesStarted": "TPT courses started",
        "nTotalCoursesCompleted": "TPT courses completed",
        "nCuredInfection": "MTB infections cured/protected",
        "nPreventedActiveTB": "Direct active TB cases prevented",
        "NNS_preventActiveTB": "NNS to prevent one active TB case",
        "NNT_started_preventActiveTB": "TPT starts per active TB case prevented",
    }
    return labels.get(metric, metric)


def _metric_units(metric: str) -> str:
    if metric.startswith("n") or metric in {"NNS_preventActiveTB", "NNS_cureInfection"}:
        return "people / cases"
    if "Rate" in metric or "prop" in metric or "relative" in metric:
        return "proportion"
    return ""


def _risk_values(value: Any) -> tuple[Any, Any, Any]:
    if value is None or value == []:
        return ("default", "default", "default")
    if isinstance(value, (list, tuple)) and len(value) == 3:
        return tuple(value)
    return (value, value, value)


def _write_key_value_sheet(wb: Workbook, title: str, rows: list[tuple[Any, Any]]) -> None:
    ws = wb.create_sheet(_safe_sheet_name(title))
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for key, value in rows:
        ws.append([_clean_cell(key), _clean_cell(value)])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_rows_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(_safe_sheet_name(title))
    _write_rows_to_existing_sheet(ws, rows)


def _write_rows_to_existing_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        rows = [{"Status": "No rows available"}]
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([_clean_cell(row.get(header)) for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _rows_from_table(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if hasattr(value, "to_dict"):
        return value.to_dict(orient="records")
    return []


def _autosize_all(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for column in range(1, ws.max_column + 1):
            width = 12
            for row in range(1, min(ws.max_row, 100) + 1):
                value = ws.cell(row, column).value
                if value is not None:
                    width = max(width, min(len(str(value)) + 2, 60))
            ws.column_dimensions[get_column_letter(column)].width = width
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                    header = str(ws.cell(1, cell.column).value or "").lower()
                    if "prevalence" in header or "coverage" in header or "probability" in header:
                        cell.number_format = "0.0%"


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return str(value)
    try:
        if hasattr(value, "item"):
            return value.item()
    except Exception:
        pass
    return value


def _safe_sheet_name(value: str) -> str:
    return value[:31]


def _git_value(args: list[str]) -> str:
    try:
        return subprocess.check_output(args, cwd=".", text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return ""
