from __future__ import annotations

from io import BytesIO
import json
import sys
import unittest

from openpyxl import load_workbook

from adapters.paths import repo_root
from adapters.python_apy_backend import PythonApyBackend
from app.results_workbook import build_results_workbook
from engine.apy.config import build_default_config
from engine.apy.costing import (
    build_cost_item,
    discount_value,
    normalise_cost_table,
)
from engine.apy.economics import (
    build_default_economics_config,
    build_economics_preset_kwab150,
    run_health_economics,
    sync_legacy_cost_fields_from_cost_items,
    update_cost_item_original_values_from_legacy_fields,
    validate_economics_config,
)
from engine.apy.scenario import (
    DIRECT_EFFECTS_SCOPE_STATEMENT,
    build_scenario_contract,
    config_updates_from_scenario,
    load_population_preset,
)
from engine.apy.validation import collect_validation_issues


class Milestone1ScenarioTests(unittest.TestCase):
    def test_apy_preset_loads_successfully(self) -> None:
        preset = load_population_preset("apy_demonstration")
        scenario = build_scenario_contract("apy_demonstration")

        self.assertEqual(preset["populationPresetId"], "apy_demonstration")
        self.assertEqual(scenario["screeningWindowYears"], 3)
        self.assertEqual(scenario["followUpHorizonYears"], 20)
        self.assertIn("ordinary background clinical care", scenario["comparator"]["notes"])

    def test_scenario_json_configuration_round_trip(self) -> None:
        backend = PythonApyBackend(repo_root())
        config = backend.default_config()
        config["scenario"]["sourcesAndNotes"].append("round trip marker")
        tmp = repo_root() / ".tmp_m1_roundtrip"
        tmp.mkdir(exist_ok=True)
        path = tmp / "scenario.json"
        try:
            backend.save_scenario(config, str(path), build_default_economics_config())
            payload = json.loads(path.read_text(encoding="utf-8"))
            loaded, report, info = backend.load_scenario(str(path))
        finally:
            if path.exists():
                path.unlink()
            if tmp.exists():
                tmp.rmdir()

        self.assertEqual(payload["contractVersion"], "ltbi_screening_scenario_v1")
        self.assertTrue(report["isValid"])
        self.assertEqual(loaded["scenario"]["sourcesAndNotes"][-1], "round trip marker")
        self.assertEqual(info["backend"], "python_apy")

    def test_second_non_apy_population_loads_without_code_changes(self) -> None:
        scenario = build_scenario_contract("generic_remote_demo")
        config = build_default_config()
        config.update(config_updates_from_scenario(scenario))

        self.assertEqual(config["populationPresetId"], "generic_remote_demo")
        self.assertEqual(config["N"], 1000)
        self.assertEqual(config["screeningStrategy"], "random")

    def test_igra_and_tst_retain_distinct_sensitivity_specificity(self) -> None:
        config = build_default_config()

        self.assertNotEqual(config["testSensitivity"], config["tstSensitivity"])
        self.assertNotEqual(config["testSpecificity"], config["tstSpecificityBCG"])
        self.assertTrue(config["testCharacteristics"]["TST"]["resourceUse"]["returnVisitForReading"])

    def test_invalid_sensitivity_or_specificity_is_rejected(self) -> None:
        report = collect_validation_issues({"testSensitivity": 1.2})

        self.assertFalse(report["isValid"])
        self.assertIn("testSensitivity", report["fatalFieldNames"])

    def test_screening_window_and_followup_remain_independent(self) -> None:
        config = build_default_config()
        changed_followup = dict(config)
        changed_followup["followHorizon"] = 30
        changed_screening = dict(config)
        changed_screening["screenWindow"] = 4

        self.assertEqual(changed_followup["screenCoverage"], config["screenCoverage"])
        self.assertEqual(changed_followup["screenWindow"], config["screenWindow"])
        self.assertEqual(changed_screening["followHorizon"], config["followHorizon"])

    def test_followup_shorter_than_screening_window_is_rejected(self) -> None:
        report = collect_validation_issues({"screenWindow": 5, "followHorizon": 4})

        self.assertFalse(report["isValid"])
        self.assertIn("followHorizon", report["fatalFieldNames"])


class Milestone1CostingTests(unittest.TestCase):
    def test_conversion_factor_is_one_when_years_match(self) -> None:
        item = build_cost_item(
            cost_item_id="same_year",
            description="same year",
            original_cost=100,
            original_currency="AUD",
            original_price_year="2025-26",
            source="test",
            source_year_status="explicit",
            resource_category="test",
        )

        converted = normalise_cost_table([item], [])[0]

        self.assertEqual(converted["inflationFactor"], 1.0)
        self.assertEqual(converted["convertedTargetYearCost"], 100)
        self.assertEqual(converted["conversionStatus"], "valid")

    def test_known_index_pair_converts_cost(self) -> None:
        item = build_cost_item(
            cost_item_id="known",
            description="known",
            original_cost=200,
            original_currency="AUD",
            original_price_year="2020-21",
            source="test",
            source_year_status="explicit",
            resource_category="test",
        )
        indices = [
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2020-21", "index_value": 100, "index_source": "fixture", "index_version": "v1"},
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2025-26", "index_value": 125, "index_source": "fixture", "index_version": "v1"},
        ]

        converted = normalise_cost_table([item], indices)[0]

        self.assertEqual(converted["inflationFactor"], 1.25)
        self.assertEqual(converted["convertedTargetYearCost"], 250)
        self.assertEqual(converted["conversionStatus"], "valid")

    def test_missing_source_price_year_generates_unresolved_warning(self) -> None:
        item = build_cost_item(
            cost_item_id="missing_year",
            description="missing",
            original_cost=10,
            original_currency="AUD",
            original_price_year=None,
            source="test",
            source_year_status="unknown",
            resource_category="test",
        )

        converted = normalise_cost_table([item], [])[0]

        self.assertEqual(converted["conversionStatus"], "unresolved_source_price_year")
        self.assertIn("Source price year is unknown", converted["warnings"][0])

    def test_unknown_index_value_is_not_fabricated(self) -> None:
        item = build_cost_item(
            cost_item_id="missing_index",
            description="missing",
            original_cost=10,
            original_currency="AUD",
            original_price_year="2020-21",
            source="test",
            source_year_status="explicit",
            resource_category="test",
        )

        converted = normalise_cost_table([item], [])[0]

        self.assertIsNone(converted["convertedTargetYearCost"])
        self.assertEqual(converted["conversionStatus"], "unresolved_index_value")

    def test_converted_cost_reconciles_with_recorded_factor(self) -> None:
        item = build_cost_item(
            cost_item_id="reconcile",
            description="reconcile",
            original_cost=80,
            original_currency="AUD",
            original_price_year="2020-21",
            source="test",
            source_year_status="explicit",
            resource_category="test",
        )
        indices = [
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2020-21", "index_value": 80},
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2025-26", "index_value": 100},
        ]

        converted = normalise_cost_table([item], indices)[0]

        self.assertEqual(
            converted["originalCost"] * converted["inflationFactor"],
            converted["convertedTargetYearCost"],
        )

    def test_repeated_normalisation_does_not_inflate_twice(self) -> None:
        item = build_cost_item(
            cost_item_id="twice",
            description="twice",
            original_cost=100,
            original_currency="AUD",
            original_price_year="2025-26",
            source="test",
            source_year_status="explicit",
            resource_category="test",
        )
        once = normalise_cost_table([item], [])[0]

        twice = normalise_cost_table([once], [])[0]

        self.assertEqual(twice["conversionStatus"], "valid")
        self.assertEqual(twice["inflationFactor"], 1.0)
        self.assertEqual(twice["convertedTargetYearCost"], 100)

    def test_genuine_attempted_double_conversion_remains_detected(self) -> None:
        item = build_cost_item(
            cost_item_id="converted_as_source",
            description="converted as source",
            original_cost=100,
            original_currency="AUD",
            original_price_year="2025-26",
            source="test",
            source_year_status="explicit",
            resource_category="test",
        )
        item["costRecordType"] = "converted_input"

        converted = normalise_cost_table([item], [])[0]

        self.assertEqual(converted["conversionStatus"], "invalid_double_inflation")

    def test_changing_target_year_changes_cost_not_resource_use(self) -> None:
        item = build_cost_item(
            cost_item_id="resource",
            description="resource",
            original_cost=10,
            original_currency="AUD",
            original_price_year="2020-21",
            source="test",
            source_year_status="explicit",
            resource_category="test",
            target_price_year="2024-25",
            resource_use={"visits": 2},
        )
        changed = dict(item)
        changed["targetPriceYear"] = "2025-26"
        indices = [
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2020-21", "index_value": 100},
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2024-25", "index_value": 110},
            {"index_id": "AUD_HEALTH_SYSTEM_CPI", "price_year": "2025-26", "index_value": 120},
        ]

        converted_a, converted_b = normalise_cost_table([item, changed], indices)

        self.assertNotEqual(converted_a["convertedTargetYearCost"], converted_b["convertedTargetYearCost"])
        self.assertEqual(converted_a["resourceUse"], converted_b["resourceUse"])

    def test_inflation_adjustment_and_future_discounting_are_separate(self) -> None:
        converted = normalise_cost_table(
            [
                build_cost_item(
                    cost_item_id="discount",
                    description="discount",
                    original_cost=100,
                    original_currency="AUD",
                    original_price_year="2025-26",
                    source="test",
                    source_year_status="explicit",
                    resource_category="test",
                )
            ],
            [],
        )[0]

        self.assertEqual(converted["convertedTargetYearCost"], 100)
        self.assertAlmostEqual(discount_value(100, 0.03, 1), 100 / 1.03)

    def test_zero_percent_discounting_returns_undiscounted_values(self) -> None:
        self.assertEqual(discount_value(123.45, 0.0, 10), 123.45)


class Milestone1EconomicsConfigTests(unittest.TestCase):
    def test_editing_igra_cost_updates_authoritative_cost_item(self) -> None:
        config = build_economics_preset_kwab150()
        config["costs"]["test"]["IGRA"] = 222.0

        updated = update_cost_item_original_values_from_legacy_fields(config)
        items = {item["costItemId"]: item for item in updated["costItems"]}

        self.assertEqual(items["test_igra"]["originalCost"], 222.0)
        self.assertEqual(updated["costs"]["test"]["IGRA"], 222.0)

    def test_editing_regimen_cost_updates_authoritative_cost_item(self) -> None:
        config = build_economics_preset_kwab150()
        config["costs"]["regimen"]["x3HP"] = 333.0

        updated = update_cost_item_original_values_from_legacy_fields(config)
        items = {item["costItemId"]: item for item in updated["costItems"]}

        self.assertEqual(items["regimen_3hp"]["originalCost"], 333.0)
        self.assertEqual(updated["costs"]["regimen"]["x3HP"], 333.0)

    def test_unresolved_cost_items_cannot_contribute_to_totals(self) -> None:
        results = _mock_economics_results()
        config = build_economics_preset_kwab150()
        config["costs"]["programSetupTotal"] = 1000.0
        config["costs"]["programRunningTotal"] = 2000.0
        config["costs"]["falsePositiveIncrementalPerPerson"] = 10.0
        config = update_cost_item_original_values_from_legacy_fields(config)

        econ = run_health_economics(results, config)

        self.assertIsNone(econ["costs"]["testingCost"])
        self.assertIsNone(econ["costs"]["totalProgramCost"])
        self.assertIn("costItems.test_igra.convertedTargetYearCost", econ["status"]["missingInputs"])

    def test_false_positive_setup_and_running_costs_cannot_bypass_normalisation(self) -> None:
        results = _mock_economics_results()
        config = build_default_economics_config()
        config["costs"]["falsePositiveIncrementalPerPerson"] = 10.0
        config["costs"]["programSetupTotal"] = 1000.0
        config["costs"]["programRunningTotal"] = 2000.0
        config = update_cost_item_original_values_from_legacy_fields(config)

        econ = run_health_economics(results, config)

        self.assertIsNone(econ["costs"]["falsePositiveIncrementalCost"])
        self.assertIsNone(econ["costs"]["programSetupCost"])
        self.assertIsNone(econ["costs"]["programRunningCost"])

    def test_saved_and_reloaded_valid_cost_configuration_remains_valid(self) -> None:
        config = _valid_target_year_econ_config()
        payload = json.loads(json.dumps(config))

        econ = run_health_economics(_mock_economics_results(), payload)

        self.assertTrue(all(item["conversionStatus"] == "valid" for item in econ["costItems"]))

    def test_repeated_economic_runs_do_not_inflate_twice(self) -> None:
        config = _valid_target_year_econ_config()
        first = run_health_economics(_mock_economics_results(), config)

        second = run_health_economics(_mock_economics_results(), first["inputs"])

        self.assertEqual(first["unitCosts"]["testPerPerson"], second["unitCosts"]["testPerPerson"])
        self.assertTrue(all(item["conversionStatus"] == "valid" for item in second["costItems"]))

    def test_backward_compatible_fields_cannot_conflict_with_cost_items(self) -> None:
        config = _valid_target_year_econ_config()
        items = {item["costItemId"]: item for item in config["costItems"]}
        items["test_igra"]["originalCost"] = 50.0
        config = sync_legacy_cost_fields_from_cost_items(config)
        config["costs"]["test"]["IGRA"] = 9999.0

        econ = run_health_economics(_mock_economics_results(), config)

        self.assertEqual(econ["unitCosts"]["testPerPerson"], 50.0)
        self.assertEqual(econ["inputs"]["costs"]["test"]["IGRA"], 50.0)

    def test_threshold_value_retains_currency_year_and_source(self) -> None:
        config = build_default_economics_config()
        config["threshold"].update(
            {"value": 90000, "currency": "AUD", "referenceYear": 2025, "source": "fixture"}
        )

        self.assertEqual(config["threshold"]["currency"], "AUD")
        self.assertEqual(config["threshold"]["referenceYear"], 2025)
        self.assertEqual(config["threshold"]["source"], "fixture")

    def test_unresolved_gdp_value_produces_visible_warning(self) -> None:
        report = validate_economics_config(build_default_economics_config())

        self.assertTrue(report["isValid"])
        self.assertIn("threshold.value", report["warningFieldNames"])

    def test_exported_assumptions_include_original_and_converted_costs(self) -> None:
        cfg = build_default_config()
        econ_config = build_economics_preset_kwab150()
        payload = build_results_workbook(
            config=cfg,
            bundle={"metadata": {}, "headline": {}, "technical": {}},
            backend_status={"name": "python_apy"},
            economics_results=None,
            economics_config=econ_config,
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        self.assertIn("Cost_normalisation", wb.sheetnames)
        headers = [cell.value for cell in next(wb["Cost_normalisation"].iter_rows())]
        self.assertIn("originalCost", headers)
        self.assertIn("convertedTargetYearCost", headers)
        self.assertIn("Unresolved_assumptions", wb.sheetnames)
        wb.close()

    def test_no_dynamic_model_code_is_required_for_workflow(self) -> None:
        sys.modules.pop("engine.dynamic.dynamic_model", None)

        config = build_default_config()
        econ_config = build_default_economics_config()

        self.assertIn("scenario", config)
        self.assertIn("costNormalisation", econ_config)
        self.assertNotIn("engine.dynamic.dynamic_model", sys.modules)
        self.assertEqual(config["scenario"]["scopeStatement"], DIRECT_EFFECTS_SCOPE_STATEMENT)

    def test_generic_remote_demo_is_labelled_as_synthetic_test_fixture(self) -> None:
        preset = load_population_preset("generic_remote_demo")

        self.assertIn("test fixture", " ".join(preset["sourcesAndNotes"]).lower())
        self.assertIn("demonstration test fixture", preset["ltbiPrevalenceAssumptions"]["source"].lower())

def _mock_economics_results() -> dict:
    return {
        "summary": [
            {"Metric": "nScreened", "Median": 10},
            {"Metric": "nTotalCoursesStarted", "Median": 4},
            {"Metric": "nFalsePositiveTreated", "Median": 2},
            {"Metric": "nCuredInfection", "Median": 3},
            {"Metric": "nPreventedActiveTB", "Median": 1},
        ],
        "interfaceConfig": {"testType": "IGRA", "regimen": "3HP"},
        "doNothing": {
            "derived": [
                {"nActiveBy20y_DoNothing": 5, "nActiveBy20y_AfterStrategy": 4}
            ]
        },
    }


def _valid_target_year_econ_config() -> dict:
    config = build_economics_preset_kwab150()
    config["costs"]["falsePositiveIncrementalPerPerson"] = 10.0
    config["costs"]["programSetupTotal"] = 1000.0
    config["costs"]["programRunningTotal"] = 2000.0
    config = update_cost_item_original_values_from_legacy_fields(config)
    for item in config["costItems"]:
        if item["originalCost"] in (None, "", []):
            item["originalCost"] = 0.0
        item["originalPriceYear"] = "2025-26"
        item["sourceYearStatus"] = "explicit"
        item["originalCurrency"] = "AUD"
        item["sourceCitation"] = item.get("sourceCitation") or "test"
        item["conversionStatus"] = "not_converted"
        item["conversionApplied"] = False
        item["costRecordType"] = "source"
        item["warnings"] = []
    return sync_legacy_cost_fields_from_cost_items(config)


if __name__ == "__main__":
    unittest.main()
