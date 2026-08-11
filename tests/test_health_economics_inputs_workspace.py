from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import load_workbook

from app.health_economics_inputs import (
    STATUS_LABEL_TO_CODE,
    assess_current_analysis_economic_readiness,
    assumptions_csv,
    assumptions_workbook,
    apply_assumptions_to_economics_config,
    editable_assumption_rows,
    fatal_validation_rows,
    inclusion_label,
    mark_workspace_applied,
    mark_workspace_validated,
    new_workspace_state,
    ordered_editor_rows,
    parse_assumptions_csv,
    reconcile_workspace_state,
    rows_from_display_rows,
    status_label,
    update_workspace_rows,
    validate_editable_assumptions,
    workspace_source_hash,
)
from app.results_workbook import build_results_workbook
from engine.apy.config import build_default_config
from engine.apy.economics import build_default_economics_config, build_economics_preset_dale2019_aud
from engine.apy.evidence import load_apy_evidence_registry
from tests.test_apy_event_ledger_economics import _synthetic_econ
from tests.test_apy_evidence_registry import _minimal_reviewed_registry, _reviewed_epi_config


ROOT = Path(__file__).resolve().parents[1]


def _complete_daly_units(rows: list[dict]) -> list[dict]:
    for row in rows:
        if row.get("assumptionId") == "daly.active_tb_disability_weight":
            row["unit"] = "disability weight"
        elif row.get("assumptionId") == "daly.active_tb_duration":
            row["unit"] = "years"
        elif row.get("assumptionId") == "daly.tb_case_fatality_risk":
            row["unit"] = "probability"
        elif row.get("assumptionId") == "daly.yll_per_tb_death":
            row["unit"] = "years"
    return rows


class HealthEconomicsInputsWorkspaceTests(unittest.TestCase):
    def test_registry_rows_load_into_editable_assumptions_table(self) -> None:
        rows = editable_assumption_rows(economics_config=build_default_economics_config())
        ids = {row["assumptionId"] for row in rows}

        self.assertIn("cost.test_igra", ids)
        self.assertIn("daly.active_tb_disability_weight", ids)
        self.assertIn("threshold.gdp_per_capita", ids)
        cost_row = next(row for row in rows if row["assumptionId"] == "cost.test_igra")
        self.assertEqual(cost_row["costBasis"], "per_person_screened")

    def test_editing_cost_updates_working_copy_and_authoritative_cost_item(self) -> None:
        econ_config = _synthetic_econ()
        original_registry = load_apy_evidence_registry()
        rows = editable_assumption_rows(_minimal_reviewed_registry(), econ_config)
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")
        row.update(
            {
                "currentValue": "42.5",
                "originalCurrency": "AUD",
                "originalPriceYear": "2025-26",
                "targetCurrency": "AUD",
                "targetPriceYear": "2025-26",
                "sourceCitation": "Synthetic edited cost source",
                "reviewStatus": "configured_reviewed",
                "provisional": False,
            }
        )

        updated = apply_assumptions_to_economics_config(econ_config, rows, config=_reviewed_epi_config())
        item = next(item for item in updated["costItems"] if item["costItemId"] == "test_igra")

        self.assertEqual(item["originalCost"], 42.5)
        self.assertEqual(item["sourceCitation"], "Synthetic edited cost source")
        self.assertEqual(updated["costs"]["test"]["IGRA"], 42.5)
        self.assertTrue(updated["assumptionEvidenceValidation"]["epidemiologyReady"])
        self.assertEqual(load_apy_evidence_registry(), original_registry)

    def test_incomplete_cost_assumptions_block_icer(self) -> None:
        econ_config = _synthetic_econ()
        rows = editable_assumption_rows(_minimal_reviewed_registry(), econ_config)
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")
        row["sourceCitation"] = ""
        row["provisional"] = False

        report = validate_editable_assumptions(rows, econ_config, config=_reviewed_epi_config())

        self.assertFalse(report["summary"]["costReady"])
        self.assertFalse(report["summary"]["icerReady"])
        self.assertIn("cost.test_igra", report["summary"]["remainingBlockers"])

    def test_cost_complete_but_daly_incomplete_allows_cost_consequence_only(self) -> None:
        econ_config = _synthetic_econ()
        rows = editable_assumption_rows(_minimal_reviewed_registry(), econ_config)
        daly_row = next(item for item in rows if item["assumptionId"] == "daly.active_tb_disability_weight")
        daly_row["sourceCitation"] = ""
        daly_row["reviewStatus"] = "unresolved"
        daly_row["provisional"] = True

        report = validate_editable_assumptions(rows, econ_config, config=_reviewed_epi_config())

        self.assertTrue(report["summary"]["costConsequenceReady"])
        self.assertFalse(report["summary"]["dalyReady"])
        self.assertFalse(report["summary"]["icerReady"])

    def test_reviewed_exclusions_require_rationale(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "daly.tpt_health_loss")
        row["inclusionStatus"] = "excluded"
        row["reviewStatus"] = "reviewed_exclusion"
        row["notes"] = ""

        report = validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertIn("daly.tpt_health_loss", report["rowMessages"])
        self.assertFalse(report["summary"]["dalyReady"])

    def test_typed_numerical_value_without_source_does_not_become_ready(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "daly.active_tb_duration")
        row["currentValue"] = "0.5"
        row["sourceCitation"] = ""
        row["reviewStatus"] = "configured_reviewed"
        row["provisional"] = False

        report = validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertFalse(report["summary"]["dalyReady"])
        self.assertIn("DALY source citation is missing.", report["rowMessages"]["daly.active_tb_duration"])

    def test_exported_assumptions_preserve_review_and_provisional_status(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")
        row["reviewStatus"] = "configured_reviewed"
        row["provisional"] = False

        csv_payload = assumptions_csv(rows)
        workbook_payload = assumptions_workbook(rows, validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config()))
        wb = load_workbook(BytesIO(workbook_payload), read_only=True)
        ws = wb["Edited_assumptions"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        status_idx = headers.index("reviewStatus")
        provisional_idx = headers.index("provisional")
        exported = {
            row_values[0].value: (row_values[status_idx].value, row_values[provisional_idx].value)
            for row_values in ws.iter_rows(min_row=2)
        }
        wb.close()

        self.assertIn("configured_reviewed", csv_payload)
        self.assertEqual(exported["cost.test_igra"], ("configured_reviewed", "false"))

    def test_uploaded_assumptions_validate_before_application(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")
        row["costBasis"] = "wrong_basis"

        uploaded = parse_assumptions_csv(assumptions_csv(rows))
        report = validate_editable_assumptions(uploaded, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertFalse(report["isValidForApplication"])
        self.assertIn("cost.test_igra", report["rowMessages"])

    def test_missing_economic_value_is_exported_blank_not_zero(self) -> None:
        payload = build_results_workbook(
            config=build_default_config(),
            bundle={"metadata": {}, "technical": {}, "headline": {}, "downloads": {}},
            economics_config=build_default_economics_config(),
        )
        wb = load_workbook(BytesIO(payload), read_only=True)
        ws = wb["Edited_assumptions"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        self.assertEqual(rows[0][0], "Status")
        self.assertEqual(rows[1][0], "No edited assumptions applied")
        self.assertNotEqual(rows[1][0], 0)

    def test_workbook_includes_original_and_applied_assumption_sheets(self) -> None:
        econ_config = apply_assumptions_to_economics_config(
            _synthetic_econ(),
            editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ()),
            config=_reviewed_epi_config(),
        )

        payload = build_results_workbook(
            config=_reviewed_epi_config(),
            bundle={"metadata": {}, "technical": {}, "headline": {}, "downloads": {}},
            economics_config=econ_config,
        )
        wb = load_workbook(BytesIO(payload), read_only=True)
        names = set(wb.sheetnames)
        wb.close()

        self.assertIn("Edited_assumptions", names)
        self.assertIn("Applied_session_assumptions", names)
        self.assertIn("Original_evidence_registry", names)
        self.assertIn("Assumption_validation", names)

    def test_preset_change_without_edits_reloads_workspace_and_clears_validation(self) -> None:
        state = new_workspace_state(build_default_economics_config())
        state = mark_workspace_validated(
            state,
            validate_editable_assumptions(state["rows"], build_default_economics_config(), config=_reviewed_epi_config()),
        )
        new_config = _synthetic_econ()

        updated = reconcile_workspace_state(state, new_config)

        self.assertEqual(updated["sourceHash"], workspace_source_hash(new_config))
        self.assertFalse(updated["hasUnsavedEdits"])
        self.assertFalse(updated["validated"])
        self.assertIsNone(updated["validation"])

    def test_preset_change_with_unsaved_edits_does_not_silently_overwrite(self) -> None:
        state = new_workspace_state(build_default_economics_config())
        rows = list(state["rows"])
        rows[0] = {**rows[0], "notes": "unsaved local edit"}
        state = update_workspace_rows(state, rows)
        state = mark_workspace_validated(
            state,
            validate_editable_assumptions(state["rows"], build_default_economics_config(), config=_reviewed_epi_config()),
        )

        updated = reconcile_workspace_state(state, _synthetic_econ())

        self.assertTrue(updated["hasUnsavedEdits"])
        self.assertTrue(updated["presetConflict"])
        self.assertEqual(updated["rows"][0]["notes"], "unsaved local edit")
        self.assertFalse(updated["validated"])
        self.assertIsNone(updated["validation"])

    def test_reset_and_discard_reload_from_current_analysis(self) -> None:
        state = new_workspace_state(build_default_economics_config())
        rows = list(state["rows"])
        rows[0] = {**rows[0], "notes": "unsaved local edit"}
        state = update_workspace_rows(state, rows)

        reset = reconcile_workspace_state(state, _synthetic_econ(), action="reset")
        discarded = reconcile_workspace_state(state, _synthetic_econ(), action="discard")

        self.assertFalse(reset["hasUnsavedEdits"])
        self.assertFalse(discarded["hasUnsavedEdits"])
        self.assertNotEqual(reset["rows"][0]["notes"], "unsaved local edit")
        self.assertEqual(reset["sourceHash"], workspace_source_hash(_synthetic_econ()))

    def test_status_and_inclusion_labels_map_to_internal_values(self) -> None:
        row = {
            "assumptionId": "daly.active_tb_duration",
            "reviewStatusLabel": "Reviewed numerical assumption",
            "inclusionStatusLabel": "Bundled into another item",
        }

        mapped = rows_from_display_rows([row])[0]

        self.assertEqual(STATUS_LABEL_TO_CODE["Reviewed numerical assumption"], "configured_reviewed")
        self.assertEqual(mapped["reviewStatus"], "configured_reviewed")
        self.assertEqual(mapped["inclusionStatus"], "bundled")
        self.assertEqual(status_label("model_derived_reviewed"), "Reviewed model-derived assumption")
        self.assertEqual(inclusion_label("excluded"), "Excluded")

    def test_conversion_status_and_converted_cost_are_displayed(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")

        self.assertEqual(row["conversionStatus"], "valid")
        self.assertEqual(row["convertedTargetYearCost"], 10)
        self.assertEqual(row["inflationFactor"], 1.0)

    def test_unresolved_inflation_remains_unresolved_in_workspace(self) -> None:
        rows = editable_assumption_rows(economics_config=build_default_economics_config())
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")

        self.assertNotEqual(row["conversionStatus"], "valid")
        self.assertIn("Original cost is missing", row["conversionWarnings"])

    def test_dale_2019_workspace_uses_embedded_registry_and_is_current_pathway_ready(self) -> None:
        econ_config = build_economics_preset_dale2019_aud("3HP")
        rows = editable_assumption_rows(economics_config=econ_config)
        row_by_id = {row["assumptionId"]: row for row in rows}

        self.assertEqual(row_by_id["cost.test_igra"]["currentValue"], 113.48)
        self.assertEqual(row_by_id["cost.test_igra"]["targetPriceYear"], "2019")
        self.assertEqual(row_by_id["cost.test_igra"]["convertedTargetYearCost"], 113.48)
        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "3HP"},
            econ_config,
            {},
            rows,
        )

        self.assertTrue(readiness["currentAnalysisCostReady"])
        self.assertTrue(readiness["currentAnalysisDALYReady"])
        self.assertTrue(readiness["currentAnalysisICERReady"])
        self.assertFalse(readiness["currentAnalysisNMBReady"])
        self.assertFalse(readiness["overallReferenceEvidenceReady"])

    def test_dale_2019_selected_pathway_not_blocked_by_unresolved_alternatives(self) -> None:
        econ_config = build_economics_preset_dale2019_aud("3HP")
        rows = editable_assumption_rows(economics_config=econ_config)
        for assumption_id in ["cost.test_tst", "cost.regimen_4r", "cost.regimen_3hr", "cost.regimen_6h", "cost.regimen_9h"]:
            row = next(item for item in rows if item["assumptionId"] == assumption_id)
            row["sourceCitation"] = ""
            row["reviewStatus"] = "unresolved"
            row["reviewStatusLabel"] = "Unresolved"
            row["provisional"] = True

        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "3HP"},
            econ_config,
            {},
            rows,
        )

        self.assertTrue(readiness["currentAnalysisICERReady"])
        self.assertFalse(readiness["fullStrategyLibraryReady"])
        self.assertNotIn("cost.test_tst", {item["assumptionId"] for item in readiness["currentBlockers"]})

    def test_dale_2019_selecting_tst_or_alternate_regimen_changes_applicable_cost(self) -> None:
        econ_config = build_economics_preset_dale2019_aud("3HP")
        rows = editable_assumption_rows(economics_config=econ_config)
        row_by_id = {row["assumptionId"]: row for row in rows}
        row_by_id["cost.test_tst"]["sourceCitation"] = ""
        row_by_id["cost.test_tst"]["reviewStatus"] = "unresolved"
        row_by_id["cost.test_tst"]["provisional"] = True
        row_by_id["cost.regimen_4r"]["sourceCitation"] = ""
        row_by_id["cost.regimen_4r"]["reviewStatus"] = "unresolved"
        row_by_id["cost.regimen_4r"]["provisional"] = True

        tst_readiness = assess_current_analysis_economic_readiness({"testType": "TST", "regimen": "3HP"}, econ_config, {}, rows)
        four_r_readiness = assess_current_analysis_economic_readiness({"testType": "IGRA", "regimen": "4R"}, econ_config, {}, rows)

        self.assertIn("cost.test_tst", {item["assumptionId"] for item in tst_readiness["currentBlockers"]})
        self.assertIn("cost.regimen_4r", {item["assumptionId"] for item in four_r_readiness["currentBlockers"]})

    def test_new_workspace_state_prefers_dale_embedded_registry(self) -> None:
        state = new_workspace_state(build_economics_preset_dale2019_aud("3HP"))
        row = next(item for item in state["rows"] if item["assumptionId"] == "cost.test_igra")

        self.assertEqual(row["originalPriceYear"], "2019")
        self.assertEqual(row["targetPriceYear"], "2019")
        self.assertEqual(row["reviewStatus"], "configured_reviewed")

    def test_apply_preserves_reviewed_bundled_false_positive_cost_as_zero_item(self) -> None:
        econ_config = build_economics_preset_dale2019_aud("3HP")
        rows = editable_assumption_rows(economics_config=econ_config)

        updated = apply_assumptions_to_economics_config(econ_config, rows, config=_reviewed_epi_config())
        by_id = {item["costItemId"]: item for item in updated["costItems"]}

        self.assertEqual(by_id["false_positive_incremental"]["originalCost"], 0.0)
        self.assertEqual(by_id["program_setup"]["originalCost"], 0.0)
        self.assertEqual(by_id["program_running"]["originalCost"], 0.0)

    def test_legacy_controls_cannot_override_applied_workspace_assumptions(self) -> None:
        page = (ROOT / "pages" / "4_Economics.py").read_text(encoding="utf-8")

        self.assertIn('st.expander("Legacy/developer economic controls"', page)
        self.assertNotIn('st.subheader("Edit Economics Inputs")', page)
        self.assertIn('disabled=workspace_applied', page)
        self.assertIn("Workspace assumptions have been applied. Legacy controls are disabled", page)

    def test_standard_page_has_one_authoritative_editing_route(self) -> None:
        page = (ROOT / "pages" / "4_Economics.py").read_text(encoding="utf-8")

        self.assertIn('st.subheader("Inputs required for this analysis")', page)
        self.assertEqual(page.count("Validate assumptions"), 1)
        self.assertEqual(page.count("Apply assumptions to current analysis"), 1)
        self.assertIn("Legacy/developer economic controls", page)

    def test_validate_rerun_retains_validation_when_rows_unchanged(self) -> None:
        econ_config = _synthetic_econ()
        state = new_workspace_state(econ_config, _minimal_reviewed_registry())
        report = validate_editable_assumptions(state["rows"], econ_config, config=_reviewed_epi_config())
        state = mark_workspace_validated(state, report)

        updated = update_workspace_rows(state, deepcopy(state["rows"]))

        self.assertTrue(updated["validated"])
        self.assertIsNotNone(updated["validation"])
        self.assertEqual(updated["validation"]["summary"], report["summary"])

    def test_genuine_edit_clears_stale_validation(self) -> None:
        econ_config = _synthetic_econ()
        state = new_workspace_state(econ_config, _minimal_reviewed_registry())
        state = mark_workspace_validated(
            state,
            validate_editable_assumptions(state["rows"], econ_config, config=_reviewed_epi_config()),
        )
        rows = deepcopy(state["rows"])
        rows[0]["notes"] = "changed by user"

        updated = update_workspace_rows(state, rows)

        self.assertFalse(updated["validated"])
        self.assertIsNone(updated["validation"])

    def test_validation_and_conversion_fields_do_not_clear_validation(self) -> None:
        econ_config = _synthetic_econ()
        state = new_workspace_state(econ_config, _minimal_reviewed_registry())
        state = mark_workspace_validated(
            state,
            validate_editable_assumptions(state["rows"], econ_config, config=_reviewed_epi_config()),
        )
        rows = deepcopy(state["rows"])
        rows[0]["validationMessage"] = "rerendered validation text"
        rows[0]["convertedTargetYearCost"] = 999999

        updated = update_workspace_rows(state, rows)

        self.assertTrue(updated["validated"])
        self.assertIsNotNone(updated["validation"])

    def test_incomplete_but_structurally_safe_assumptions_can_be_applied(self) -> None:
        econ_config = _synthetic_econ()
        rows = editable_assumption_rows(_minimal_reviewed_registry(), econ_config)
        row = next(item for item in rows if item["assumptionId"] == "daly.active_tb_duration")
        row["sourceCitation"] = ""

        report = validate_editable_assumptions(rows, econ_config, config=_reviewed_epi_config())
        updated = apply_assumptions_to_economics_config(econ_config, rows, config=_reviewed_epi_config())

        self.assertTrue(report["isValidForApplication"])
        self.assertFalse(report["summary"]["icerReady"])
        self.assertIn("assumptionEvidenceValidation", updated)

    def test_negative_cost_prevents_application(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "cost.test_igra")
        row["currentValue"] = "-1"

        report = validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertFalse(report["isValidForApplication"])
        self.assertIn("cost.test_igra", {item["assumptionId"] for item in fatal_validation_rows(report)})

    def test_probability_above_one_prevents_application(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "daly.active_tb_disability_weight")
        row["currentValue"] = "1.2"

        report = validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertFalse(report["isValidForApplication"])

    def test_target_currency_year_mismatch_prevents_application(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "threshold.gdp_per_capita")
        row["targetPriceYear"] = "2024-25"

        report = validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertFalse(report["isValidForApplication"])

    def test_invalid_bundling_prevents_application(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        row = next(item for item in rows if item["assumptionId"] == "cost.tpt_adr_management")
        row["inclusionStatus"] = "bundled"
        row["bundledIntoAssumptionId"] = "cost.not_real"

        report = validate_editable_assumptions(rows, _synthetic_econ(), config=_reviewed_epi_config())

        self.assertFalse(report["isValidForApplication"])

    def test_igra_3hp_icer_not_blocked_by_unresolved_alternative_costs(self) -> None:
        registry = _complete_daly_units(_minimal_reviewed_registry())
        for assumption_id in ["cost.test_tst", "cost.regimen_4r", "cost.regimen_3hr", "cost.regimen_6h", "cost.regimen_9h"]:
            row = next(item for item in registry if item["assumptionId"] == assumption_id)
            row["sourceCitation"] = ""
            row["reviewStatus"] = "unresolved"
            row["provisional"] = True

        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "3HP"},
            _synthetic_econ({"threshold": 100000}),
            {},
            registry,
        )

        self.assertTrue(readiness["currentAnalysisICERReady"])
        self.assertFalse(readiness["fullStrategyLibraryReady"])
        self.assertNotIn("cost.test_tst", {item["assumptionId"] for item in readiness["currentBlockers"]})
        self.assertIn("cost.test_tst", {item["assumptionId"] for item in readiness["alternativeStrategyBlockers"]})

    def test_selecting_tst_makes_tst_cost_applicable(self) -> None:
        registry = _complete_daly_units(_minimal_reviewed_registry())
        row = next(item for item in registry if item["assumptionId"] == "cost.test_tst")
        row["sourceCitation"] = ""
        row["reviewStatus"] = "unresolved"
        row["provisional"] = True

        readiness = assess_current_analysis_economic_readiness(
            {"testType": "TST", "regimen": "3HP"},
            _synthetic_econ({"threshold": 100000}),
            {},
            registry,
        )

        self.assertFalse(readiness["currentAnalysisCostReady"])
        self.assertIn("cost.test_tst", {item["assumptionId"] for item in readiness["currentBlockers"]})

    def test_selecting_another_regimen_makes_only_that_regimen_applicable(self) -> None:
        registry = _complete_daly_units(_minimal_reviewed_registry())
        for assumption_id in ["cost.regimen_3hp", "cost.regimen_4r"]:
            row = next(item for item in registry if item["assumptionId"] == assumption_id)
            row["sourceCitation"] = ""
            row["reviewStatus"] = "unresolved"
            row["provisional"] = True

        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "4R"},
            _synthetic_econ({"threshold": 100000}),
            {},
            registry,
        )
        current_ids = {item["assumptionId"] for item in readiness["currentBlockers"]}

        self.assertIn("cost.regimen_4r", current_ids)
        self.assertNotIn("cost.regimen_3hp", current_ids)

    def test_missing_daly_blocks_icer_but_missing_threshold_blocks_only_nmb(self) -> None:
        registry = _complete_daly_units(_minimal_reviewed_registry())
        daly_row = next(item for item in registry if item["assumptionId"] == "daly.active_tb_duration")
        daly_row["sourceCitation"] = ""
        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "3HP"},
            _synthetic_econ({"threshold": 100000}),
            {},
            registry,
        )
        self.assertFalse(readiness["currentAnalysisICERReady"])

        registry = _complete_daly_units(_minimal_reviewed_registry())
        threshold_row = next(item for item in registry if item["assumptionId"] == "threshold.gdp_per_capita")
        threshold_row["sourceCitation"] = ""
        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "3HP"},
            _synthetic_econ(),
            {},
            registry,
        )
        self.assertTrue(readiness["currentAnalysisICERReady"])
        self.assertFalse(readiness["currentAnalysisNMBReady"])

    def test_standard_editor_orders_editable_fields_before_technical_fields(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        ordered = ordered_editor_rows(rows[:1], advanced=True)[0]
        keys = list(ordered)

        self.assertLess(keys.index("description"), keys.index("assumptionId"))
        self.assertLess(keys.index("currentValue"), keys.index("doubleCountingGroup"))

    def test_standard_editor_keeps_hidden_row_key_for_safe_mapping(self) -> None:
        rows = editable_assumption_rows(_minimal_reviewed_registry(), _synthetic_econ())
        ordered = ordered_editor_rows(rows[:1], advanced=False)[0]
        keys = list(ordered)

        self.assertIn("assumptionId", ordered)
        self.assertLess(keys.index("description"), keys.index("assumptionId"))
        self.assertNotIn("doubleCountingGroup", ordered)

    def test_every_current_blocker_maps_to_editable_row_and_is_not_truncated(self) -> None:
        registry = _minimal_reviewed_registry()
        for row in registry:
            if row["assumptionId"].startswith("daly."):
                row["sourceCitation"] = ""
                row["reviewStatus"] = "unresolved"
                row["provisional"] = True

        readiness = assess_current_analysis_economic_readiness(
            {"testType": "IGRA", "regimen": "3HP"},
            _synthetic_econ({"threshold": 100000}),
            {},
            registry,
        )
        editable_ids = {row["assumptionId"] for row in editable_assumption_rows(registry, _synthetic_econ())}

        self.assertEqual(len(readiness["currentBlockers"]), len([item for item in readiness["currentBlockers"]]))
        self.assertGreater(len(readiness["currentBlockers"]), 3)
        self.assertTrue(all(blocker["editableRowId"] in editable_ids for blocker in readiness["currentBlockers"]))

    def test_converted_costs_are_read_only_and_derived_in_page(self) -> None:
        page = (ROOT / "pages" / "4_Economics.py").read_text(encoding="utf-8")

        self.assertIn("*DERIVED_CONVERSION_COLUMNS", page)
        self.assertIn('"convertedTargetYearCost"', page)
        self.assertIn("Price-year conversion audit", page)


if __name__ == "__main__":
    unittest.main()
