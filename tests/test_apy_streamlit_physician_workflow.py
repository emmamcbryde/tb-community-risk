from __future__ import annotations

from io import BytesIO
import json
import os
import sys
import types
import unittest
from pathlib import Path

from openpyxl import load_workbook

from adapters.paths import repo_root
from adapters.python_apy_backend import PythonApyBackend
from app.epidemiology_inputs import (
    apply_epidemiology_updates,
    apply_ltbi_state_development_compatibility,
    apply_ltbi_state_assumption_update,
    fraction_to_percent,
    ltbi_state_display_rows,
    percent_to_fraction,
    risk_override_from_percentages,
)
from app.results_workbook import build_results_workbook
from app.run_analysis_controls import (
    TECHNICAL_DEMONSTRATION_ROUTE,
    prepare_run_config_for_recent_ltbi_route,
    recent_ltbi_decision_required,
    technical_demonstration_summary,
)
from engine.apy.calibration import calibrate_from_config
from engine.apy.ltbi_state import resolve_ltbi_state_assumptions
from engine.apy.config import build_default_config
from engine.apy.runner import run_scenario_with_do_nothing


class PhysicianWorkflowHelperTests(unittest.TestCase):
    def test_percentage_fraction_conversion(self) -> None:
        self.assertAlmostEqual(percent_to_fraction(7.5), 0.075)
        self.assertAlmostEqual(fraction_to_percent(0.0125), 1.25)
        with self.assertRaises(ValueError):
            percent_to_fraction(0)
        with self.assertRaises(ValueError):
            percent_to_fraction(100)

    def test_default_and_custom_prevalence_updates(self) -> None:
        cfg = build_default_config()
        custom = apply_epidemiology_updates(
            cfg,
            use_default_ltbi_prevalence=False,
            ltbi_prevalence_percent=1.0,
            use_default_active_tb_prevalence=False,
            active_tb_prevalence_percent=0.2,
        )
        self.assertEqual(custom["ltbiPrevalence"], 0.01)
        self.assertEqual(custom["activeTBPrevalence"], 0.002)
        restored = apply_epidemiology_updates(
            custom,
            use_default_ltbi_prevalence=True,
            ltbi_prevalence_percent=9.0,
            use_default_active_tb_prevalence=True,
            active_tb_prevalence_percent=1.0,
        )
        self.assertIsNone(restored["ltbiPrevalence"])
        self.assertIsNone(restored["activeTBPrevalence"])

    def test_risk_factor_scalar_and_age_group_inputs(self) -> None:
        self.assertEqual(risk_override_from_percentages("Single overall", [20]), 0.2)
        self.assertEqual(
            risk_override_from_percentages("Three age groups", [1, 2, 3]),
            [0.01, 0.02, 0.03],
        )
        with self.assertRaises(ValueError):
            risk_override_from_percentages("Three age groups", [1, 2])

    def test_ltbi_state_ui_helper_updates_authoritative_nested_field(self) -> None:
        cfg = build_default_config()
        updated = apply_ltbi_state_assumption_update(
            cfg,
            baseline_recent_percent=33.0,
            transition_rate_per_year=0.2,
            source="test UI source",
            status="configured",
            notes="test UI notes",
        )
        rows = ltbi_state_display_rows(updated)

        self.assertEqual(updated["ltbiStateAssumptions"]["baselineRecentLTBIProportion"], 0.33)
        self.assertEqual(updated["baselineRecentLTBIProportion"], 0.33)
        self.assertEqual(updated["ltbiStateAssumptions"]["source"], "test UI source")
        self.assertIn(
            {"Assumption": "Baseline recent-LTBI proportion", "Value": 0.33},
            rows,
        )

    def test_ltbi_state_ui_rows_show_development_compatibility_flags(self) -> None:
        rows = ltbi_state_display_rows(
            apply_ltbi_state_development_compatibility(build_default_config())
        )

        self.assertIn(
            {"Assumption": "Development compatibility mode", "Value": True},
            rows,
        )
        self.assertIn({"Assumption": "Provisional result", "Value": True}, rows)

    def test_unresolved_recent_ltbi_blocks_normal_reference_analysis(self) -> None:
        cfg = build_default_config()

        self.assertTrue(recent_ltbi_decision_required(cfg))
        with self.assertRaisesRegex(ValueError, "Recent versus remote LTBI is unresolved"):
            prepare_run_config_for_recent_ltbi_route(cfg, selected_route=None)

    def test_technical_demonstration_explicitly_enables_compatibility_mode(self) -> None:
        cfg = build_default_config()
        before = resolve_ltbi_state_assumptions(cfg)
        run_cfg = prepare_run_config_for_recent_ltbi_route(
            cfg,
            selected_route=TECHNICAL_DEMONSTRATION_ROUTE,
        )
        after = resolve_ltbi_state_assumptions(run_cfg)

        self.assertFalse(before["developmentCompatibilityMode"])
        self.assertTrue(after["developmentCompatibilityMode"])
        self.assertEqual(after["baselineRecentLTBIProportion"], 0.0)
        self.assertEqual(after["status"], "unresolved_development_compatibility")
        self.assertEqual(
            after["baselineRecentLTBIProportionStatus"],
            "unresolved",
        )
        self.assertTrue(after["provisional"])

    def test_technical_demonstration_summary_preserves_unreviewed_status(self) -> None:
        summary = technical_demonstration_summary(build_default_config())

        self.assertTrue(summary["developmentCompatibilityMode"])
        self.assertTrue(summary["provisional"])
        self.assertEqual(summary["baselineRecentLTBIProportion"], 0.0)
        self.assertEqual(summary["status"], "unresolved_development_compatibility")
        self.assertNotIn("configured_reviewed", summary.values())

    def test_reviewed_recent_ltbi_assumption_allows_ordinary_run_config(self) -> None:
        cfg = apply_ltbi_state_assumption_update(
            build_default_config(),
            baseline_recent_percent=20.0,
            transition_rate_per_year=0.2,
            source="Reviewed unit-test recent LTBI source",
            status="configured_reviewed",
            notes="Reviewed for interface guardrail test.",
        )

        self.assertFalse(recent_ltbi_decision_required(cfg))
        run_cfg = prepare_run_config_for_recent_ltbi_route(cfg, selected_route=None)
        state = resolve_ltbi_state_assumptions(run_cfg)
        self.assertFalse(state["developmentCompatibilityMode"])
        self.assertFalse(state["provisional"])
        self.assertEqual(state["baselineRecentLTBIProportion"], 0.2)

    def test_run_page_does_not_show_raw_model_failure_for_unresolved_ltbi(self) -> None:
        page_text = (repo_root() / "pages" / "2_Run_Model.py").read_text(encoding="utf-8")

        self.assertIn("Recent versus remote LTBI assumption", page_text)
        self.assertIn("Run a technical demonstration", page_text)
        self.assertIn("Review or enter the assumption", page_text)
        self.assertNotIn("Model run failed", page_text)
        self.assertNotIn("Run provisional development analysis using the 0% compatibility placeholder", page_text)

    def test_new_analysis_pages_clear_prior_technical_demonstration_route(self) -> None:
        start_text = (repo_root() / "pages" / "0_Start.py").read_text(encoding="utf-8")
        scenario_text = (repo_root() / "pages" / "1_Scenario.py").read_text(encoding="utf-8")

        self.assertGreaterEqual(start_text.count('pop("recent_ltbi_run_route", None)'), 2)
        self.assertGreaterEqual(scenario_text.count('pop("recent_ltbi_run_route", None)'), 4)


class PythonApyPrevalencePathTests(unittest.TestCase):
    def test_custom_ltbi_prevalence_reaches_calibration(self) -> None:
        cfg = apply_ltbi_state_assumption_update(
            build_default_config(),
            baseline_recent_percent=25.0,
            transition_rate_per_year=0.2,
            source="Reviewed unit-test LTBI-state fixture",
            status="configured",
            notes="",
        )
        cfg["ltbiPrevalence"] = 0.01
        cfg["activeTBPrevalence"] = (10 / 770) * 0.01 / (47 / 624)
        calibration = calibrate_from_config(cfg)
        self.assertAlmostEqual(calibration["targetInfPrev"], 0.01)
        self.assertAlmostEqual(calibration["expectedInfPrev"], 0.01, places=4)

    def test_low_prevalence_changes_python_yield_without_matlab(self) -> None:
        self.assertNotIn("matlab.engine", sys.modules)
        default_cfg = apply_ltbi_state_assumption_update(
            build_default_config(),
            baseline_recent_percent=25.0,
            transition_rate_per_year=0.2,
            source="Reviewed unit-test LTBI-state fixture",
            status="configured",
            notes="",
        )
        default_cfg.update({"N": 300, "nReps": 20, "seed": 4, "screenCoverage": 1.0})
        low_cfg = dict(default_cfg)
        low_cfg["ltbiPrevalence"] = 0.01
        low_cfg["activeTBPrevalence"] = (10 / 770) * 0.01 / (47 / 624)
        default_out = run_scenario_with_do_nothing(default_cfg)
        low_out = run_scenario_with_do_nothing(low_cfg)
        default_inf = _median(default_out, "nInfected")
        low_inf = _median(low_out, "nInfected")
        default_positive = _median(default_out, "nTestPositiveNonActive")
        low_positive = _median(low_out, "nTestPositiveNonActive")
        self.assertGreater(default_inf, low_inf)
        self.assertGreater(default_positive, low_positive)
        self.assertNotIn("matlab.engine", sys.modules)

    def test_python_backend_save_load_round_trip_preserves_custom_prevalence(self) -> None:
        backend = PythonApyBackend(repo_root())
        cfg = backend.default_config()
        cfg["ltbiPrevalence"] = 0.02
        cfg["activeTBPrevalence"] = 0.003
        cfg["riskPrev"]["contact"] = [0.01, 0.02, 0.03]
        tmp = repo_root() / ".tmp_test_scenario_roundtrip"
        tmp.mkdir(exist_ok=True)
        path = tmp / "scenario.json"
        try:
            backend.save_scenario(cfg, str(path))
            loaded, _, info = backend.load_scenario(str(path))
        finally:
            if path.exists():
                path.unlink()
            if tmp.exists():
                tmp.rmdir()
        self.assertEqual(loaded["ltbiPrevalence"], 0.02)
        self.assertEqual(loaded["activeTBPrevalence"], 0.003)
        self.assertEqual(loaded["riskPrev"]["contact"], [0.01, 0.02, 0.03])
        self.assertEqual(info["backend"], "python_apy")

    def test_python_backend_save_load_round_trip_preserves_ltbi_state(self) -> None:
        backend = PythonApyBackend(repo_root())
        cfg = apply_ltbi_state_assumption_update(
            backend.default_config(),
            baseline_recent_percent=22.0,
            transition_rate_per_year=0.25,
            source="saved source",
            status="configured",
            notes="saved notes",
        )
        tmp = repo_root() / ".tmp_test_ltbi_state_roundtrip"
        tmp.mkdir(exist_ok=True)
        path = tmp / "scenario.json"
        try:
            backend.save_scenario(cfg, str(path))
            loaded, report, _ = backend.load_scenario(str(path))
        finally:
            if path.exists():
                path.unlink()
            if tmp.exists():
                tmp.rmdir()

        self.assertTrue(report["isValid"])
        self.assertEqual(loaded["ltbiStateAssumptions"]["baselineRecentLTBIProportion"], 0.22)
        self.assertEqual(loaded["ltbiStateAssumptions"]["recentToRemoteTransitionRatePerYear"], 0.25)
        self.assertEqual(loaded["ltbiStateAssumptions"]["source"], "saved source")
        self.assertEqual(loaded["ltbiStateAssumptions"]["status"], "configured")


class WorkbookExportTests(unittest.TestCase):
    def test_workbook_contains_required_sheets_and_values(self) -> None:
        cfg = apply_ltbi_state_development_compatibility(build_default_config())
        cfg.update({"N": 80, "nReps": 3, "seed": 3, "ltbiPrevalence": 0.02})
        bundle = PythonApyBackend(repo_root()).run_scenario_bundle(cfg)
        payload = build_results_workbook(
            config=cfg,
            bundle=bundle,
            backend_status={"name": "python_apy", "experimental": True},
            economics_results=None,
            economics_config=None,
            results_stale=False,
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        expected = {
            "Read_me",
            "Scenario_inputs",
            "Risk_factor_inputs",
            "Headline_results",
            "Summary_results",
            "Natural_history",
            "Technical_metadata",
            "Economics",
        }
        self.assertTrue(expected.issubset(set(wb.sheetnames)))
        scenario_values = list(wb["Scenario_inputs"].iter_rows(values_only=True))
        self.assertIn(("LTBI prevalence source", "Custom", None), scenario_values)
        self.assertIn(("LTBI prevalence input", 0.02, "percentage"), scenario_values)
        self.assertIn(("LTBI-state model", "continuous_markov_recent_remote", None), scenario_values)
        self.assertIn(("recent-state implied mean residence time", 5, "years"), scenario_values)
        self.assertIn(("baseline recent LTBI proportion source", None, None), scenario_values)
        self.assertIn(("baseline recent LTBI proportion status", "unresolved", None), scenario_values)
        self.assertIn(("LTBI transition model status", "configured_reviewed", None), scenario_values)
        self.assertIn(("LTBI state development compatibility mode", True, None), scenario_values)
        self.assertIn(("LTBI state provisional result", True, None), scenario_values)
        economics_values = list(wb["Economics"].iter_rows(values_only=True))
        self.assertIn(("Economics not run", None, "No zero values have been substituted for missing economics outputs."), economics_values)
        wb.close()

    def test_technical_demonstration_workbook_retains_provisional_flag(self) -> None:
        cfg = prepare_run_config_for_recent_ltbi_route(
            build_default_config(),
            selected_route=TECHNICAL_DEMONSTRATION_ROUTE,
        )
        cfg.update({"N": 40, "nReps": 2, "seed": 11, "ltbiPrevalence": 0.02})
        bundle = PythonApyBackend(repo_root()).run_scenario_bundle(cfg)
        payload = build_results_workbook(
            config=cfg,
            bundle=bundle,
            backend_status={"name": "python_apy", "experimental": True},
            economics_results=None,
            economics_config=None,
            results_stale=False,
        )

        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        scenario_values = list(wb["Scenario_inputs"].iter_rows(values_only=True))
        technical_values = list(wb["Technical_metadata"].iter_rows(values_only=True))
        self.assertIn(("LTBI state development compatibility mode", True, None), scenario_values)
        self.assertIn(("LTBI state provisional result", True, None), scenario_values)
        self.assertIn(("ltbiState.provisional", True), technical_values)
        wb.close()

    def test_stale_state_function_marks_results_and_economics(self) -> None:
        fake_streamlit = types.SimpleNamespace(
            session_state={
                "results_bundle": {"metadata": {}},
                "economics_results": {"summaryRows": []},
                "compare_baseline_bundle": {"metadata": {}},
            },
            cache_resource=lambda **_: (lambda fn: fn),
        )
        old_streamlit = sys.modules.get("streamlit")
        sys.modules["streamlit"] = fake_streamlit
        try:
            import importlib
            import app.state as state

            importlib.reload(state)
            state.mark_config_changed()
            self.assertTrue(fake_streamlit.session_state["dirty_config"])
            self.assertTrue(fake_streamlit.session_state["results_stale"])
            self.assertTrue(fake_streamlit.session_state["dirty_economics"])
            self.assertTrue(fake_streamlit.session_state["compare_results_stale"])
        finally:
            if old_streamlit is not None:
                sys.modules["streamlit"] = old_streamlit
            else:
                sys.modules.pop("streamlit", None)
            import importlib
            import app.state as state

            importlib.reload(state)


class HostedDeploymentSafetyTests(unittest.TestCase):
    def test_new_session_defaults_to_python_and_blocks_matlab_without_opt_in(self) -> None:
        fake_streamlit = types.SimpleNamespace(
            session_state={},
            cache_resource=lambda **_: (lambda fn: fn),
        )
        old_streamlit = sys.modules.get("streamlit")
        old_env = os.environ.pop("APY_ENABLE_MATLAB_BACKEND", None)
        sys.modules["streamlit"] = fake_streamlit
        try:
            import importlib
            import app.state as state

            importlib.reload(state)
            state.init_session_state()
            self.assertEqual(state.get_backend_name(), "python_apy")
            self.assertFalse(state.matlab_backend_enabled())
            with self.assertRaises(ValueError):
                state.set_backend_name("matlab")
        finally:
            if old_env is not None:
                os.environ["APY_ENABLE_MATLAB_BACKEND"] = old_env
            if old_streamlit is not None:
                sys.modules["streamlit"] = old_streamlit
            else:
                sys.modules.pop("streamlit", None)
            import importlib
            import app.state as state

            importlib.reload(state)

    def test_python_hosted_workflow_completes_without_matlab(self) -> None:
        sys.modules.pop("matlab.engine", None)
        backend = PythonApyBackend(repo_root())
        config = backend.default_config()
        config.update({"N": 100, "nReps": 5, "seed": 9})
        config = apply_ltbi_state_assumption_update(
            config,
            baseline_recent_percent=25.0,
            transition_rate_per_year=0.2,
            source="Reviewed unit-test LTBI-state fixture",
            status="configured",
            notes="",
        )
        config = apply_epidemiology_updates(
            config,
            use_default_ltbi_prevalence=False,
            ltbi_prevalence_percent=2.0,
            use_default_active_tb_prevalence=False,
            active_tb_prevalence_percent=(10 / 770) * 0.02 / (47 / 624) * 100,
        )
        report = backend.validate_config(config)
        self.assertTrue(report["isValid"])
        bundle = backend.run_scenario_bundle(config, validation_report=report)
        payload = build_results_workbook(
            config=config,
            bundle=bundle,
            backend_status=backend.status(),
            economics_results=None,
            economics_config=None,
            results_stale=False,
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        self.assertIn("Scenario_inputs", wb.sheetnames)
        wb.close()
        self.assertNotIn("matlab.engine", sys.modules)


def _median(model_out: dict, metric: str) -> float:
    summary = model_out["results"]["summary"]
    return float(summary.loc[summary["Metric"] == metric, "Median"].iloc[0])


if __name__ == "__main__":
    unittest.main()
