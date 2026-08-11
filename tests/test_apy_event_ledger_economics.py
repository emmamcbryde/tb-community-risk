from __future__ import annotations

from io import BytesIO
import math
import unittest

from openpyxl import load_workbook

from app.results_workbook import build_results_workbook
from engine.apy.config import build_default_config
from engine.apy.economics import build_default_economics_config, build_economics_preset_dale2019_aud, run_health_economics
from engine.apy.event_ledger import EVENT_LEDGER_CONTRACT_VERSION
from engine.apy.event_ledger_economics import (
    HEALTH_ECONOMICS_CONTRACT_VERSION,
    classify_incremental_result,
    run_event_ledger_health_economics,
)
from engine.apy.expected_value import run_expected_value
from engine.apy.ltbi_state import apply_ltbi_state_edit, enable_development_compatibility_mode
from engine.apy.runner import run_replicates
from engine.apy.results_bundle import build_results_bundle


class ApyEventLedgerEconomicsTests(unittest.TestCase):
    def test_zero_intervention_has_zero_incremental_cost_and_dalys(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 100, "screenCoverage": 0}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ({"program_setup": 0, "program_running": 0}))
        reps = econ["replicateResults"]
        primary = reps[reps["discountRate"] == 0.03].iloc[0]

        self.assertAlmostEqual(primary["incrementalCost"], 0)
        self.assertAlmostEqual(primary["dalysAverted"], 0)
        self.assertEqual(primary["classification"], "no material difference")

    def test_valid_synthetic_cost_and_daly_calculation(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 80, "screeningStrategy": "random"}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ())
        annual = econ["annualByArm"]
        year0 = annual[
            (annual["arm"] == "intervention")
            & (annual["discountRate"] == 0.0)
            & (annual["modelYear"] == 0)
        ].iloc[0]

        self.assertEqual(econ["contractVersion"], HEALTH_ECONOMICS_CONTRACT_VERSION)
        self.assertAlmostEqual(year0["screeningTestCost"], year0["screened"] * 10)
        self.assertAlmostEqual(year0["programSetupCost"], 1000)
        self.assertEqual(econ["validation"]["isValid"], True)

    def test_unresolved_cost_prevents_complete_total(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 80}))
        econ_config = _synthetic_econ()
        for item in econ_config["costItems"]:
            if item["costItemId"] == "test_igra":
                item["originalPriceYear"] = None
                item["sourceYearStatus"] = "unknown"
        econ = run_event_ledger_health_economics(result, econ_config)

        self.assertIn("costItems.test_igra", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertFalse(econ["status"]["isComplete"])
        self.assertIn("ICER", econ["status"]["notCalculated"])

    def test_missing_annual_cost_is_not_skipped_by_replicate_aggregation(self) -> None:
        config = _synthetic_econ()
        _unresolve_cost(config, "test_igra")

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        annual = econ["annualByArm"]
        reps = econ["replicateResults"]
        row = reps[(reps["discountRate"] == 0.0) & (reps["replicateId"] == 0)].iloc[0]

        self.assertTrue((annual["activeTBDiseaseCost"].dropna() > 0).any())
        self.assertFalse(row["costPairComplete"])
        self.assertIsNone(row["incrementalCost"])
        self.assertIsNone(row["replicateICER"])
        self.assertEqual(row["classification"], "incomplete / not calculated")

    def test_unresolved_active_tb_cost_with_cases_blocks_totals(self) -> None:
        config = _synthetic_econ()
        _unresolve_cost(config, "active_tb_disease")

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertFalse(row["costPairComplete"])
        self.assertIsNone(row["incrementalCost"])

    def test_unresolved_active_tb_daly_input_blocks_dalys_and_icer(self) -> None:
        config = _synthetic_econ()
        config["dalyAssumptions"]["activeTBDisabilityWeight"] = _assumption(None)

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertFalse(row["dalyPairComplete"])
        self.assertIsNone(row["dalysAverted"])
        self.assertIsNone(row["replicateICER"])

    def test_zero_resource_use_does_not_require_unit_cost(self) -> None:
        config = _synthetic_econ()
        _unresolve_cost(config, "test_igra")

        econ = run_event_ledger_health_economics(_toy_ledger(screened=0, starts=0), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertTrue(row["costPairComplete"])
        self.assertIsNotNone(row["incrementalCost"])

    def test_ratio_of_means_uses_identical_complete_pair_set(self) -> None:
        config = _synthetic_econ()
        _unresolve_cost(config, "test_igra")
        ledger = _toy_ledger(replicates=[
            {"replicateId": 0, "screened": 0, "starts": 0, "comp_tb": 2, "int_tb": 2, "prevented": 0},
            {"replicateId": 1, "screened": 10, "starts": 5, "comp_tb": 2, "int_tb": 1, "prevented": 1},
        ])

        econ = run_event_ledger_health_economics(ledger, config)
        summary = econ["summaries"]
        primary = summary[
            (summary["discountRate"] == 0.0)
            & (summary["metric"] == "primaryICER_ratioOfMeans")
        ].iloc[0]

        self.assertEqual(primary["totalPairedReplicates"], 2)
        self.assertEqual(primary["completePairedReplicates"], 1)
        self.assertEqual(primary["excludedPairedReplicates"], 1)

    def test_active_tb_cost_reconciliation_and_no_double_subtraction(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 120, "screeningStrategy": "random"}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ({"program_setup": 0, "program_running": 0}))
        annual = econ["annualByArm"]
        rows = econ["replicateResults"]
        r = rows[(rows["discountRate"] == 0.0)].iloc[0]
        tb = annual[annual["discountRate"] == 0.0].groupby("arm")["activeTBDiseaseCost"].sum()
        programme = annual[
            (annual["discountRate"] == 0.0)
            & (annual["arm"] == "intervention")
        ][["screeningTestCost", "tptRegimenCost", "falsePositiveIncrementalCost", "programSetupCost", "programRunningCost", "adrManagementCost"]].sum().sum()

        self.assertAlmostEqual(tb["comparator"] - tb["intervention"], (r["activeTBCasesPrevented"] * 1000))
        self.assertAlmostEqual(r["incrementalCost"], programme + tb["intervention"] - tb["comparator"])

    def test_programme_running_basis_and_setup_timing(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 100, "screeningWindowYears": 2}))
        econ = run_event_ledger_health_economics(
            result,
            _synthetic_econ({"program_running_basis": "annual_during_screening_window", "program_running": 50}),
        )
        annual = econ["annualByArm"]

        setup_rows = annual[(annual["programSetupCost"] > 0) & (annual["discountRate"] == 0.0)]
        running_years = annual[(annual["programRunningCost"] > 0) & (annual["discountRate"] == 0.0)]["modelYear"].unique()
        self.assertEqual(set(setup_rows["modelYear"]), {0})
        self.assertTrue(set(running_years).issubset({0, 1}))

        bad = _synthetic_econ({"program_running": 50, "program_running_basis": ""})
        unresolved = run_event_ledger_health_economics(result, bad)
        self.assertIn("costItems.program_running.resourceUse.costBasis", {item["field"] for item in unresolved["unresolvedInputs"]})

        total = run_event_ledger_health_economics(
            result,
            _synthetic_econ(
                {
                    "program_running": 120,
                    "program_running_basis": "total_over_screening_programme",
                    "program_running_allocation": "proportional_to_screening_volume",
                }
            ),
        )
        running = total["annualByArm"][
            (total["annualByArm"]["arm"] == "intervention")
            & (total["annualByArm"]["discountRate"] == 0.0)
        ]["programRunningCost"].sum()
        self.assertAlmostEqual(running, 120)

    def test_setup_and_running_costs_survive_zero_event_years(self) -> None:
        config = _synthetic_econ({"program_running": 25})
        ledger = _toy_ledger(screening_window=3, screened_by_year={0: 5, 1: 0, 2: 5})

        econ = run_event_ledger_health_economics(ledger, config)
        annual = econ["annualByArm"]
        running = annual[
            (annual["arm"] == "intervention")
            & (annual["discountRate"] == 0.0)
            & (annual["modelYear"].isin([0, 1, 2]))
        ][["modelYear", "programRunningCost"]]

        self.assertEqual(set(running["modelYear"]), {0, 1, 2})
        self.assertTrue((running["programRunningCost"] == 25).all())

    def test_cost_target_year_must_match_analysis(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 80}))
        config = _synthetic_econ()
        for item in config["costItems"]:
            if item["costItemId"] == "test_igra":
                item["targetPriceYear"] = "2024-25"

        econ = run_event_ledger_health_economics(result, config)

        self.assertIn("costItems.test_igra.targetPriceYear", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertIsNone(econ["costs"]["testingCost"])

    def test_threshold_currency_year_mismatch_prevents_nmb(self) -> None:
        config = _synthetic_econ({"threshold": 50000})
        config["threshold"]["referenceYear"] = "2024-25"

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertIn("threshold.alignment", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertIsNone(row["netMonetaryBenefit"])

    def test_incorrect_cost_basis_prevents_component_calculation(self) -> None:
        config = _synthetic_econ()
        for item in config["costItems"]:
            if item["costItemId"] == "test_igra":
                item["resourceUse"]["costBasis"] = "per_completed_course"

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertIn("costItems.test_igra.resourceUse.costBasis", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertFalse(row["costPairComplete"])

    def test_false_positive_and_adr_costs_do_not_create_benefit(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 100, "testSpecificity": 0}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ({"adr": 7}))
        annual = econ["annualByArm"]
        intervention = annual[(annual["arm"] == "intervention") & (annual["discountRate"] == 0.0)]

        self.assertGreater(intervention["falsePositiveIncrementalCost"].sum(), 0)
        self.assertAlmostEqual(
            intervention["adrManagementCost"].sum(),
            intervention["tpt_adr_stop_total"].sum() * 7,
        )
        self.assertLessEqual(intervention["active_tb_cases_prevented"].sum(), intervention["infection_effectively_treated_total"].sum())

    def test_dale_2019_working_defaults_calculate_icer_but_not_nmb(self) -> None:
        econ = run_event_ledger_health_economics(
            _toy_ledger(screened=500, starts=200, adr_stops=10),
            build_economics_preset_dale2019_aud("3HP"),
        )
        primary = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.03].iloc[0]

        self.assertTrue(primary["economicPairComplete"])
        self.assertIsNotNone(primary["incrementalCost"])
        self.assertIsNotNone(primary["dalysAverted"])
        self.assertIsNotNone(primary["replicateICER"])
        self.assertIsNone(primary["netMonetaryBenefit"])
        self.assertIn("threshold.value", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertTrue(econ["metadata"]["isProvisional"])

    def test_dale_2019_false_positive_people_still_incur_ordinary_costs(self) -> None:
        econ = run_event_ledger_health_economics(
            _toy_ledger(screened=10, starts=5, adr_stops=1),
            build_economics_preset_dale2019_aud("3HP"),
        )
        annual = econ["annualByArm"]
        intervention = annual[(annual["arm"] == "intervention") & (annual["discountRate"] == 0.0)]

        self.assertGreater(intervention["screeningTestCost"].sum(), 0)
        self.assertGreater(intervention["tptRegimenCost"].sum(), 0)
        self.assertGreater(intervention["adrManagementCost"].sum(), 0)
        self.assertEqual(intervention["falsePositiveIncrementalCost"].sum(), 0)
        self.assertAlmostEqual(intervention["adrManagementCost"].sum(), 39.4059)

    def test_dale_2019_3hr_leaves_adr_management_unresolved(self) -> None:
        econ = run_event_ledger_health_economics(_toy_ledger(adr_stops=1), build_economics_preset_dale2019_aud("3HR"))
        primary = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.03].iloc[0]

        self.assertFalse(primary["costPairComplete"])
        self.assertIn("adrManagementCost", primary["exclusionReasons"])

    def test_dale_2019_workbook_records_costs_and_authoritative_icer(self) -> None:
        cfg = build_economics_preset_dale2019_aud("3HP")
        ledger = _toy_ledger(screened=500, starts=200, adr_stops=10)
        econ = run_event_ledger_health_economics(ledger, cfg)
        payload = build_results_workbook(
            config=_reviewed_epi({"testType": "IGRA", "regimen": "3HP"}),
            bundle=ledger,
            economics_results=econ,
            economics_config=cfg,
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        cost_ws = wb["Cost_normalisation"]
        headers = [cell.value for cell in next(cost_ws.iter_rows(min_row=1, max_row=1))]
        rows = {
            row[headers.index("costItemId")].value: row
            for row in cost_ws.iter_rows(min_row=2)
        }
        igra = rows["test_igra"]
        self.assertEqual(igra[headers.index("originalPriceYear")].value, "2019")
        self.assertEqual(igra[headers.index("targetPriceYear")].value, "2019")
        self.assertEqual(igra[headers.index("inflationFactor")].value, 1)
        self.assertAlmostEqual(igra[headers.index("convertedTargetYearCost")].value, 113.48)

        summary_ws = wb["Economic_summary"]
        summary_headers = [cell.value for cell in next(summary_ws.iter_rows(min_row=1, max_row=1))]
        workbook_icer = None
        for row in summary_ws.iter_rows(min_row=2):
            if row[summary_headers.index("discountRate")].value == 0.03 and row[summary_headers.index("metric")].value == "primaryICER_ratioOfMeans":
                workbook_icer = row[summary_headers.index("mean")].value
                break
        primary = econ["summaries"][
            (econ["summaries"]["discountRate"] == 0.03)
            & (econ["summaries"]["metric"] == "primaryICER_ratioOfMeans")
        ].iloc[0]["mean"]
        wb.close()

        self.assertAlmostEqual(workbook_icer, primary)

    def test_missing_adr_cost_and_harm_require_reviewed_decisions(self) -> None:
        config = _synthetic_econ({"adr": None, "adr_loss": None})
        _unresolve_cost(config, "tpt_adr_management")
        config["dalyAssumptions"]["includeADRHealthLoss"] = False
        config["dalyAssumptions"]["adrHealthLossExclusionStatus"] = "unresolved"
        config["dalyAssumptions"]["adrHealthLossExclusionRationale"] = ""

        econ = run_event_ledger_health_economics(_toy_ledger(adr_stops=1), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertFalse(row["costPairComplete"])
        self.assertFalse(row["dalyPairComplete"])
        self.assertIn("dalyAssumptions.includeADRHealthLoss", {item["field"] for item in econ["unresolvedInputs"]})

    def test_unresolved_post_tb_exclusion_prevents_complete_dalys(self) -> None:
        config = _synthetic_econ()
        config["dalyAssumptions"]["postTBSequelaeStatus"] = "unresolved"
        config["dalyAssumptions"]["postTBSequelaeExclusionRationale"] = ""

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertFalse(row["dalyPairComplete"])
        self.assertIn("dalyAssumptions.includePostTBSequelae", {item["field"] for item in econ["unresolvedInputs"]})

    def test_invalid_daly_ranges_are_rejected(self) -> None:
        config = _synthetic_econ()
        config["dalyAssumptions"]["activeTBDisabilityWeight"] = _assumption(1.5)
        config["dalyAssumptions"]["tbCaseFatalityRisk"] = _assumption(-0.1)

        econ = run_event_ledger_health_economics(_toy_ledger(), config)

        self.assertIn("dalyAssumptions.activeTBDisabilityWeight", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertIn("dalyAssumptions.tbCaseFatalityRisk", {item["field"] for item in econ["unresolvedInputs"]})

    def test_discounting_known_year_and_no_quantity_change(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 80}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ())
        annual = econ["annualByArm"]
        y10 = annual[(annual["modelYear"] == 10) & (annual["discountRate"] == 0.03)]
        if y10.empty:
            self.skipTest("Fixture has no year-10 economic events.")
        row = y10.iloc[0]
        self.assertAlmostEqual(row["discountFactor"], 1 / (1.03 ** 10))
        same = annual[
            (annual["modelYear"] == row["modelYear"])
            & (annual["arm"] == row["arm"])
            & (annual["discountRate"] == 0.0)
            & (annual["replicateId"] == row["replicateId"])
        ].iloc[0]
        self.assertEqual(row["active_tb_cases"], same["active_tb_cases"])

    def test_daly_equations_and_treatment_harms(self) -> None:
        result = run_expected_value(_reviewed_epi({"N": 80}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ({"tpt_loss": 0.5}))
        annual = econ["annualByArm"]
        row = annual[(annual["active_tb_cases"] > 0) & (annual["discountRate"] == 0.0)].iloc[0]

        self.assertAlmostEqual(row["activeTBYLD"], row["active_tb_cases"] * 0.2 * 0.5)
        self.assertAlmostEqual(row["activeTBYLL"], row["active_tb_cases"] * 0.1 * 10)
        reps = econ["replicateResults"]
        r = reps[reps["discountRate"] == 0.0].iloc[0]
        self.assertAlmostEqual(r["dalysAverted"], r["comparatorDALYs"] - r["interventionDALYs"])

    def test_icer_ratio_of_means_not_mean_replicate_icer(self) -> None:
        result = run_replicates(_reviewed_epi({"N": 180, "nReps": 5, "seed": 5}), keep_example_cohort=False)
        econ = run_event_ledger_health_economics(result, _synthetic_econ())
        reps = econ["replicateResults"]
        reps = reps[(reps["discountRate"] == 0.03) & reps["replicateICER"].notna()]
        summary = econ["summaries"]
        primary = summary[
            (summary["discountRate"] == 0.03)
            & (summary["metric"] == "primaryICER_ratioOfMeans")
        ].iloc[0]["mean"]

        self.assertAlmostEqual(primary, reps["incrementalCost"].mean() / reps["dalysAverted"].mean())
        if len(reps) > 1 and not math.isclose(reps["replicateICER"].mean(), primary):
            self.assertNotAlmostEqual(primary, reps["replicateICER"].mean())

    def test_dominance_and_nmb(self) -> None:
        self.assertEqual(classify_incremental_result(None, 0), "incomplete / not calculated")
        self.assertEqual(classify_incremental_result(float("nan"), 0), "incomplete / not calculated")
        self.assertEqual(classify_incremental_result(-1, 1), "dominant")
        self.assertEqual(classify_incremental_result(1, -1), "dominated")
        self.assertEqual(classify_incremental_result(-1, -1), "cost saving with health loss")
        self.assertEqual(classify_incremental_result(0, 1), "health gain at no additional cost")
        self.assertEqual(classify_incremental_result(1, 0), "increased cost with no material health difference")
        econ = run_event_ledger_health_economics(
            run_expected_value(_reviewed_epi({"N": 80})),
            _synthetic_econ({"threshold": 50000}),
        )
        reps = econ["replicateResults"]
        row = reps[reps["discountRate"] == 0.03].iloc[0]
        self.assertAlmostEqual(row["netMonetaryBenefit"], 50000 * row["dalysAverted"] - row["incrementalCost"])

        no_nmb = run_event_ledger_health_economics(run_expected_value(_reviewed_epi({"N": 80})), _synthetic_econ())
        self.assertIn("threshold.value", {item["field"] for item in no_nmb["unresolvedInputs"]})

    def test_negative_or_zero_health_gain_has_no_primary_numeric_icer(self) -> None:
        econ = run_event_ledger_health_economics(
            _toy_ledger(prevented=0, comp_tb=1, int_tb=1),
            _synthetic_econ({"threshold": 50000}),
        )
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]
        primary = econ["summaries"][
            (econ["summaries"]["discountRate"] == 0.0)
            & (econ["summaries"]["metric"] == "primaryICER_ratioOfMeans")
        ].iloc[0]

        self.assertIsNone(row["replicateICER"])
        self.assertIsNone(primary["mean"])

    def test_missing_nmb_excluded_and_deterministic_has_no_probability_row(self) -> None:
        deterministic = run_event_ledger_health_economics(_toy_ledger(model_type="expected_value"), _synthetic_econ())
        metrics = set(deterministic["summaries"]["metric"])

        self.assertNotIn("probabilityPositiveNMB_fixedParameterSimulation", metrics)

        stochastic = run_event_ledger_health_economics(_toy_ledger(replicates=[
            {"replicateId": 0, "screened": 0, "starts": 0, "comp_tb": 2, "int_tb": 2, "prevented": 0},
            {"replicateId": 1, "screened": 10, "starts": 5, "comp_tb": 2, "int_tb": 1, "prevented": 1},
        ]), _synthetic_econ())
        self.assertNotIn("probabilityPositiveNMB_fixedParameterSimulation", set(stochastic["summaries"]["metric"]))

    def test_outside_horizon_events_visible_but_excluded_from_totals(self) -> None:
        config = _synthetic_econ({"economic_horizon": 1})

        econ = run_event_ledger_health_economics(_toy_ledger(active_year=5), config)
        annual = econ["annualByArm"]
        outside = annual[(annual["modelYear"] == 5) & (annual["discountRate"] == 0.0)]
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]

        self.assertFalse(outside["includedInEconomicAnalysis"].any())
        self.assertTrue((outside["economicExclusionReason"] != "").all())
        self.assertAlmostEqual(row["comparatorCost"], 0.0)

    def test_total_programme_cost_includes_all_program_components(self) -> None:
        econ = run_event_ledger_health_economics(_toy_ledger(), _synthetic_econ({"adr": 7}))
        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]
        expected = sum(
            row[f"intervention_{component}Discounted"]
            for component in [
                "screeningTestCost",
                "tptRegimenCost",
                "falsePositiveIncrementalCost",
                "programSetupCost",
                "programRunningCost",
                "adrManagementCost",
            ]
        )

        self.assertAlmostEqual(row["intervention_totalProgrammeCostDiscounted"], expected)

    def test_provisional_epidemiology_propagates(self) -> None:
        result = run_expected_value(enable_development_compatibility_mode({"N": 80}))
        econ = run_event_ledger_health_economics(result, _synthetic_econ())

        self.assertTrue(econ["metadata"]["isProvisional"])
        self.assertFalse(econ["metadata"]["conclusionPermitted"])

    def test_existing_public_economics_routes_to_event_ledger(self) -> None:
        bundle = build_results_bundle(run_replicates(_reviewed_epi({"N": 80, "nReps": 2}), keep_example_cohort=False))
        econ = run_health_economics(bundle, _synthetic_econ())

        self.assertEqual(econ["contractVersion"], HEALTH_ECONOMICS_CONTRACT_VERSION)
        self.assertEqual(econ["source"], "event_ledger_health_economics_v2")

    def test_workbook_values_match_authoritative_economics(self) -> None:
        results = run_replicates(_reviewed_epi({"N": 80, "nReps": 2, "seed": 22}), keep_example_cohort=False)
        bundle = build_results_bundle(results)
        econ = run_event_ledger_health_economics(bundle, _synthetic_econ())
        payload = build_results_workbook(
            config=results["interfaceConfig"],
            bundle=bundle,
            backend_status={"name": "python_apy"},
            economics_results=econ,
            economics_config=_synthetic_econ(),
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        rows = list(wb["Economic_replicates"].iter_rows(values_only=True))
        headers = rows[0]
        first = dict(zip(headers, rows[1]))
        expected = econ["replicateResults"].iloc[0]

        self.assertIn("Economic_annual_by_arm", wb.sheetnames)
        self.assertEqual(first["incrementalCost"], expected["incrementalCost"])
        wb.close()

    def test_workbook_blanks_unavailable_aggregate_results(self) -> None:
        config = _synthetic_econ()
        _unresolve_cost(config, "test_igra")
        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        payload = build_results_workbook(
            config=_reviewed_epi({}),
            bundle={"metadata": {}, "headline": {}, "technical": {"eventLedger": _toy_ledger()["technical"]["eventLedger"]}},
            backend_status={"name": "python_apy"},
            economics_results=econ,
            economics_config=config,
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        rows = list(wb["Economic_replicates"].iter_rows(values_only=True))
        first = dict(zip(rows[0], rows[1]))

        self.assertIn("costPairComplete", rows[0])
        self.assertIsNone(first["incrementalCost"])
        wb.close()


def _reviewed_epi(overrides: dict) -> dict:
    cfg = apply_ltbi_state_edit(
        build_default_config(),
        baseline_recent_proportion=0.35,
        transition_rate_per_year=0.2,
        source="Reviewed unit-test baseline recent LTBI source",
        status="configured_reviewed",
        notes="Synthetic fixture.",
    )
    cfg.update(overrides)
    return cfg


def _synthetic_econ(overrides: dict | None = None) -> dict:
    overrides = overrides or {}
    config = build_default_economics_config()
    if "economic_horizon" in overrides:
        config["metadata"]["economicHorizonYears"] = overrides["economic_horizon"]
    values = {
        "test_igra": 10,
        "test_tst": 8,
        "regimen_3hp": 100,
        "regimen_4r": 90,
        "regimen_3hr": 95,
        "regimen_6h": 80,
        "regimen_9h": 70,
        "active_tb_disease": 1000,
        "false_positive_incremental": 5,
        "program_setup": overrides.get("program_setup", 1000),
        "program_running": overrides.get("program_running", 0),
        "tpt_adr_management": overrides.get("adr", 0),
    }
    for item in config["costItems"]:
        item_id = item["costItemId"]
        if item_id not in values:
            continue
        item["originalCost"] = values[item_id]
        item["originalPriceYear"] = "2025-26"
        item["sourceYearStatus"] = "explicit"
        item["sourceCitation"] = "Synthetic unit-test cost fixture"
        item["resourceUse"]["costBasis"] = item["resourceUse"].get("costBasis") or "per_event"
        if item_id == "program_running":
            item["resourceUse"]["costBasis"] = overrides.get("program_running_basis", "annual_during_screening_window")
            if "program_running_allocation" in overrides:
                item["resourceUse"]["allocationMethod"] = overrides["program_running_allocation"]
    config["dalyAssumptions"] = {
        "activeTBDisabilityWeight": _assumption(0.2),
        "activeTBDurationYears": _assumption(0.5),
        "tbCaseFatalityRisk": _assumption(0.1),
        "yllPerTBDeath": _assumption(10),
        "includeTPTHealthLoss": "tpt_loss" in overrides,
        "tptHealthLossExclusionStatus": "reviewed_exclusion" if "tpt_loss" not in overrides else "configured_reviewed",
        "tptHealthLossExclusionRationale": "Synthetic reviewed exclusion.",
        "dalyLossPerTPTStarted": _assumption(overrides.get("tpt_loss")),
        "includeADRHealthLoss": "adr_loss" in overrides,
        "adrHealthLossExclusionStatus": "reviewed_exclusion" if "adr_loss" not in overrides else "configured_reviewed",
        "adrHealthLossExclusionRationale": "Synthetic reviewed ADR exclusion.",
        "dalyLossPerADRStop": _assumption(0.01),
        "includePostTBSequelae": False,
        "postTBSequelaeStatus": "reviewed_exclusion",
        "postTBSequelaeExclusionRationale": "Synthetic reviewed post-TB exclusion.",
    }
    if "threshold" in overrides:
        config["threshold"].update(
            {
                "value": overrides["threshold"],
                "currency": "AUD",
                "referenceYear": "2025-26",
                "source": "Synthetic unit-test threshold fixture",
                "status": "configured_reviewed",
            }
        )
    return config


def _unresolve_cost(config: dict, cost_item_id: str) -> None:
    for item in config["costItems"]:
        if item["costItemId"] == cost_item_id:
            item["originalCost"] = None
            item["originalPriceYear"] = None
            item["sourceYearStatus"] = "unknown"
            item["convertedTargetYearCost"] = None
            item["conversionStatus"] = "unresolved_source_price_year"
            item["conversionApplied"] = False


def _assumption(value):
    return {
        "value": value,
        "source": "Synthetic unit-test DALY fixture" if value is not None else "",
        "status": "configured_reviewed" if value is not None else "unresolved",
        "notes": "Synthetic fixture.",
        "provisional": False if value is not None else True,
    }


def _toy_ledger(
    *,
    model_type: str = "agent_based",
    screened: float = 10,
    starts: float = 5,
    adr_stops: float = 1,
    comp_tb: float = 2,
    int_tb: float = 1,
    prevented: float = 1,
    active_year: int = 5,
    screening_window: int = 3,
    screened_by_year: dict[int, float] | None = None,
    replicates: list[dict] | None = None,
) -> dict:
    reps = replicates or [
        {
            "replicateId": 0,
            "screened": screened,
            "starts": starts,
            "adr_stops": adr_stops,
            "comp_tb": comp_tb,
            "int_tb": int_tb,
            "prevented": prevented,
        }
    ]
    annual = []
    value_type = "expected" if model_type == "expected_value" else "simulated_count"
    for rep in reps:
        rep_id = rep.get("replicateId", 0)
        seed = None if model_type == "expected_value" else 100 + int(rep_id)
        schedule = screened_by_year or {0: rep.get("screened", screened)}
        for arm in ["comparator", "intervention"]:
            for year, n_screened in schedule.items():
                if arm == "intervention":
                    _append_event(annual, model_type, value_type, arm, rep_id, seed, year, "screened", n_screened)
                    _append_event(annual, model_type, value_type, arm, rep_id, seed, year, "tpt_started_total", rep.get("starts", starts) if year == min(schedule) else 0)
                    _append_event(annual, model_type, value_type, arm, rep_id, seed, year, "tpt_started_false_positive", 1 if year == min(schedule) and rep.get("starts", starts) else 0)
                    _append_event(annual, model_type, value_type, arm, rep_id, seed, year, "tpt_adr_stop_total", rep.get("adr_stops", adr_stops) if year == min(schedule) else 0)
                    _append_event(annual, model_type, value_type, arm, rep_id, seed, year, "infection_effectively_treated_total", 1 if year == min(schedule) and rep.get("starts", starts) else 0)
            active = rep.get("comp_tb", comp_tb) if arm == "comparator" else rep.get("int_tb", int_tb)
            _append_event(annual, model_type, value_type, arm, rep_id, seed, active_year, "active_tb_cases", active)
            if arm == "intervention":
                _append_event(annual, model_type, value_type, arm, rep_id, seed, active_year, "active_tb_cases_prevented", rep.get("prevented", prevented))
    return {
        "technical": {
            "eventLedger": {
                "metadata": {
                    "contractVersion": EVENT_LEDGER_CONTRACT_VERSION,
                    "scenarioId": "toy_economics",
                    "modelType": model_type,
                    "backend": "python_apy",
                    "screeningWindowYears": screening_window,
                    "followUpHorizonYears": 20,
                },
                "definitions": [],
                "replicateTotals": [],
                "annualEvents": annual,
                "validation": {"isValid": True, "errors": [], "warnings": []},
                "summaries": [],
            }
        }
    }


def _append_event(rows, model_type, value_type, arm, rep_id, seed, year, name, value):
    rows.append(
        {
            "scenarioId": "toy_economics",
            "modelType": model_type,
            "backend": "python_apy",
            "arm": arm,
            "replicateId": rep_id,
            "pairedReplicateId": rep_id,
            "replicateSeed": seed,
            "valueType": value_type,
            "screeningWindowYears": 3,
            "followUpHorizonYears": 20,
            "contractVersion": EVENT_LEDGER_CONTRACT_VERSION,
            "modelVersion": "toy",
            "modelYear": year,
            "timeInterval": f"[{year}, {year + 1})",
            "withinFollowUp": True,
            "eventName": name,
            "value": value,
        }
    )


if __name__ == "__main__":
    unittest.main()
