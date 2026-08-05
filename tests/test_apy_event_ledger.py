from __future__ import annotations

import json
import unittest
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from adapters.serialization import to_json_like
from app.results_workbook import build_results_workbook
from engine.apy.event_ledger import EVENT_LEDGER_CONTRACT_VERSION, matlab_total_ledger_from_raw_rows
from engine.apy.expected_value import run_expected_value
from engine.apy.results_bundle import build_results_bundle
from engine.apy.runner import run_replicates


class ApyEventLedgerTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.ev = run_expected_value({"N": 200, "screeningStrategy": "random", "seed": 1})
        cls.abm = run_replicates(
            {"N": 200, "nReps": 4, "screeningStrategy": "random", "seed": 1},
            keep_example_cohort=False,
        )

    def test_event_ledger_contract_metadata_and_validation(self) -> None:
        ledger = self.ev["eventLedger"]

        self.assertEqual(ledger["metadata"]["contractVersion"], EVENT_LEDGER_CONTRACT_VERSION)
        self.assertEqual(ledger["metadata"]["modelType"], "expected_value")
        self.assertTrue(ledger["validation"]["isValid"])
        self.assertIn("year 0 is [0,1)", ledger["metadata"]["yearBinConvention"])

    def test_abm_ledger_preserves_replicates_and_seeds(self) -> None:
        totals = self.abm["eventLedger"]["replicateTotals"]

        self.assertTrue(self.abm["eventLedger"]["validation"]["isValid"])
        self.assertEqual(set(totals["modelType"]), {"agent_based"})
        self.assertEqual(set(totals["valueType"]), {"simulated_count"})
        self.assertEqual(len(set(totals["replicateId"])), 4)
        self.assertEqual(len(set(totals["replicateSeed"])), 4)

    def test_comparator_intervention_tb_identity_holds_per_replicate(self) -> None:
        wide = _long_to_wide(self.abm["eventLedger"]["replicateTotals"])
        for replicate_id, group in wide.groupby("replicateId"):
            comparator = group[group["arm"] == "comparator"].iloc[0]
            intervention = group[group["arm"] == "intervention"].iloc[0]
            self.assertEqual(
                comparator["active_tb_cases"],
                intervention["active_tb_cases"] + intervention["active_tb_cases_prevented"],
                msg=f"replicate {replicate_id}",
            )

    def test_annual_totals_reconcile_with_cumulative_totals(self) -> None:
        for ledger in [self.ev["eventLedger"], self.abm["eventLedger"]]:
            self.assertTrue(ledger["validation"]["isValid"])
            totals = _long_to_wide(ledger["replicateTotals"])
            annual = ledger["annualEvents"]
            grouped = annual.groupby(
                ["modelType", "replicateId", "pairedReplicateId", "arm", "eventName"],
                dropna=False,
            )["value"].sum()
            for _, row in totals.iterrows():
                key = (row["modelType"], row["replicateId"], row["pairedReplicateId"], row["arm"])
                self.assertAlmostEqual(
                    grouped.get((*key, "screened"), 0.0),
                    row["screened"],
                    places=7,
                )

    def test_zero_screening_coverage_has_no_screening_or_tpt_events(self) -> None:
        ledger = run_expected_value({"N": 100, "screenCoverage": 0})["eventLedger"]
        wide = _long_to_wide(ledger["replicateTotals"])
        intervention = wide[wide["arm"] == "intervention"].iloc[0]
        comparator = wide[wide["arm"] == "comparator"].iloc[0]

        for event in ["screened", "tpt_started_total", "infection_effectively_treated_total", "active_tb_cases_prevented"]:
            self.assertEqual(intervention[event], 0)
        self.assertAlmostEqual(intervention["active_tb_cases"], comparator["active_tb_cases"])

    def test_test_sensitivity_boundaries(self) -> None:
        sens0 = _intervention_totals(run_expected_value({"N": 100, "testSensitivity": 0})["eventLedger"])
        sens1 = _intervention_totals(run_expected_value({"N": 100, "testSensitivity": 1})["eventLedger"])

        self.assertAlmostEqual(sens0["true_positive_latent"], 0)
        self.assertAlmostEqual(sens0["test_positive_active"], 0)
        self.assertAlmostEqual(sens1["false_negative_latent"], 0)
        self.assertAlmostEqual(sens1["test_negative_active"], 0)

    def test_test_specificity_boundaries(self) -> None:
        spec1 = _intervention_totals(run_expected_value({"N": 100, "testSpecificity": 1})["eventLedger"])
        spec0 = _intervention_totals(run_expected_value({"N": 100, "testSpecificity": 0})["eventLedger"])

        self.assertAlmostEqual(spec1["false_positive"], 0)
        self.assertAlmostEqual(spec0["false_positive"], spec0["uninfected_screened"])

    def test_treatment_start_and_efficacy_boundaries(self) -> None:
        no_start = _intervention_totals(run_expected_value({"N": 100, "pStartTPT": 0})["eventLedger"])
        no_eff = _intervention_totals(run_expected_value({"N": 100, "regimenEffFull": 0})["eventLedger"])

        for event in ["tpt_started_total", "tpt_completed_total", "infection_effectively_treated_total", "active_tb_cases_prevented"]:
            self.assertAlmostEqual(no_start[event], 0)
        self.assertGreater(no_eff["tpt_started_total"], 0)
        self.assertGreater(no_eff["tpt_completed_total"], 0)
        self.assertAlmostEqual(no_eff["infection_effectively_treated_total"], 0)
        self.assertAlmostEqual(no_eff["active_tb_cases_prevented"], 0)

    def test_completion_zero_retains_partial_course_rule(self) -> None:
        totals = _intervention_totals(
            run_expected_value(
                {
                    "N": 100,
                    "regimenPComplete": 0,
                    "partialShortCourseMode": "linear",
                    "partialDoseFractionOther": 0.6,
                }
            )["eventLedger"]
        )

        self.assertAlmostEqual(totals["infection_effectively_treated_full"], 0)
        self.assertGreater(totals["infection_effectively_treated_partial"], 0)

    def test_igra_and_tst_characteristics_and_bcg_false_positives(self) -> None:
        igra = _intervention_totals(run_expected_value({"N": 100, "testType": "IGRA"})["eventLedger"])
        tst = _intervention_totals(run_expected_value({"N": 100, "testType": "TST"})["eventLedger"])

        self.assertNotAlmostEqual(igra["true_positive_latent"], tst["true_positive_latent"])
        self.assertGreater(tst["false_positive_bcg"], 0)
        self.assertGreaterEqual(tst["false_positive_due_to_bcg"], 0)
        self.assertAlmostEqual(tst["false_positive"], tst["false_positive_bcg"] + tst["false_positive_no_bcg"])

    def test_fixed_abm_seed_and_expected_value_seed_behaviour(self) -> None:
        a = run_replicates({"N": 80, "nReps": 2, "seed": 123}, keep_example_cohort=False)
        b = run_replicates({"N": 80, "nReps": 2, "seed": 123}, keep_example_cohort=False)
        ev_a = run_expected_value({"N": 100, "seed": 123})
        ev_b = run_expected_value({"N": 100, "seed": 999})

        self.assertEqual(
            to_json_like(a["eventLedger"]["replicateTotals"]),
            to_json_like(b["eventLedger"]["replicateTotals"]),
        )
        self.assertEqual(
            to_json_like(ev_a["eventLedger"]["replicateTotals"]),
            to_json_like(ev_b["eventLedger"]["replicateTotals"]),
        )

    def test_event_ledger_outputs_are_json_serialisable(self) -> None:
        json.dumps(to_json_like(self.ev["eventLedger"]))
        json.dumps(to_json_like(self.abm["eventLedger"]))

    def test_legacy_raw_apy_fields_remain_available(self) -> None:
        raw = self.abm["raw"]

        for field in ["nScreened", "nCuredInfection", "nPreventedActiveTB", "nTruePositiveLatent"]:
            self.assertIn(field, raw.columns)

    def test_results_bundle_exposes_download_tables(self) -> None:
        bundle = build_results_bundle(self.abm)

        self.assertIn("eventLedger", bundle["technical"])
        self.assertTrue(bundle["technical"]["eventLedger"]["validation"]["isValid"])
        self.assertGreater(len(bundle["downloads"]["eventLedgerTotals"]), 0)
        self.assertGreater(len(bundle["downloads"]["eventLedgerAnnual"]), 0)
        self.assertGreater(len(bundle["downloads"]["eventDefinitions"]), 0)

    def test_workbook_export_contains_event_ledger_sheets(self) -> None:
        bundle = build_results_bundle(self.abm)
        payload = build_results_workbook(
            config=self.abm["interfaceConfig"],
            bundle=bundle,
            backend_status={"name": "python_apy"},
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        try:
            for sheet in [
                "Event_ledger_totals",
                "Event_ledger_annual",
                "Event_definitions",
                "Event_ledger_validation",
            ]:
                self.assertIn(sheet, wb.sheetnames)
        finally:
            wb.close()

    def test_matlab_adapter_reports_missing_unambiguous_fields(self) -> None:
        adapted = matlab_total_ledger_from_raw_rows(
            config=self.abm["interfaceConfig"],
            raw_rows=[{"rep": 1, "nScreened": 10}],
        )

        self.assertFalse(adapted["available"])
        self.assertFalse(adapted["annualAvailable"])
        self.assertIn("nTruePositiveLatent", adapted["missingFields"])

    def test_expected_value_close_to_agent_based_mean(self) -> None:
        ev = _intervention_totals(self.ev["eventLedger"])
        abm_mean = (
            _long_to_wide(self.abm["eventLedger"]["replicateTotals"])
            .query("arm == 'intervention'")
            .mean(numeric_only=True)
        )

        for event in [
            "screened",
            "true_positive_latent",
            "false_positive",
            "tpt_started_total",
            "tpt_completed_total",
            "infection_effectively_treated_total",
            "active_tb_cases_prevented",
        ]:
            tolerance = max(8.0, 0.45 * max(abs(ev[event]), 1.0))
            self.assertLessEqual(abs(ev[event] - abm_mean[event]), tolerance, event)


def _intervention_totals(ledger: dict) -> pd.Series:
    wide = _long_to_wide(ledger["replicateTotals"])
    return wide[wide["arm"] == "intervention"].iloc[0]


def _long_to_wide(totals: pd.DataFrame) -> pd.DataFrame:
    frame = totals.copy()
    if "replicateSeed" in frame.columns:
        frame["replicateSeed"] = frame["replicateSeed"].fillna("")
    id_cols = [
        col
        for col in [
            "modelType",
            "backend",
            "modelVersion",
            "arm",
            "replicateId",
            "pairedReplicateId",
            "replicateSeed",
            "valueType",
            "screeningWindow",
            "followUpHorizon",
        ]
        if col in frame.columns
    ]
    return frame.pivot(
        index=id_cols,
        columns="eventName",
        values="value",
    ).reset_index()


if __name__ == "__main__":
    unittest.main()
