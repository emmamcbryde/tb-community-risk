from __future__ import annotations

import ast
from io import BytesIO
from pathlib import Path
import unittest

from openpyxl import load_workbook

from app.parameter_workspace import (
    HIGHER_BURDEN_DALY_OUTCOME_PRESET,
    PARAMETER_GROUPS,
    PRIMARY_DALY_OUTCOME_PRESET,
    apply_parameter_workspace,
    build_parameter_workspace,
    reset_all_parameters,
    reset_parameter_group,
    unified_default_session_state,
    validate_parameter_workspace,
)
from app.results_workbook import build_results_workbook
from engine.apy.working_defaults import (
    UNIFIED_WORKING_DEFAULT_LABEL,
    UNIFIED_WORKING_DEFAULT_PRESET_ID,
    build_unified_working_default_preset,
)


ROOT = Path(__file__).resolve().parents[1]


class APYWorkingDefaultsTests(unittest.TestCase):
    def test_unified_preset_loads_all_components_atomically(self) -> None:
        state = unified_default_session_state()

        config = state["config"]
        econ = state["economics_config"]
        workspace = state["parameter_workspace"]

        self.assertEqual(config["workingDefaultPresetId"], UNIFIED_WORKING_DEFAULT_PRESET_ID)
        self.assertEqual(config["populationPresetId"], "apy_demonstration")
        self.assertEqual(config["testType"], "IGRA")
        self.assertEqual(config["regimen"], "3HP")
        self.assertEqual(config["screeningWindowYears"], 3)
        self.assertEqual(config["followUpHorizonYears"], 20)
        self.assertEqual(config["analysisMethod"], "expected_value")
        self.assertEqual(econ["metadata"]["presetName"], "Dale 2019 AUD working defaults")
        self.assertEqual(econ["metadata"]["targetCurrency"], "AUD")
        self.assertEqual(econ["metadata"]["targetPriceYear"], "2019")
        self.assertEqual(econ["dalyAssumptions"]["outcomePreset"], PRIMARY_DALY_OUTCOME_PRESET)
        self.assertEqual(workspace["presetLabel"], UNIFIED_WORKING_DEFAULT_LABEL)
        self.assertEqual(set(workspace["groups"]), set(PARAMETER_GROUPS))

    def test_parameter_groups_and_changed_from_default_tracking(self) -> None:
        state = unified_default_session_state()
        rows = state["parameter_workspace"]["rows"]
        edited = [dict(row) for row in rows]
        population_row = next(row for row in edited if row["parameterId"] == "demography.population_size")
        population_row["currentValue"] = 12345

        config, econ = apply_parameter_workspace(state["config"], state["economics_config"], edited)
        workspace = build_parameter_workspace(config, econ)

        self.assertEqual(config["N"], 12345)
        changed_ids = {row["parameterId"] for row in workspace["rows"] if row["changedFromDefault"]}
        self.assertIn("demography.population_size", changed_ids)
        self.assertIn("Demography", {row["group"] for row in workspace["rows"]})
        self.assertIn("TB epidemiology", {row["group"] for row in workspace["rows"]})
        self.assertIn("Health-service provision and intervention", {row["group"] for row in workspace["rows"]})
        self.assertIn("Costs and outcomes", {row["group"] for row in workspace["rows"]})
        self.assertIn("Analysis settings", {row["group"] for row in workspace["rows"]})

    def test_group_reset_and_reset_all_restore_defaults(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        next(row for row in rows if row["parameterId"] == "demography.population_size")["currentValue"] = 999
        next(row for row in rows if row["parameterId"] == "service.coverage")["currentValue"] = 0.75

        demography_reset = reset_parameter_group(rows, "Demography")
        self.assertEqual(
            next(row for row in demography_reset if row["parameterId"] == "demography.population_size")["currentValue"],
            next(row for row in demography_reset if row["parameterId"] == "demography.population_size")["defaultValue"],
        )
        self.assertEqual(next(row for row in demography_reset if row["parameterId"] == "service.coverage")["currentValue"], 0.75)

        all_reset = reset_all_parameters(rows)
        self.assertTrue(all(row["currentValue"] == row["defaultValue"] for row in all_reset))

    def test_dale_costs_and_local_sa_health_costs_are_populated(self) -> None:
        econ = build_unified_working_default_preset()["economicsConfig"]
        items = {item["costItemId"]: item for item in econ["costItems"]}

        self.assertEqual(items["test_igra"]["originalCost"], 113.48)
        self.assertEqual(items["test_tst"]["originalCost"], 116.07)
        self.assertEqual(items["regimen_3hp"]["originalCost"], 165.5072)
        self.assertEqual(items["regimen_4r"]["originalCost"], 123.3172)
        self.assertEqual(items["regimen_3hr"]["originalCost"], 134.2272)
        self.assertEqual(items["regimen_6h"]["originalCost"], 187.7508)
        self.assertEqual(items["regimen_9h"]["originalCost"], 254.8544)
        self.assertEqual(items["active_tb_disease"]["originalCost"], 19079.60)
        self.assertEqual(items["program_setup"]["originalCost"], 0.0)
        self.assertEqual(items["program_running"]["originalCost"], 0.0)
        self.assertEqual(econ["localSAHealthPathwayCosts"]["returnForResults"]["value"], 50.0)
        self.assertEqual(econ["localSAHealthPathwayCosts"]["clinicalReview"]["value"], 50.0)
        self.assertEqual(econ["localSAHealthPathwayCosts"]["activeTBExclusionWorkup"]["value"], 150.0)
        self.assertEqual(items["return_for_results"]["originalCost"], 50.0)
        self.assertEqual(items["clinical_review"]["originalCost"], 50.0)
        self.assertEqual(items["active_tb_exclusion_workup"]["originalCost"], 150.0)
        self.assertEqual(items["travel_outreach_staff_support"]["originalCost"], 0.0)

    def test_local_pathway_overrides_update_authoritative_cost_items(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        next(row for row in rows if row["parameterId"] == "cost.return_for_results")["currentValue"] = 75.0

        _, econ = apply_parameter_workspace(state["config"], state["economics_config"], rows)
        item = next(item for item in econ["costItems"] if item["costItemId"] == "return_for_results")

        self.assertEqual(item["originalCost"], 75.0)
        self.assertEqual(econ["localSAHealthPathwayCosts"]["returnForResults"]["value"], 75.0)

    def test_discount_overrides_update_authoritative_profiles(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        next(row for row in rows if row["parameterId"] == "analysis.cost_discount")["currentValue"] = 0.1
        next(row for row in rows if row["parameterId"] == "analysis.health_discount")["currentValue"] = 0.02
        next(row for row in rows if row["parameterId"] == "analysis.comparison_discount")["currentValue"] = 0.01

        _, econ = apply_parameter_workspace(state["config"], state["economics_config"], rows)
        profiles = econ["discounting"]["profiles"]

        self.assertEqual(profiles["primary"]["costRate"], 0.1)
        self.assertEqual(profiles["primary"]["healthRate"], 0.02)
        self.assertEqual(profiles["comparison"]["costRate"], 0.01)
        self.assertEqual(profiles["comparison"]["healthRate"], 0.01)

    def test_risk_factor_overrides_are_structured_probabilities(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        diabetes = next(row for row in rows if row["parameterId"] == "demography.risk.diabetes")

        self.assertEqual(diabetes["editableType"], "probability")
        diabetes["currentValue"] = 0.9
        config, _ = apply_parameter_workspace(state["config"], state["economics_config"], rows)
        self.assertEqual(config["riskPrev"]["diabetes"], 0.9)

        diabetes["currentValue"] = 1.2
        validation = validate_parameter_workspace(rows)
        self.assertFalse(validation["isValid"])
        self.assertIn("demography.risk.diabetes", {row["parameterId"] for row in validation["messages"]})

    def test_age_distribution_override_updates_authoritative_table(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        next(row for row in rows if row["parameterId"] == "demography.age.0_4")["currentValue"] = 0.1
        next(row for row in rows if row["parameterId"] == "demography.age.5_14")["currentValue"] = 0.2
        next(row for row in rows if row["parameterId"] == "demography.age.15_plus")["currentValue"] = 0.7

        config, _ = apply_parameter_workspace(state["config"], state["economics_config"], rows)

        self.assertEqual(config["ageDistributionFile"], "")
        self.assertEqual(
            config["ageDistributionTable"],
            [
                {"age": "0-4", "proportion": 0.1},
                {"age": "5-14", "proportion": 0.2},
                {"age": "15+", "proportion": 0.7},
            ],
        )

    def test_age_distribution_override_must_sum_to_one(self) -> None:
        workspace = unified_default_session_state()["parameter_workspace"]
        rows = [dict(row) for row in workspace["rows"]]
        next(row for row in rows if row["parameterId"] == "demography.age.0_4")["currentValue"] = 0.1
        next(row for row in rows if row["parameterId"] == "demography.age.5_14")["currentValue"] = 0.2
        next(row for row in rows if row["parameterId"] == "demography.age.15_plus")["currentValue"] = 0.6

        validation = validate_parameter_workspace(rows)

        self.assertFalse(validation["isValid"])
        self.assertIn("demography.age", {row["parameterId"] for row in validation["messages"]})

    def test_editable_rows_have_operational_status(self) -> None:
        workspace = unified_default_session_state()["parameter_workspace"]
        editable_rows = [row for row in workspace["rows"] if row["editableType"] != "read_only"]

        self.assertTrue(editable_rows)
        for row in editable_rows:
            self.assertIn(
                row["operationalStatus"],
                {"authoritative_model_input", "authoritative_economic_input"},
                row["parameterId"],
            )

    def test_advanced_fields_are_rendered_as_editable_controls(self) -> None:
        text = (ROOT / "pages" / "0_Start.py").read_text(encoding="utf-8")

        self.assertIn("parameter_advanced_editor_", text)
        self.assertNotIn("st.dataframe(advanced_rows", text)

    def test_start_page_exposes_provisional_default_run_acknowledgement(self) -> None:
        text = (ROOT / "pages" / "0_Start.py").read_text(encoding="utf-8")

        self.assertIn("Recent versus remote LTBI assumption remains provisional", text)
        self.assertIn("Run provisional working defaults", text)
        self.assertIn("0% compatibility", text)

    def test_start_and_strategy_pages_show_resolved_demographic_profile(self) -> None:
        start_text = (ROOT / "pages" / "0_Start.py").read_text(encoding="utf-8")
        strategy_text = (ROOT / "pages" / "1_Scenario.py").read_text(encoding="utf-8")

        self.assertIn("APY demographic profile", start_text)
        self.assertIn("APY demographic profile", strategy_text)
        self.assertIn("demographic_summary_rows", start_text)
        self.assertIn("demographic_summary_rows", strategy_text)
        self.assertIn("Restore APY demographic defaults", strategy_text)

    def test_start_page_wraps_arrow_safe_tables_in_streamlit_dataframe(self) -> None:
        text = (ROOT / "pages" / "0_Start.py").read_text(encoding="utf-8")
        tree = ast.parse(text)
        for parent in ast.walk(tree):
            for child in ast.iter_child_nodes(parent):
                child.parent = parent
        arrow_calls = [
            node
            for node in ast.walk(tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Name)
            and node.func.id == "arrow_safe_dataframe"
        ]

        self.assertGreaterEqual(len(arrow_calls), 2)
        self.assertTrue(all(not call.keywords for call in arrow_calls))
        for call in arrow_calls:
            self.assertIsInstance(call.parent, ast.Call)
            parent_func = call.parent.func
            self.assertIsInstance(parent_func, ast.Attribute)
            self.assertEqual(parent_func.attr, "dataframe")

    def test_arrow_safe_dataframe_is_never_passed_streamlit_display_keywords(self) -> None:
        display_keywords = {"use_container_width", "hide_index", "width", "height", "column_config"}
        offenders = []
        for path in [*ROOT.glob("app/**/*.py"), *ROOT.glob("pages/**/*.py")]:
            tree = ast.parse(path.read_text(encoding="utf-8"))
            for parent in ast.walk(tree):
                for child in ast.iter_child_nodes(parent):
                    child.parent = parent
            for node in ast.walk(tree):
                if (
                    isinstance(node, ast.Call)
                    and isinstance(node.func, ast.Name)
                    and node.func.id == "arrow_safe_dataframe"
                ):
                    bad = sorted({keyword.arg for keyword in node.keywords if keyword.arg in display_keywords})
                    if bad:
                        offenders.append((path.relative_to(ROOT).as_posix(), bad))
        self.assertEqual(offenders, [])

    def test_programme_setup_base_and_new_programme_options_apply(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        setup = next(row for row in rows if row["parameterId"] == "cost.program_setup_preset")
        setup["currentValue"] = "New-programme implementation scenario"

        _, econ = apply_parameter_workspace(state["config"], state["economics_config"], rows)
        item = next(item for item in econ["costItems"] if item["costItemId"] == "program_setup")

        self.assertEqual(item["originalCost"], 500000.0)
        self.assertEqual(econ["localSAHealthPathwayCosts"]["selectedProgramSetupPreset"], "new_programme_implementation")

    def test_standard_workspace_uses_dalys_and_hides_qalys(self) -> None:
        workspace = unified_default_session_state()["parameter_workspace"]
        labels = "\n".join(str(row["label"]) for row in workspace["rows"])
        self.assertIn("Outcome preset", labels)
        self.assertIn("Active-TB disability weight", labels)
        self.assertNotIn("QALY", labels)

    def test_standard_workspace_uses_plain_language_targeting_labels(self) -> None:
        workspace = unified_default_session_state()["parameter_workspace"]
        targeting = next(row for row in workspace["rows"] if row["parameterId"] == "service.targeting")

        self.assertEqual(targeting["currentValue"], "Prioritise people most likely to avoid active TB")
        self.assertNotEqual(targeting["currentValue"], "prevent")
        self.assertIn("No risk-based prioritisation", targeting["selectOptions"])

    def test_invalid_select_override_is_rejected(self) -> None:
        workspace = unified_default_session_state()["parameter_workspace"]
        targeting = next(row for row in workspace["rows"] if row["parameterId"] == "service.targeting")
        targeting["currentValue"] = "invented targeting option"

        validation = validate_parameter_workspace(workspace["rows"])

        self.assertFalse(validation["isValid"])
        self.assertIn("service.targeting", {row["parameterId"] for row in validation["messages"]})

    def test_higher_burden_option_adds_post_tb_dalys_without_combining_qalys(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        outcome = next(row for row in rows if row["parameterId"] == "outcome.daly_preset")
        outcome["currentValue"] = HIGHER_BURDEN_DALY_OUTCOME_PRESET

        _, econ = apply_parameter_workspace(state["config"], state["economics_config"], rows)
        daly = econ["dalyAssumptions"]

        self.assertTrue(daly["includePostTBSequelae"])
        self.assertEqual(daly["postTBDALYsPerActiveTBCase"]["value"], 5.8)
        self.assertEqual(daly["standardOutcomeMetric"], "DALYs")
        self.assertFalse(daly["qalyStandardInterfaceVisible"])

    def test_expected_outcomes_mode_does_not_require_simulation_count(self) -> None:
        workspace = unified_default_session_state()["parameter_workspace"]
        n_reps = next(row for row in workspace["rows"] if row["parameterId"] == "analysis.n_reps")
        n_reps["currentValue"] = ""

        validation = validate_parameter_workspace(workspace["rows"])
        self.assertTrue(validation["isValid"])

    def test_quick_simulation_is_labelled_preview_and_sets_five_replicates(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        next(row for row in rows if row["parameterId"] == "analysis.method")["currentValue"] = "Simulated community variation"
        next(row for row in rows if row["parameterId"] == "analysis.simulation_mode")["currentValue"] = "Quick preview: 5 simulations"

        config, _ = apply_parameter_workspace(state["config"], state["economics_config"], rows)

        self.assertEqual(config["analysisMethod"], "agent_based")
        self.assertEqual(config["simulationMode"], "quick_preview")
        self.assertEqual(config["nReps"], 5)
        self.assertIn("preview", config["simulationModeLabel"].lower())

    def test_workbook_exports_defaults_and_overrides(self) -> None:
        state = unified_default_session_state()
        rows = [dict(row) for row in state["parameter_workspace"]["rows"]]
        next(row for row in rows if row["parameterId"] == "service.coverage")["currentValue"] = 0.4
        config, econ = apply_parameter_workspace(state["config"], state["economics_config"], rows)

        payload = build_results_workbook(
            config=config,
            economics_config=econ,
            bundle={"metadata": {}, "headline": {}, "technical": {}},
        )
        wb = load_workbook(BytesIO(payload), read_only=True)
        self.assertIn("Parameter_workspace", wb.sheetnames)
        self.assertIn("Parameter_defaults", wb.sheetnames)
        self.assertIn("Parameter_overrides", wb.sheetnames)
        rows = list(wb["Parameter_overrides"].iter_rows(values_only=True))
        headers = rows[0]
        data = [dict(zip(headers, row)) for row in rows[1:]]
        coverage = next(row for row in data if row["parameterId"] == "service.coverage")
        self.assertEqual(coverage["currentValue"], 0.4)
        self.assertTrue(coverage["changedFromDefault"])
        wb.close()


if __name__ == "__main__":
    unittest.main()
