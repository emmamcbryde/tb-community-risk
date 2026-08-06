from __future__ import annotations

import json
import py_compile
import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from app.results_workbook import build_results_workbook
from adapters.serialization import to_json_like
from engine.apy.decision_analysis import (
    DECISION_ANALYSIS_CONTRACT_VERSION,
    build_scenario_config,
    run_scenario_comparison,
)
from engine.apy.sensitivity import (
    SENSITIVITY_CONTRACT_VERSION,
    THRESHOLD_CONTRACT_VERSION,
    crossings_from_evaluated_grid,
    load_sensitivity_specs,
    monotonicity_diagnostic,
    run_one_way_sensitivity,
    run_threshold_analysis,
)
from engine.apy.early_review import EARLY_REVIEW_CONTRACT_VERSION, run_early_screening_review
from tests.test_apy_event_ledger_economics import _reviewed_epi, _synthetic_econ


class ApyDecisionAnalysisScenarioComparisonTests(unittest.TestCase):
    def test_expected_value_scenario_comparison_contract_and_repeatability(self) -> None:
        base = _reviewed_epi({"N": 80, "screeningStrategy": "prevent", "screenCoverage": 0.25})
        scenarios = [
            {"scenarioId": "igra_prevent", "label": "IGRA prevent", "changes": {"test": "IGRA"}},
            {"scenarioId": "tst_prevent", "label": "TST prevent", "changes": {"test": "TST"}},
        ]

        first = run_scenario_comparison(base, _synthetic_econ(), scenarios, model_type="expected_value")
        second = run_scenario_comparison(base, _synthetic_econ(), scenarios, model_type="expected_value")

        self.assertEqual(first["contractVersion"], DECISION_ANALYSIS_CONTRACT_VERSION)
        self.assertTrue(first["validation"]["isValid"])
        self.assertEqual(first["scenarios"][0]["valueType"], "expected")
        self.assertEqual(first["scenarioSummaries"][0]["configurationHash"], second["scenarioSummaries"][0]["configurationHash"])
        self.assertEqual(
            first["scenarioSummaries"][0]["active_tb_cases_prevented"],
            second["scenarioSummaries"][0]["active_tb_cases_prevented"],
        )
        json.dumps(to_json_like(first))

    def test_scenario_changes_are_validated_adapters_not_arbitrary_paths(self) -> None:
        base = _reviewed_epi({"N": 20})

        with self.assertRaisesRegex(ValueError, "Unsupported or unvalidated"):
            build_scenario_config(
                base,
                {"scenarioId": "bad", "changes": {"ltbiStateAssumptions.status": "configured_reviewed"}},
            )

        cfg, changed, _ = build_scenario_config(
            base,
            {"scenarioId": "good", "changes": {"testType": "TST", "tstSpecificityBCG": 0.5}},
        )
        self.assertEqual(cfg["testType"], "TST")
        self.assertEqual(cfg["testCharacteristics"]["TST"]["specificityBCG"], 0.5)
        self.assertEqual(changed["tstSpecificityBCG"], 0.5)

    def test_incomplete_economics_does_not_create_strategy_economic_comparison(self) -> None:
        base = _reviewed_epi({"N": 60})
        econ = _synthetic_econ()
        for item in econ["costItems"]:
            if item["costItemId"] == "test_igra":
                item["originalCost"] = None
                item["conversionStatus"] = "unresolved_source_price_year"
                item["convertedTargetYearCost"] = None
        result = run_scenario_comparison(
            base,
            econ,
            [{"scenarioId": "igra", "changes": {"test": "IGRA"}}],
            model_type="expected_value",
        )

        summary = result["scenarioSummaries"][0]
        self.assertIsNone(summary["incrementalCost"])
        self.assertEqual(summary["icerClassification"], "incomplete / not calculated")
        self.assertTrue(result["validation"]["warnings"])

    def test_stochastic_strategy_comparison_preserves_baseline_cohort_fingerprint(self) -> None:
        base = _reviewed_epi({"N": 120, "screenCoverage": 0.25, "screeningStrategy": "prevent"})
        scenarios = [
            {"scenarioId": "igra", "changes": {"test": "IGRA"}},
            {"scenarioId": "tst", "changes": {"test": "TST"}},
        ]

        result = run_scenario_comparison(
            base,
            _synthetic_econ(),
            scenarios,
            model_type="agent_based",
            n_reps=3,
            master_seed=123,
        )

        self.assertTrue(result["validation"]["isValid"])
        self.assertTrue(result["pairedComparisons"][0]["pairedCohortFingerprintMatch"])
        self.assertEqual(result["scenarios"][0]["valueType"], "simulated_count")
        self.assertIsNotNone(result["scenarios"][0]["seed"])


class ApyDecisionAnalysisSensitivityTests(unittest.TestCase):
    def test_default_sensitivity_specs_are_unresolved_without_invented_ranges(self) -> None:
        specs = load_sensitivity_specs()

        self.assertTrue(specs)
        self.assertTrue(all(spec["lowValue"] is None and spec["highValue"] is None for spec in specs))

        result = run_one_way_sensitivity(
            _reviewed_epi({"N": 40}),
            _synthetic_econ(),
            specs[:1],
            ["active_tb_cases_prevented"],
        )

        self.assertEqual(result["contractVersion"], SENSITIVITY_CONTRACT_VERSION)
        self.assertFalse(result["results"][0]["complete"])
        self.assertIn("Reviewed low/high range required", result["results"][0]["unresolvedReason"])

    def test_one_way_sensitivity_runs_only_explicit_supplied_ranges(self) -> None:
        spec = {
            "parameterId": "epi.tpt_initiation.synthetic",
            "label": "Synthetic TPT initiation",
            "adapter": "pStartTPT",
            "baseValue": 0.85,
            "lowValue": 0.0,
            "highValue": 1.0,
            "unit": "probability",
            "scale": "probability",
            "source": "Synthetic unit-test range",
            "reviewStatus": "configured_reviewed",
            "provisional": False,
        }

        result = run_one_way_sensitivity(
            _reviewed_epi({"N": 60, "screeningStrategy": "random"}),
            _synthetic_econ(),
            [spec],
            ["tpt_started_total", "active_tb_cases_prevented", "incrementalCost"],
        )

        rows = result["results"]
        self.assertTrue(all(row["complete"] for row in rows))
        starts = next(row for row in rows if row["outcome"] == "tpt_started_total")
        self.assertEqual(starts["lowOutcome"], 0)
        self.assertGreater(starts["highOutcome"], starts["baseOutcome"])
        self.assertTrue(result["tornadoRows"])

    def test_sensitivity_cost_adapter_updates_authoritative_cost_item(self) -> None:
        spec = {
            "parameterId": "cost.program_running.synthetic",
            "label": "Synthetic running cost",
            "adapter": "programRunningCost",
            "baseValue": 0,
            "lowValue": 0,
            "highValue": 100,
            "unit": "AUD",
            "scale": "monetary",
            "source": "Synthetic unit-test range",
            "reviewStatus": "configured_reviewed",
            "provisional": False,
        }

        result = run_one_way_sensitivity(
            _reviewed_epi({"N": 50, "screeningWindowYears": 2}),
            _synthetic_econ({"program_running": 0}),
            [spec],
            ["incrementalCost"],
        )

        row = result["results"][0]
        self.assertTrue(row["complete"])
        self.assertGreater(row["highOutcome"], row["lowOutcome"])

    def test_threshold_analysis_preserves_grid_and_requires_valid_threshold_for_nmb(self) -> None:
        spec = {
            "parameterId": "epi.tpt_initiation.synthetic",
            "adapter": "pStartTPT",
            "label": "Synthetic TPT initiation",
        }

        unavailable = run_threshold_analysis(
            _reviewed_epi({"N": 40}),
            _synthetic_econ({"threshold": None}),
            spec,
            "nmb",
            {"low": 0, "high": 1, "gridPoints": 5},
        )
        self.assertFalse(unavailable["validation"]["isValid"])

        result = run_threshold_analysis(
            _reviewed_epi({"N": 20}),
            _synthetic_econ({"threshold": 50000}),
            spec,
            "active_tb_cases_prevented",
            {"low": 0, "high": 1, "gridPoints": 3, "target": -999},
        )

        self.assertEqual(result["contractVersion"], THRESHOLD_CONTRACT_VERSION)
        self.assertEqual(len(result["grid"]), 3)
        self.assertEqual(result["crossingCount"], 0)
        self.assertIn("isMonotonic", result["monotonicity"])

    def test_nonmonotonic_and_multiple_crossing_grid_is_detected(self) -> None:
        grid = [
            {"parameterValue": 0, "difference": -1, "complete": True},
            {"parameterValue": 1, "difference": 1, "complete": True},
            {"parameterValue": 2, "difference": -1, "complete": True},
            {"parameterValue": 3, "difference": 1, "complete": True},
        ]

        self.assertFalse(monotonicity_diagnostic(grid)["isMonotonic"])
        crossings = crossings_from_evaluated_grid(grid)
        self.assertEqual(len(crossings), 3)
        self.assertTrue(all(crossing["bracketed"] for crossing in crossings))


class ApyDecisionAnalysisEarlyReviewTests(unittest.TestCase):
    def test_no_prior_means_no_posterior(self) -> None:
        result = run_early_screening_review(
            _reviewed_epi({"N": 100}),
            _synthetic_econ(),
            {"screenedToDate": 10, "testPositiveToDate": 1, "plannedTotalScreened": 30},
        )

        self.assertFalse(result["validation"]["isValid"])
        self.assertIn("prior", {issue["field"] for issue in result["validation"]["errors"]})

    def test_posterior_weights_normalize_and_low_high_yield_shift(self) -> None:
        base = _reviewed_epi({"N": 30, "screeningStrategy": "random", "screenCoverage": 0.3})
        common = {
            "plannedTotalScreened": 9,
            "screenedToDate": 6,
            "eligiblePopulation": 30,
            "prior": {"type": "beta", "mean": 0.1, "effectiveSampleSize": 20, "source": "Synthetic prior"},
            "prevalenceGrid": [0.03, 0.1, 0.17],
        }

        with patch("engine.apy.early_review._run_coverage_projection", _fake_prevalence_projection):
            low = run_early_screening_review(base, _synthetic_econ(), {**common, "testPositiveToDate": 0})
            high = run_early_screening_review(base, _synthetic_econ(), {**common, "testPositiveToDate": 5})

        self.assertEqual(low["contractVersion"], EARLY_REVIEW_CONTRACT_VERSION)
        self.assertAlmostEqual(sum(row["posteriorWeight"] for row in low["priorPosteriorTable"]), 1.0)
        self.assertLess(low["posterior"]["summary"]["mean"], low["prior"]["summary"]["mean"])
        self.assertGreater(high["posterior"]["summary"]["mean"], high["prior"]["summary"]["mean"])

    def test_false_positives_possible_at_zero_prevalence_when_specificity_below_one(self) -> None:
        base = _reviewed_epi({"N": 20, "testType": "IGRA", "testSpecificity": 0.8, "screeningStrategy": "random"})
        review = _review_input({"prevalenceGrid": [0.0, 0.1], "testPositiveToDate": 1})

        result = run_early_screening_review(base, None, review)
        zero = next(row for row in result["priorPosteriorTable"] if row["prevalence"] == 0.0)

        self.assertGreater(zero["predictedPositiveProbability"], 0)

    def test_perfect_test_approximates_prevalence_likelihood_for_random_screening(self) -> None:
        base = _reviewed_epi(
            {
                "N": 20,
                "testSensitivity": 1.0,
                "testSpecificity": 1.0,
                "screeningStrategy": "random",
                "screeningWindowYears": 0.01,
            }
        )
        with patch("engine.apy.early_review._run_coverage_projection", _fake_prevalence_projection):
            result = run_early_screening_review(base, None, _review_input({"prevalenceGrid": [0.02, 0.1]}))

        probs = {row["prevalence"]: row["predictedPositiveProbability"] for row in result["priorPosteriorTable"]}
        self.assertLess(probs[0.02], probs[0.1])

    def test_tst_bcg_composition_changes_predicted_false_positive_yield(self) -> None:
        high_bcg = _reviewed_epi(
            {
                "N": 20,
                "testType": "TST",
                "screeningStrategy": "random",
                "riskPrev": {"BCG": 1.0},
            }
        )
        low_bcg = _reviewed_epi(
            {
                "N": 20,
                "testType": "TST",
                "screeningStrategy": "random",
                "riskPrev": {"BCG": 0.0},
            }
        )
        review = _review_input({"prevalenceGrid": [0.0]})

        high = run_early_screening_review(high_bcg, None, review)
        low = run_early_screening_review(low_bcg, None, review)

        self.assertGreater(
            high["priorPosteriorTable"][0]["predictedPositiveProbability"],
            low["priorPosteriorTable"][0]["predictedPositiveProbability"],
        )

    def test_targeted_screening_yield_can_differ_from_population_prevalence(self) -> None:
        base = _reviewed_epi({"N": 30, "screeningStrategy": "prevent"})
        with patch("engine.apy.early_review._run_coverage_projection", _fake_targeted_projection):
            result = run_early_screening_review(base, None, _review_input({"prevalenceGrid": [0.1]}))

        predicted = result["priorPosteriorTable"][0]["predictedPositiveProbability"]
        self.assertNotAlmostEqual(predicted, 0.1)

    def test_stop_versus_continue_retains_completed_screening_and_adds_remaining(self) -> None:
        base = _reviewed_epi({"N": 30, "screeningStrategy": "random"})
        with patch("engine.apy.early_review._run_coverage_projection", _fake_prevalence_projection):
            result = run_early_screening_review(
                base,
                _synthetic_econ({"threshold": 50000}),
                _review_input(
                    {
                        "screenedToDate": 6,
                        "plannedTotalScreened": 12,
                        "eligiblePopulation": 30,
                        "prevalenceGrid": [0.08, 0.12],
                    }
                ),
            )

        self.assertTrue(result["timingApproximation"])
        first = result["projection"][0]
        self.assertGreater(first["additionalPeopleScreened"], 0)
        self.assertAlmostEqual(first["additionalCoverage"], (12 - 6) / 30)

    def test_screened_equals_planned_total_has_no_remaining_screening(self) -> None:
        with patch("engine.apy.early_review._run_coverage_projection", _fake_prevalence_projection):
            result = run_early_screening_review(
                _reviewed_epi({"N": 30, "screeningStrategy": "random"}),
                None,
                _review_input({"screenedToDate": 12, "plannedTotalScreened": 12, "eligiblePopulation": 30}),
            )

        self.assertAlmostEqual(result["projection"][0]["additionalPeopleScreened"], 0.0)


class ApyDecisionAnalysisPageSmokeTests(unittest.TestCase):
    def test_decision_analysis_page_compiles(self) -> None:
        py_compile.compile("pages/5_Decision_Analysis.py", doraise=True)


class ApyDecisionAnalysisWorkbookTests(unittest.TestCase):
    def test_workbook_exports_decision_analysis_values_from_bundle(self) -> None:
        comparison = run_scenario_comparison(
            _reviewed_epi({"N": 20, "screenCoverage": 0}),
            None,
            [{"scenarioId": "zero", "label": "Zero screening", "changes": {"screenCoverage": 0}}],
            model_type="expected_value",
        )
        payload = build_results_workbook(
            config=_reviewed_epi({"N": 20}),
            bundle={"metadata": {}, "headline": {}, "technical": {}, "downloads": {}},
            decision_analysis_results={"scenarioComparison": comparison},
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)

        self.assertIn("Decision_scenarios", wb.sheetnames)
        self.assertIn("Scenario_comparison", wb.sheetnames)
        rows = list(wb["Scenario_comparison"].iter_rows(values_only=True))
        exported = dict(zip(rows[0], rows[1]))
        expected = comparison["scenarioSummaries"][0]

        self.assertEqual(exported["scenarioId"], expected["scenarioId"])
        self.assertEqual(exported["screened"], expected["screened"])
        self.assertIsNone(exported["incrementalCost"])
        wb.close()

    def test_workbook_exports_early_review_prior_and_projection(self) -> None:
        with patch("engine.apy.early_review._run_coverage_projection", _fake_prevalence_projection):
            early = run_early_screening_review(
                _reviewed_epi({"N": 30}),
                None,
                _review_input({"screenedToDate": 6, "plannedTotalScreened": 12, "eligiblePopulation": 30}),
            )
        payload = build_results_workbook(
            config=_reviewed_epi({"N": 30}),
            bundle={"metadata": {}, "headline": {}, "technical": {}, "downloads": {}},
            decision_analysis_results={"earlyReview": early},
        )
        wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)

        self.assertIn("Early_review_prior_posterior", wb.sheetnames)
        self.assertIn("Early_review_projection", wb.sheetnames)
        rows = list(wb["Early_review_prior_posterior"].iter_rows(values_only=True))
        exported = dict(zip(rows[0], rows[1]))
        self.assertEqual(exported["prevalence"], early["priorPosteriorTable"][0]["prevalence"])
        wb.close()


def _review_input(overrides: dict | None = None) -> dict:
    base = {
        "reviewId": "synthetic_review",
        "scenarioId": "synthetic_scenario",
        "screenedToDate": 10,
        "testPositiveToDate": 1,
        "plannedTotalScreened": 30,
        "eligiblePopulation": 100,
        "prior": {
            "type": "discrete_grid",
            "weights": [0.5, 0.5],
            "source": "Synthetic unit-test prior",
        },
        "prevalenceGrid": [0.05, 0.15],
        "prevalenceThresholds": [0.1],
    }
    if overrides:
        base.update(overrides)
        if "prevalenceGrid" in overrides and "prior" not in overrides:
            n = len(overrides["prevalenceGrid"])
            base["prior"] = {"type": "discrete_grid", "weights": [1 / n] * n, "source": "Synthetic unit-test prior"}
    return base


def _fake_prevalence_projection(base_config, economics_config, prevalence, coverage, scenario_id):
    screened = float(base_config.get("N", 100)) * float(coverage)
    positives = screened * (0.02 + 0.8 * float(prevalence))
    return {
        "comparison": None,
        "summary": {
            "screened": screened,
            "test_positive_total": positives,
            "true_positive_recent": positives * 0.4,
            "true_positive_remote": positives * 0.4,
            "false_positive": positives * 0.2,
            "tpt_started_total": positives * 0.8,
            "tpt_completed_total": positives * 0.6,
            "infection_effectively_treated_total": positives * 0.5,
            "active_tb_cases_prevented": positives * 0.1,
            "incrementalCost": positives * 10 if economics_config is not None else None,
            "dalysAverted": positives * 0.01 if economics_config is not None else None,
            "nmb": positives if economics_config is not None else None,
        },
    }


def _fake_targeted_projection(base_config, economics_config, prevalence, coverage, scenario_id):
    out = _fake_prevalence_projection(base_config, economics_config, prevalence, coverage, scenario_id)
    out["summary"]["test_positive_total"] = out["summary"]["screened"] * 0.25
    return out


if __name__ == "__main__":
    unittest.main()
