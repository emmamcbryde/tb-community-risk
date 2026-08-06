from __future__ import annotations

from copy import deepcopy
from io import BytesIO
import unittest

from openpyxl import load_workbook

from app.results_workbook import build_results_workbook
from engine.apy.config import build_default_config
from engine.apy.costing import normalise_cost_table
from engine.apy.economics import build_default_economics_config
from engine.apy.evidence import (
    assess_apy_reference_readiness,
    load_apy_evidence_registry,
    validate_cost_item_evidence,
)
from engine.apy.event_ledger_economics import run_event_ledger_health_economics
from engine.apy.ltbi_state import apply_ltbi_state_edit
from engine.apy.validation import collect_validation_issues
from tests.test_apy_event_ledger_economics import _synthetic_econ, _toy_ledger


class ApyEvidenceRegistryTests(unittest.TestCase):
    def test_registry_loads_required_assumptions(self) -> None:
        rows = load_apy_evidence_registry()
        ids = {row["assumptionId"] for row in rows}

        self.assertIn("epi.baseline_recent_ltbi_proportion", ids)
        self.assertIn("cost.test_igra", ids)
        self.assertIn("daly.active_tb_disability_weight", ids)
        self.assertIn("threshold.gdp_per_capita", ids)

    def test_same_year_costs_still_require_source_and_review_status(self) -> None:
        item = _valid_cost_item("test_igra")
        item["sourceCitation"] = ""
        row = _reviewed_registry_row("cost.test_igra")

        errors = validate_cost_item_evidence(item, row)

        self.assertIn("Source citation is missing.", errors)

        item["sourceCitation"] = "Synthetic reviewed source"
        row["reviewStatus"] = "unreviewed_repository_input"
        errors = validate_cost_item_evidence(item, row)
        self.assertIn("Cost evidence requires reviewed numerical status.", errors)

    def test_foreign_currency_costs_cannot_enter_totals_without_conversion(self) -> None:
        config = _synthetic_econ()
        for item in config["costItems"]:
            if item["costItemId"] == "test_igra":
                item["originalCurrency"] = "USD"

        econ = run_event_ledger_health_economics(_toy_ledger(), config)

        self.assertIn(
            "costItems.test_igra.originalCurrency",
            {item["field"] for item in econ["unresolvedInputs"]},
        )
        self.assertFalse(econ["status"]["isComplete"])

    def test_empty_source_and_unreviewed_source_year_prevent_readiness(self) -> None:
        item = _valid_cost_item("test_igra")
        item["sourceYearStatus"] = "inferred"
        row = _reviewed_registry_row("cost.test_igra")

        errors = validate_cost_item_evidence(item, row)

        self.assertIn("Source price year must be explicit or reviewed-inferred.", errors)

    def test_essential_costs_cannot_be_removed_by_generic_exclusion(self) -> None:
        config = _synthetic_econ()
        for item in config["costItems"]:
            if item["costItemId"] == "test_igra":
                item["originalCost"] = None
        config["optionalCostExclusions"] = {
            "test_igra": {
                "status": "reviewed_exclusion",
                "rationale": "Generic exclusion is not enough for an essential selected test cost.",
                "exclusionType": "omitted_nonessential",
            }
        }

        econ = run_event_ledger_health_economics(_toy_ledger(), config)

        self.assertIn("costItems.test_igra", {item["field"] for item in econ["unresolvedInputs"]})
        self.assertFalse(econ["status"]["isComplete"])

    def test_valid_explicit_bundling_prevents_double_counting_conflict(self) -> None:
        registry = _minimal_reviewed_registry()

        readiness = assess_apy_reference_readiness(_reviewed_epi_config(), _reviewed_econ_config(), registry)

        self.assertEqual(readiness["bundlingConflicts"], [])

    def test_circular_or_conflicting_bundling_is_rejected(self) -> None:
        registry = _minimal_reviewed_registry()
        by_id = {row["assumptionId"]: row for row in registry}
        by_id["cost.program_setup"]["inclusionStatus"] = "bundled"
        by_id["cost.program_setup"]["bundledIntoAssumptionId"] = "cost.program_running"
        by_id["cost.program_setup"]["currentValue"] = ""
        by_id["cost.program_running"]["inclusionStatus"] = "bundled"
        by_id["cost.program_running"]["bundledIntoAssumptionId"] = "cost.program_setup"
        by_id["cost.program_running"]["currentValue"] = ""

        readiness = assess_apy_reference_readiness(_reviewed_epi_config(), _reviewed_econ_config(), registry)

        self.assertTrue(readiness["bundlingConflicts"])

    def test_legacy_repository_values_remain_provisional_until_reviewed(self) -> None:
        readiness = assess_apy_reference_readiness(build_default_config(), build_default_economics_config())

        self.assertFalse(readiness["overallClinicianReady"])
        self.assertIn("cost.test_igra", readiness["provisionalAssumptionIds"])

    def test_synthetic_validation_fixture_is_not_apy_reference_evidence(self) -> None:
        rows = load_apy_evidence_registry()
        recent = next(row for row in rows if row["assumptionId"] == "epi.baseline_recent_ltbi_proportion")

        self.assertNotEqual(recent["currentValue"], "0.35")
        self.assertEqual(recent["reviewStatus"], "unresolved")

    def test_reviewed_daly_exclusions_require_rationale(self) -> None:
        registry = _minimal_reviewed_registry()
        row = next(row for row in registry if row["assumptionId"] == "daly.tpt_health_loss")
        row["inclusionStatus"] = "excluded"
        row["reviewStatus"] = "reviewed_exclusion"
        row["notes"] = ""

        readiness = assess_apy_reference_readiness(_reviewed_epi_config(), _reviewed_econ_config(), registry)

        self.assertFalse(readiness["dalyReady"])
        self.assertIn("daly.tpt_health_loss", readiness["unresolvedAssumptionIds"])

    def test_unresolved_recent_ltbi_prevents_overall_clinician_readiness(self) -> None:
        readiness = assess_apy_reference_readiness(build_default_config(), _reviewed_econ_config(), _minimal_reviewed_registry())

        self.assertFalse(readiness["epidemiologyReady"])
        self.assertFalse(readiness["overallClinicianReady"])
        self.assertIn("epi.baseline_recent_ltbi_proportion", readiness["unresolvedAssumptionIds"])

    def test_threshold_absence_does_not_prevent_cost_and_daly_totals(self) -> None:
        config = _synthetic_econ({"threshold": None})
        config["threshold"]["value"] = None

        econ = run_event_ledger_health_economics(_toy_ledger(), config)
        readiness = assess_apy_reference_readiness(_reviewed_epi_config(), config, _minimal_reviewed_registry())

        row = econ["replicateResults"][econ["replicateResults"]["discountRate"] == 0.0].iloc[0]
        self.assertIsNotNone(row["incrementalCost"])
        self.assertIsNotNone(row["dalysAverted"])
        self.assertFalse(readiness["thresholdReady"])

    def test_workbook_evidence_rows_match_registry(self) -> None:
        registry = load_apy_evidence_registry()
        payload = build_results_workbook(
            config=build_default_config(),
            bundle={"metadata": {}, "technical": {}, "headline": {}, "downloads": {}},
            economics_config=build_default_economics_config(),
        )
        wb = load_workbook(BytesIO(payload), read_only=True)
        ws = wb["Assumption_evidence_registry"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assumption_idx = headers.index("assumptionId")
        exported_ids = {
            row[assumption_idx].value
            for row in ws.iter_rows(min_row=2)
            if row[assumption_idx].value
        }
        wb.close()

        self.assertEqual(exported_ids, {row["assumptionId"] for row in registry})

    def test_registry_assessment_does_not_change_economic_calculation(self) -> None:
        econ_config = _synthetic_econ()
        before = run_event_ledger_health_economics(_toy_ledger(), deepcopy(econ_config))
        assess_apy_reference_readiness(_reviewed_epi_config(), econ_config)
        after = run_event_ledger_health_economics(_toy_ledger(), deepcopy(econ_config))

        before_row = before["replicateResults"][before["replicateResults"]["discountRate"] == 0.0].iloc[0]
        after_row = after["replicateResults"][after["replicateResults"]["discountRate"] == 0.0].iloc[0]
        self.assertEqual(before_row["incrementalCost"], after_row["incrementalCost"])
        self.assertEqual(before_row["dalysAverted"], after_row["dalysAverted"])

    def test_existing_clinician_ready_validation_can_use_registry_gate(self) -> None:
        report = collect_validation_issues(
            _reviewed_epi_config(),
            clinician_ready=True,
            economics_config=_reviewed_econ_config(),
            evidence_registry=load_apy_evidence_registry(),
        )

        self.assertFalse(report["isValid"])
        self.assertIn("evidenceRegistry.cost.test_igra", report["fatalFieldNames"])


def _valid_cost_item(cost_item_id: str) -> dict:
    item = next(item for item in normalise_cost_table(_synthetic_econ()["costItems"]) if item["costItemId"] == cost_item_id)
    item["sourceCitation"] = "Synthetic reviewed source"
    item["sourceYearStatus"] = "explicit"
    return item


def _reviewed_registry_row(assumption_id: str) -> dict:
    prefix = assumption_id.split(".")[0]
    category = "epidemiology" if prefix == "epi" else prefix
    return {
        "assumptionId": assumption_id,
        "category": category,
        "description": assumption_id,
        "currentValue": "1",
        "unit": "",
        "originalCurrency": "AUD",
        "targetCurrency": "AUD",
        "originalPriceYear": "2025-26",
        "targetPriceYear": "2025-26",
        "sourceCitation": "Synthetic reviewed source",
        "sourceLocation": "unit test",
        "pageTableItem": "",
        "repositoryPath": "",
        "repositoryVariable": "",
        "derivationMethod": "",
        "inflationIndexId": "",
        "inclusionStatus": "included",
        "reviewStatus": "configured_reviewed",
        "provisional": False,
        "bundledIntoAssumptionId": "",
        "doubleCountingGroup": "",
        "notes": "Synthetic reviewed registry fixture.",
        "unresolvedReason": "",
    }


def _minimal_reviewed_registry() -> list[dict]:
    rows = []
    for assumption_id in [
        "epi.baseline_recent_ltbi_proportion",
        "epi.recent_to_remote_transition_rate",
        "epi.ltbi_prevalence",
        "epi.active_tb_calibration_target_2y",
        "epi.igra_sensitivity",
        "epi.igra_specificity",
        "epi.tst_sensitivity",
        "epi.tst_specificity_bcg",
        "epi.tst_specificity_no_bcg",
        "epi.tpt_initiation",
        "epi.regimen_3hp_completion",
        "epi.regimen_4r_completion",
        "epi.regimen_3hr_completion",
        "epi.regimen_6h_completion",
        "epi.regimen_9h_completion",
        "epi.regimen_3hp_adr_stop",
        "epi.regimen_4r_adr_stop",
        "epi.regimen_3hr_adr_stop",
        "epi.regimen_6h_adr_stop",
        "epi.regimen_9h_adr_stop",
        "epi.regimen_full_efficacy",
        "epi.regimen_partial_efficacy",
        "epi.active_tb_progression",
        "cost.test_igra",
        "cost.test_tst",
        "cost.regimen_3hp",
        "cost.regimen_4r",
        "cost.regimen_3hr",
        "cost.regimen_6h",
        "cost.regimen_9h",
        "cost.active_tb_disease",
        "cost.false_positive_incremental",
        "cost.program_setup",
        "cost.program_running",
        "daly.active_tb_disability_weight",
        "daly.active_tb_duration",
        "daly.tb_case_fatality_risk",
        "daly.yll_per_tb_death",
        "threshold.gdp_per_capita",
    ]:
        rows.append(_reviewed_registry_row(assumption_id))
    for assumption_id in ["daly.tpt_health_loss", "daly.adr_health_loss", "daly.post_tb_sequelae"]:
        rows.append(
            {
                **_reviewed_registry_row(assumption_id),
                "category": "daly",
                "inclusionStatus": "excluded",
                "reviewStatus": "reviewed_exclusion",
                "currentValue": "",
                "notes": "Synthetic reviewed exclusion rationale.",
            }
        )
    rows.append(
        {
            **_reviewed_registry_row("cost.tpt_adr_management"),
            "category": "cost",
            "inclusionStatus": "bundled",
            "currentValue": "",
            "bundledIntoAssumptionId": "cost.regimen_3hp",
            "notes": "ADR management cost reviewed as bundled into regimen cost.",
        }
    )
    return rows


def _reviewed_epi_config() -> dict:
    return apply_ltbi_state_edit(
        build_default_config(),
        baseline_recent_proportion=0.2,
        transition_rate_per_year=0.2,
        source="Synthetic reviewed APY recent-fraction source",
        status="configured_reviewed",
        notes="Synthetic reviewed source for readiness testing.",
    )


def _reviewed_econ_config() -> dict:
    config = _synthetic_econ()
    for item in config["costItems"]:
        item["sourceCitation"] = "Synthetic reviewed source"
        item["sourceYearStatus"] = "explicit"
    config["threshold"].update(
        {
            "value": 100000,
            "currency": "AUD",
            "referenceYear": "2025-26",
            "source": "Synthetic reviewed threshold source",
            "status": "configured_reviewed",
        }
    )
    return config


if __name__ == "__main__":
    unittest.main()
