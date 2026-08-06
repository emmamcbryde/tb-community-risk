from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.results_workbook import build_results_workbook
from engine.apy.decision_analysis import run_scenario_comparison
from engine.apy.early_review import run_early_screening_review
from engine.apy.sensitivity import run_one_way_sensitivity, run_threshold_analysis
from tests.test_apy_event_ledger_economics import _reviewed_epi, _synthetic_econ


def main() -> None:
    cfg = _reviewed_epi({"N": 24, "screenCoverage": 0.25, "screeningStrategy": "random", "nReps": 2, "seed": 11})
    econ = _synthetic_econ({"threshold": 50000, "program_setup": 0, "program_running": 0})
    scenarios = [
        {"scenarioId": "igra_random", "label": "IGRA random", "changes": {"test": "IGRA", "screeningStrategy": "random"}},
        {"scenarioId": "tst_random", "label": "TST random", "changes": {"test": "TST", "screeningStrategy": "random"}},
    ]
    deterministic = run_scenario_comparison(cfg, econ, scenarios, model_type="expected_value")
    assert deterministic["validation"]["isValid"], deterministic["validation"]
    stochastic = run_scenario_comparison(cfg, econ, scenarios, model_type="agent_based", n_reps=2, master_seed=11)
    assert stochastic["validation"]["isValid"], stochastic["validation"]
    assert stochastic["pairedComparisons"][0]["pairedCohortFingerprintMatch"]

    sens = run_one_way_sensitivity(
        cfg,
        econ,
        [
            {
                "parameterId": "synthetic.pStartTPT",
                "label": "Synthetic TPT start",
                "adapter": "pStartTPT",
                "baseValue": 0.85,
                "lowValue": 0.0,
                "highValue": 1.0,
                "unit": "probability",
                "scale": "probability",
                "source": "Synthetic validation range, not APY reference evidence",
                "reviewStatus": "configured_reviewed",
                "provisional": False,
            }
        ],
        ["tpt_started_total", "active_tb_cases_prevented", "incrementalCost"],
    )
    assert sens["validation"]["isValid"], sens["validation"]
    assert sens["tornadoRows"]

    threshold = run_threshold_analysis(
        cfg,
        econ,
        {"parameterId": "ltbiPrevalence", "adapter": "ltbiPrevalence", "label": "LTBI prevalence"},
        "active_tb_cases_prevented",
        {"low": 0.02, "high": 0.12, "gridPoints": 3, "target": -999},
    )
    assert threshold["validation"]["isValid"], threshold["validation"]
    assert len(threshold["grid"]) == 3

    early_common = {
        "reviewId": "synthetic_e2e_review",
        "screenedToDate": 6,
        "plannedTotalScreened": 12,
        "eligiblePopulation": 24,
        "prior": {
            "type": "discrete_grid",
            "weights": [0.5, 0.5],
            "source": "Synthetic validation prior, not APY reference evidence",
        },
        "prevalenceGrid": [0.0, 0.1],
    }
    early_low = run_early_screening_review(cfg, econ, {**early_common, "testPositiveToDate": 0})
    early_high = run_early_screening_review(cfg, econ, {**early_common, "testPositiveToDate": 5})
    assert early_low["validation"]["isValid"], early_low["validation"]
    assert early_high["validation"]["isValid"], early_high["validation"]
    assert early_low["posterior"]["summary"]["mean"] < early_high["posterior"]["summary"]["mean"]

    workbook = build_results_workbook(
        config=cfg,
        bundle={"metadata": {}, "headline": {}, "technical": {}, "downloads": {}},
        economics_config=econ,
        decision_analysis_results={
            "scenarioComparison": deterministic,
            "sensitivity": sens,
            "threshold": threshold,
            "earlyReview": early_low,
        },
    )
    wb = load_workbook(BytesIO(workbook), read_only=True)
    required = {
        "Decision_scenarios",
        "Scenario_comparison",
        "Sensitivity_results",
        "Threshold_results",
        "Early_review_prior_posterior",
    }
    assert required.issubset(set(wb.sheetnames)), wb.sheetnames
    wb.close()
    print("Synthetic clinician decision workflow validation passed.")
    print("Deterministic scenarios:", len(deterministic["scenarioSummaries"]))
    print("Stochastic paired comparisons:", len(stochastic["pairedComparisons"]))
    print("Sensitivity tornado rows:", len(sens["tornadoRows"]))
    print("Threshold grid rows:", len(threshold["grid"]))
    print("Early-review low/high posterior means:", early_low["posterior"]["summary"]["mean"], early_high["posterior"]["summary"]["mean"])


if __name__ == "__main__":
    main()
