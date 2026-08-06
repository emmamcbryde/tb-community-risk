from __future__ import annotations

import argparse
from copy import deepcopy
from pathlib import Path
import sys
from typing import Any
from xml.sax.saxutils import escape
import zipfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from engine.apy.config import build_default_config
from engine.apy.ltbi_state import enable_development_compatibility_mode
from engine.apy.economics import (
    build_economics_preset_kwab150,
    run_health_economics,
)
from engine.apy.health_economics import (
    build_default_health_economic_assumptions,
    calculate_health_outcomes,
    calculate_icers,
    calculate_post_tb_sequelae,
)
from engine.apy.runner import run_scenario_with_do_nothing


DEFAULT_STRATEGIES = ["random", "ltbi", "cure", "prevent"]
DEFAULT_OUTPUT_DIR = Path("outputs/apy_paper_daly_icer")
DEFAULT_WORKBOOK_PATH = Path("paper/excel/APY_economics_decision_tree_formula_linked.xlsx")
POST_TB_WORKBOOK_PATH = Path("paper/excel/APY_economics_decision_tree_formula_linked_post_tb.xlsx")
POST_TB_REPORT_PATH = Path("APY_LTBI_health_economics_report_updated_post_tb.docx")
LEGACY_OFFLINE_STATUS = "legacy_offline_not_clinician_ready"
LEGACY_OFFLINE_NOTE = (
    "Offline paper-output script using legacy DALY/QALY sensitivity assumptions; "
    "not the authoritative Milestone 3 clinician-facing economic result."
)
OUTPUT_FILES = {
    "strategy": "apy_daly_icer_strategy_summary.csv",
    "health": "apy_daly_icer_health_outcomes.csv",
    "costs": "apy_daly_icer_costs.csv",
    "sensitivity": "apy_daly_icer_sensitivity.csv",
    "source_audit": "apy_daly_qaly_source_audit.csv",
    "notes": "apy_daly_icer_notes.md",
    "source_audit_notes": "apy_daly_qaly_source_audit_notes.md",
    "post_tb_scenarios": "apy_post_tb_sequelae_scenarios.csv",
    "post_tb_notes": "apy_post_tb_sequelae_notes.md",
}


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    acute_outputs = load_or_generate_acute_outputs(args, output_dir)
    strategy_df = acute_outputs["strategy"]
    health_df = acute_outputs["health"]
    cost_df = acute_outputs["costs"]
    sensitivity_df = acute_outputs["sensitivity"]
    source_audit_df = acute_outputs["source_audit"]

    strategy_df.to_csv(output_dir / OUTPUT_FILES["strategy"], index=False)
    health_df.to_csv(output_dir / OUTPUT_FILES["health"], index=False)
    cost_df.to_csv(output_dir / OUTPUT_FILES["costs"], index=False)
    sensitivity_df.to_csv(output_dir / OUTPUT_FILES["sensitivity"], index=False)
    source_audit_df.to_csv(output_dir / OUTPUT_FILES["source_audit"], index=False)
    write_notes(output_dir / OUTPUT_FILES["notes"], strategy_df, sensitivity_df, args)
    write_source_audit_notes(output_dir / OUTPUT_FILES["source_audit_notes"])

    assumptions = build_default_health_economic_assumptions()
    post_tb_df = build_post_tb_scenarios(strategy_df, assumptions)
    post_tb_df.to_csv(output_dir / OUTPUT_FILES["post_tb_scenarios"], index=False)
    write_post_tb_notes(output_dir / OUTPUT_FILES["post_tb_notes"], post_tb_df, assumptions)

    if should_write_paper_artifacts(args, output_dir):
        build_post_tb_workbook(
            REPO_ROOT / DEFAULT_WORKBOOK_PATH,
            REPO_ROOT / POST_TB_WORKBOOK_PATH,
            strategy_df,
            post_tb_df,
            assumptions,
        )
        build_post_tb_report(REPO_ROOT / POST_TB_REPORT_PATH, post_tb_df, assumptions)

    print(f"Wrote APY DALY/QALY/ICER outputs to {output_dir}")
    print(f"Wrote post-TB sequelae outputs to {output_dir / OUTPUT_FILES['post_tb_scenarios']}")
    print(strategy_df.head().to_string(index=False))
    return 0


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run legacy offline APY paper DALY/QALY/ICER outputs. These outputs "
            "are not clinician-ready Milestone 3 economic conclusions."
        ),
    )
    parser.add_argument("--output-dir", default="outputs/apy_paper_daly_icer")
    parser.add_argument("--quick", action="store_true")
    parser.add_argument("--n", type=int, default=1500)
    parser.add_argument("--n-reps", type=int, default=2000)
    parser.add_argument("--seed", type=int, default=1)
    parser.add_argument("--screen-coverage", type=float, default=0.30)
    parser.add_argument("--strategies", default=",".join(DEFAULT_STRATEGIES))
    parser.add_argument("--include-qaly", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--include-daly", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument(
        "--rerun-model",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Force regeneration of acute APY outputs even if the expected CSVs already exist.",
    )
    parser.add_argument(
        "--write-paper-artifacts",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="Write the post-TB workbook and Word report. Defaults to on only for the standard paper output folder.",
    )
    return parser.parse_args(argv)


def parse_strategies(raw: str) -> list[str]:
    values = [item.strip().lower() for item in raw.split(",") if item.strip()]
    return values or DEFAULT_STRATEGIES


def should_write_paper_artifacts(args: argparse.Namespace, output_dir: Path) -> bool:
    if args.write_paper_artifacts is not None:
        return bool(args.write_paper_artifacts)
    return output_dir.resolve() == (REPO_ROOT / DEFAULT_OUTPUT_DIR).resolve()


def load_or_generate_acute_outputs(
    args: argparse.Namespace,
    output_dir: Path,
) -> dict[str, pd.DataFrame]:
    if not args.rerun_model and acute_outputs_exist(output_dir):
        return {
            "strategy": pd.read_csv(output_dir / OUTPUT_FILES["strategy"]),
            "health": pd.read_csv(output_dir / OUTPUT_FILES["health"]),
            "costs": pd.read_csv(output_dir / OUTPUT_FILES["costs"]),
            "sensitivity": pd.read_csv(output_dir / OUTPUT_FILES["sensitivity"]),
            "source_audit": load_optional_csv(output_dir / OUTPUT_FILES["source_audit"]),
        }

    n = 100 if args.quick else args.n
    n_reps = 5 if args.quick else args.n_reps
    strategies = parse_strategies(args.strategies)

    strategy_rows: list[dict[str, Any]] = []
    health_rows: list[dict[str, Any]] = []
    cost_rows: list[dict[str, Any]] = []
    sensitivity_rows: list[dict[str, Any]] = []
    source_audit_rows: list[dict[str, Any]] = []

    for strategy in strategies:
        scenario = run_one_strategy(
            strategy=strategy,
            n=n,
            n_reps=n_reps,
            seed=args.seed,
            screen_coverage=args.screen_coverage,
            include_daly=args.include_daly,
            include_qaly=args.include_qaly,
        )
        strategy_rows.append(scenario["strategyRow"])
        health_rows.extend(scenario["healthRows"])
        cost_rows.extend(scenario["costRows"])
        sensitivity_rows.extend(scenario["sensitivityRows"])
        source_audit_rows.append(scenario["sourceAuditRow"])

    return {
        "strategy": pd.DataFrame(strategy_rows),
        "health": pd.DataFrame(health_rows),
        "costs": pd.DataFrame(cost_rows),
        "sensitivity": pd.DataFrame(sensitivity_rows),
        "source_audit": pd.DataFrame(source_audit_rows),
    }


def acute_outputs_exist(output_dir: Path) -> bool:
    required = (
        OUTPUT_FILES["strategy"],
        OUTPUT_FILES["health"],
        OUTPUT_FILES["costs"],
        OUTPUT_FILES["sensitivity"],
    )
    return all((output_dir / name).exists() for name in required)


def load_optional_csv(path: Path) -> pd.DataFrame:
    if path.exists():
        return pd.read_csv(path)
    return pd.DataFrame()


def build_post_tb_scenarios(
    strategy_df: pd.DataFrame,
    assumptions: dict[str, Any],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for row in strategy_df.to_dict(orient="records"):
        acute_qalys = first_not_none(
            row.get("qalysGained_DaleMortalityAdjusted"),
            row.get("qalysGained"),
        )
        acute_net_cost = row.get("netCostVsBaseline")
        acute_dalys = row.get("dalysAverted")
        tb_cases = row.get("tbCasesPrevented")

        acute_only = calculate_post_tb_sequelae(
            tb_cases,
            acute_dalys,
            acute_qalys,
            acute_net_cost,
            assumptions,
        )
        with_post_tb = calculate_post_tb_sequelae(
            tb_cases,
            acute_dalys,
            acute_qalys,
            acute_net_cost,
            {"postTB": {"includePostTBSequelae": True}},
        )
        acute_values = acute_only["postTBScenarios"]
        post_tb_values = with_post_tb["postTBScenarios"]
        rows.extend(
            [
                build_post_tb_view_row(row, assumptions, acute_values, post_tb_values, "acute_only"),
                build_post_tb_view_row(
                    row,
                    assumptions,
                    acute_values,
                    post_tb_values,
                    "acute_plus_post_tb_daly_tail",
                ),
                build_post_tb_view_row(
                    row,
                    assumptions,
                    acute_values,
                    post_tb_values,
                    "acute_plus_post_tb_qaly_tail",
                ),
            ]
        )
    return pd.DataFrame(rows)


def build_post_tb_view_row(
    strategy_row: dict[str, Any],
    assumptions: dict[str, Any],
    acute_values: dict[str, Any],
    post_tb_values: dict[str, Any],
    view: str,
) -> dict[str, Any]:
    if view == "acute_only":
        displayed_dalys = acute_values.get("acuteDALYsAverted")
        displayed_qalys = acute_values.get("acuteQALYsGained")
        displayed_cost = acute_values.get("acuteNetCost")
        displayed_post_tb_dalys = 0.0
        displayed_post_tb_qalys = 0.0
        displayed_post_tb_costs = 0.0
    elif view == "acute_plus_post_tb_daly_tail":
        displayed_dalys = post_tb_values.get("totalDALYsAvertedIncludingPostTB")
        displayed_qalys = acute_values.get("acuteQALYsGained")
        displayed_cost = post_tb_values.get("netCostIncludingPostTB")
        displayed_post_tb_dalys = post_tb_values.get("postTBDALYsAverted")
        displayed_post_tb_qalys = 0.0
        displayed_post_tb_costs = post_tb_values.get("postTBCostsAverted")
    else:
        displayed_dalys = acute_values.get("acuteDALYsAverted")
        displayed_qalys = post_tb_values.get("totalQALYsGainedIncludingPostTB")
        displayed_cost = post_tb_values.get("netCostIncludingPostTB")
        displayed_post_tb_dalys = 0.0
        displayed_post_tb_qalys = post_tb_values.get("postTBQALYsGained")
        displayed_post_tb_costs = post_tb_values.get("postTBCostsAverted")

    return {
        "scenario_id": strategy_row.get("scenario_id"),
        "screeningStrategy": strategy_row.get("screeningStrategy"),
        "view": view,
        "tbCasesPrevented": acute_values.get("tbCasesPrevented"),
        "acuteDALYsAverted": acute_values.get("acuteDALYsAverted"),
        "postTBDALYsAverted": displayed_post_tb_dalys,
        "totalDALYsAvertedIncludingPostTB": displayed_dalys,
        "acuteQALYsGained": acute_values.get("acuteQALYsGained"),
        "postTBQALYsGained": displayed_post_tb_qalys,
        "totalQALYsGainedIncludingPostTB": displayed_qalys,
        "postTBCostsAverted": displayed_post_tb_costs,
        "acuteNetCost": acute_values.get("acuteNetCost"),
        "netCostIncludingPostTB": displayed_cost,
        "costPerDALYIncludingPostTB": safe_divide(displayed_cost, displayed_dalys),
        "costPerQALYIncludingPostTB": safe_divide(displayed_cost, displayed_qalys),
        "postTBAnnualCareCost": assumptions["postTB"]["postTBAnnualCareCost"],
        "postTBDurationYears": assumptions["postTB"]["postTBDurationYears"],
        "postTBExcessMortalityMultiplier": assumptions["postTB"][
            "postTBExcessMortalityMultiplier"
        ],
        "pPTLD": assumptions["postTB"]["pPTLD"],
        "pPTLD_low": assumptions["postTB"]["pPTLD_low"],
        "pPTLD_high": assumptions["postTB"]["pPTLD_high"],
        "notes": assumptions["postTB"]["notes"],
    }


def run_one_strategy(
    strategy: str,
    n: int,
    n_reps: int,
    seed: int,
    screen_coverage: float,
    include_daly: bool = True,
    include_qaly: bool = True,
) -> dict[str, Any]:
    config = enable_development_compatibility_mode(build_default_config())
    config.update(
        {
            "scenarioLabel": f"APY paper DALY ICER - {strategy}",
            "N": int(n),
            "nReps": int(n_reps),
            "seed": int(seed),
            "screenWindow": 2,
            "followHorizon": 20,
            "screenCoverage": float(screen_coverage),
            "screeningStrategy": strategy,
            "testType": "IGRA",
            "regimen": "3HP",
            "pStartTPT": 0.85,
            "regimenPComplete": 0.80,
            "regimenADRstop": 0.05,
            "regimenEffFull": 0.85,
        }
    )
    model_out = run_scenario_with_do_nothing(config)
    econ_config = build_economics_preset_kwab150()
    econ_config["costs"]["falsePositiveIncrementalPerPerson"] = 0.0
    econ_config["costs"]["programSetupTotal"] = 0.0
    econ_config["costs"]["programRunningTotal"] = 0.0
    economics = run_health_economics(model_out, econ_config)
    assumptions = build_default_health_economic_assumptions()
    health = calculate_health_outcomes(model_out, assumptions)
    icer = calculate_icers(economics, health, assumptions)

    scenario_id = f"apy_paper_{strategy}_igra_3hp"
    strategy_row = build_strategy_row(
        scenario_id,
        strategy,
        model_out,
        economics,
        health,
        icer,
        include_daly,
        include_qaly,
    )
    sensitivity_rows = build_sensitivity_rows(
        scenario_id,
        strategy,
        model_out,
        economics,
        include_daly,
        include_qaly,
    )
    source_audit_row = build_source_audit_row(scenario_id, strategy, health, icer)
    return {
        "strategyRow": strategy_row,
        "healthRows": flatten_section_rows(scenario_id, strategy, "health", health["healthOutcomes"]),
        "costRows": flatten_section_rows(scenario_id, strategy, "costs", economics["costs"]),
        "sensitivityRows": sensitivity_rows,
        "sourceAuditRow": source_audit_row,
    }


def build_strategy_row(
    scenario_id: str,
    strategy: str,
    model_out: dict[str, Any],
    economics: dict[str, Any],
    health: dict[str, Any],
    icer: dict[str, Any],
    include_daly: bool,
    include_qaly: bool,
) -> dict[str, Any]:
    results = model_out["results"]
    config = results["interfaceConfig"]
    dynamic = model_out["bundle"]["technical"]["dynamicComparison"]
    costs = economics["costs"]
    health_values = health["healthOutcomes"]
    icers = icer["icers"]
    nmb = icer["nmb"]
    row = {
        "scenario_id": scenario_id,
        "screeningStrategy": strategy,
        "testType": config["testType"],
        "regimen": config["regimen"],
        "N": config["N"],
        "nReps": config["nReps"],
        "screenCoverage": config["screenCoverage"],
        "screenWindow": config["screenWindow"],
        "followHorizon": config["followHorizon"],
        "analysis_status": LEGACY_OFFLINE_STATUS,
        "analysis_note": LEGACY_OFFLINE_NOTE,
        "m3_authoritative_economics_contract": economics.get("contractVersion"),
        "nScreened_median": metric(results, "nScreened"),
        "nTotalCoursesStarted_median": metric(results, "nTotalCoursesStarted"),
        "nTotalCoursesCompleted_median": metric(results, "nTotalCoursesCompleted"),
        "nFalsePositiveTreated_median": metric(results, "nFalsePositiveTreated"),
        "nADRstop_median": metric(results, "nADRstop"),
        "nCuredInfection_median": metric(results, "nCuredInfection"),
        "nPreventedActiveTB_median": metric(results, "nPreventedActiveTB"),
        "nActiveBy20y_DoNothing_median": dynamic.get("cumulative_baseline_active_tb_cases"),
        "nActiveBy20y_AfterStrategy_median": dynamic.get("cumulative_intervention_active_tb_cases"),
        "relativeReduction20y_median": dynamic.get("relative_reduction_cumulative_active_tb_cases"),
        "testingCost": costs.get("testingCost"),
        "treatmentCost": costs.get("treatmentCost"),
        "tbDiseaseCostsAverted": costs.get("tbDiseaseCostsAverted"),
        "totalProgramCost": costs.get("totalProgramCost"),
        "netCostVsBaseline": costs.get("netCostVsBaseline"),
        "tbCasesPrevented": health_values.get("tbCasesPrevented") if include_daly or include_qaly else None,
        "tbDeathsPrevented": health_values.get("tbDeathsPrevented") if include_daly else None,
        "yldAverted": health_values.get("yldAverted") if include_daly else None,
        "yllAverted": health_values.get("yllAverted") if include_daly else None,
        "dalysAverted": health_values.get("dalysAverted") if include_daly else None,
        "morbidityOnlyQalysGained": health_values.get("morbidityOnlyQalysGained") if include_qaly else None,
        "mortalityQalysGained": health_values.get("mortalityQalysGained") if include_qaly else None,
        "activeTbMorbidityQalysGained_Dale": health_values.get("activeTbMorbidityQalysGained_Dale") if include_qaly else None,
        "activeTbMorbidityQalysGained_GBD": health_values.get("activeTbMorbidityQalysGained_GBD") if include_qaly else None,
        "qalysGained_DaleMortalityAdjusted": health_values.get("qalysGained_DaleMortalityAdjusted") if include_qaly else None,
        "qalysGained_GBDAlignedMortalityAdjusted": health_values.get("qalysGained_GBDAlignedMortalityAdjusted") if include_qaly else None,
        "qalysGained": health_values.get("qalysGained") if include_qaly else None,
        "costPerTBCasePrevented": icers.get("costPerTBCasePrevented"),
        "costPerDALYAverted": icers.get("costPerDALYAverted") if include_daly else None,
        "costPerQALYGained": icers.get("costPerQALYGained") if include_qaly else None,
        "costPerQALYGained_GBDAligned": icers.get("costPerQALYGained_GBDAligned") if include_qaly else None,
        "nmbDALY_45000": nmb.get("netMonetaryBenefitDALY_low") if include_daly else None,
        "nmbDALY_75000": nmb.get("netMonetaryBenefitDALY_high") if include_daly else None,
        "nmbQALY_45000": nmb.get("netMonetaryBenefitQALY_low") if include_qaly else None,
        "nmbQALY_75000": nmb.get("netMonetaryBenefitQALY_high") if include_qaly else None,
        "nmbQALY_GBDAligned_45000": nmb.get("netMonetaryBenefitQALY_GBDAligned_low") if include_qaly else None,
        "nmbQALY_GBDAligned_75000": nmb.get("netMonetaryBenefitQALY_GBDAligned_high") if include_qaly else None,
    }
    row["dominance_status"] = dominance_status(row)
    row["notes"] = (
        "Direct person-level prevented cases; no downstream transmission effects. "
        f"{LEGACY_OFFLINE_NOTE}"
    )
    return row


def build_sensitivity_rows(
    scenario_id: str,
    strategy: str,
    model_out: dict[str, Any],
    economics: dict[str, Any],
    include_daly: bool,
    include_qaly: bool,
) -> list[dict[str, Any]]:
    scenarios = [
        ("no_ltbi_treatment_utility_decrement", {}),
        (
            "ltbi_treatment_decrement_0_0133",
            {"qaly": {"ltbiTreatmentDecrement": 0.0133}},
        ),
        ("yld_only_no_yll", {"daly": {"includeYLL": False}}),
        ("conservative_yll_10", {"daly": {"yllPerTBDeath": 10.0}}),
        ("higher_yll_30", {"daly": {"yllPerTBDeath": 30.0}}),
        ("lower_tb_case_fatality", {"daly": {"tbCaseFatalityRisk": 0.037}}),
        ("higher_tb_case_fatality", {"daly": {"tbCaseFatalityRisk": 0.111}}),
    ]
    rows = []
    for sensitivity_name, overrides in scenarios:
        assumptions = build_default_health_economic_assumptions()
        deep_update(assumptions, overrides)
        health = calculate_health_outcomes(model_out, assumptions)
        icer = calculate_icers(economics, health, assumptions)
        rows.append(
            {
                "scenario_id": scenario_id,
                "screeningStrategy": strategy,
                "sensitivity": sensitivity_name,
                "tbCasesPrevented": health["healthOutcomes"].get("tbCasesPrevented"),
                "dalysAverted": health["healthOutcomes"].get("dalysAverted") if include_daly else None,
                "morbidityOnlyQalysGained": health["healthOutcomes"].get("morbidityOnlyQalysGained") if include_qaly else None,
                "mortalityQalysGained": health["healthOutcomes"].get("mortalityQalysGained") if include_qaly else None,
                "qalysGained": health["healthOutcomes"].get("qalysGained") if include_qaly else None,
                "qalysGained_GBDAlignedMortalityAdjusted": health["healthOutcomes"].get("qalysGained_GBDAlignedMortalityAdjusted") if include_qaly else None,
                "costPerDALYAverted": icer["icers"].get("costPerDALYAverted") if include_daly else None,
                "costPerQALYGained": icer["icers"].get("costPerQALYGained") if include_qaly else None,
                "costPerQALYGained_GBDAligned": icer["icers"].get("costPerQALYGained_GBDAligned") if include_qaly else None,
                "nmbDALY_45000": icer["nmb"].get("netMonetaryBenefitDALY_low") if include_daly else None,
                "nmbDALY_75000": icer["nmb"].get("netMonetaryBenefitDALY_high") if include_daly else None,
                "nmbQALY_45000": icer["nmb"].get("netMonetaryBenefitQALY_low") if include_qaly else None,
                "nmbQALY_75000": icer["nmb"].get("netMonetaryBenefitQALY_high") if include_qaly else None,
                "nmbQALY_GBDAligned_45000": icer["nmb"].get("netMonetaryBenefitQALY_GBDAligned_low") if include_qaly else None,
                "nmbQALY_GBDAligned_75000": icer["nmb"].get("netMonetaryBenefitQALY_GBDAligned_high") if include_qaly else None,
                "notes": "Sensitivity assumption; not a final policy setting.",
            }
        )
    return rows


def build_source_audit_row(
    scenario_id: str,
    strategy: str,
    health: dict[str, Any],
    icer: dict[str, Any],
) -> dict[str, Any]:
    values = health["healthOutcomes"]
    assumptions = health["assumptions"]
    daly = assumptions["daly"]
    qaly = assumptions["qaly"]
    tb_cases = values.get("tbCasesPrevented")
    dale_morbidity_per_case = (
        (qaly["healthyUtility"] - qaly["activeTBUtility"])
        * qaly.get("qalyActiveTbDurationYears", qaly["activeTBDurationYears"])
    )
    gbd_morbidity_per_case = (
        daly["activeTBDisabilityWeight"] * daly["activeTBDurationYears"]
    )
    yld_per_case = gbd_morbidity_per_case
    yll_per_case = daly["tbCaseFatalityRisk"] * daly["yllPerTBDeath"]
    mortality_qaly_per_case = (
        daly["tbCaseFatalityRisk"]
        * values.get("qualityAdjustedLifeExpectancyPerTBDeath")
    )
    dalys = values.get("dalysAverted")
    dale_qalys = values.get("qalysGained_DaleMortalityAdjusted")
    gbd_qalys = values.get("qalysGained_GBDAlignedMortalityAdjusted")
    morbidity_qalys = values.get("morbidityOnlyQalysGained")
    return {
        "scenario_id": scenario_id,
        "screeningStrategy": strategy,
        "tbCasesPrevented": tb_cases,
        "dale_bauer_morbidity_qaly_per_tb_case": dale_morbidity_per_case,
        "gbd_aligned_morbidity_qaly_per_tb_case": gbd_morbidity_per_case,
        "yld_per_tb_case": yld_per_case,
        "yll_per_tb_case": yll_per_case,
        "mortality_qalys_per_tb_case": mortality_qaly_per_case,
        "morbidityOnlyQalysGained": morbidity_qalys,
        "total_DaleMortalityAdjusted_QALYs": dale_qalys,
        "total_GBDAlignedMortalityAdjusted_QALYs": gbd_qalys,
        "total_DALYs": dalys,
        "daly_to_morbidity_only_qaly_ratio": safe_divide(dalys, morbidity_qalys),
        "daly_to_dale_mortality_adjusted_qaly_ratio": safe_divide(dalys, dale_qalys),
        "daly_to_gbd_aligned_qaly_ratio": safe_divide(dalys, gbd_qalys),
        "costPerQALYGained_GBDAligned": icer["icers"].get("costPerQALYGained_GBDAligned"),
    }


def metric(results: dict[str, Any], metric_name: str) -> float | None:
    rows = results["summary"]
    matched = rows.loc[rows["Metric"] == metric_name, "Median"]
    if matched.empty:
        return None
    return matched.iloc[0]


def flatten_section_rows(
    scenario_id: str,
    strategy: str,
    section: str,
    values: dict[str, Any],
) -> list[dict[str, Any]]:
    return [
        {
            "scenario_id": scenario_id,
            "screeningStrategy": strategy,
            "section": section,
            "metric": key,
            "value": value,
        }
        for key, value in values.items()
        if isinstance(value, (int, float)) or value is None
    ]


def dominance_status(row: dict[str, Any]) -> str:
    net_cost = row.get("netCostVsBaseline")
    cases = row.get("tbCasesPrevented")
    if net_cost is None or cases is None:
        return "not assessed"
    if net_cost < 0 and cases > 0:
        return "dominant_cost_saving"
    if cases <= 0:
        return "no_health_gain"
    return "incremental_cost"


def deep_update(target: dict[str, Any], updates: dict[str, Any]) -> None:
    for key, value in updates.items():
        if isinstance(value, dict) and isinstance(target.get(key), dict):
            deep_update(target[key], value)
        else:
            target[key] = value


def safe_divide(a: Any, b: Any) -> float | None:
    if a is None or b in (None, 0):
        return None
    try:
        return float(a) / float(b)
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def write_notes(
    path: Path,
    strategy_df: pd.DataFrame,
    sensitivity_df: pd.DataFrame,
    args: argparse.Namespace,
) -> None:
    best_daly = best_strategy(strategy_df, "costPerDALYAverted")
    best_qaly = best_strategy(strategy_df, "costPerQALYGained")
    lines = [
        "# APY Paper DALY/QALY/ICER Outputs",
        "",
        "Purpose: offline health-economic outputs for the APY paper/report.",
        "",
        "Model basis: Python APY closed-cohort, person-level prevention model using direct prevented TB cases. Downstream transmission effects are not included.",
        "",
        "Costing: Australian health-care system perspective, 2019 AUD, KWAB150/Dale-informed unit costs, 3% discounting assumption recorded for interpretation.",
        "",
        "Program setup, program running, and false-positive incremental costs are set to zero in this offline paper script because the KWAB150 preset does not specify them. This should be sensitivity-tested if programme implementation costs are added.",
        "",
        "DALY assumptions: active TB disability weight 0.333, active TB duration 0.5 years, configurable TB case fatality risk and YLL per TB death. YLL assumptions require explicit review.",
        "",
        "QALY assumptions: healthy utility 0.8733, active TB utility 0.8182, SAE utility 0.8685, and LTBI treatment decrement sensitivity based on Dale-informed assumptions.",
        "",
        "Primary QALYs now include mortality benefits using the same TB case fatality and YLL assumptions as DALYs. Dale-compatible QALYs use the Dale/Bauer morbidity utility decrement; GBD-aligned QALYs are reported as a sensitivity using the active TB disability weight for morbidity.",
        "",
        "ICER formulas: incremental cost divided by DALYs averted, QALYs gained, or active TB cases prevented. Net monetary benefit equals health gain times WTP threshold minus incremental cost.",
        "",
        "Limitations: not an online tool output; not a full transmission model; not a final policy gate; mortality/YLL assumptions are sensitivity inputs.",
        "",
        f"Lowest cost per DALY averted among calculable base-case strategies: {best_daly}.",
        "",
        f"Lowest cost per QALY gained among calculable base-case strategies: {best_qaly}.",
        "",
        "The APY model supports sequencing and prioritisation, not exclusion from care.",
        "",
        f"Run settings: quick={args.quick}, N={100 if args.quick else args.n}, nReps={5 if args.quick else args.n_reps}, seed={args.seed}, screenCoverage={args.screen_coverage}.",
        "",
        f"Sensitivity rows generated: {len(sensitivity_df)}.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def write_source_audit_notes(path: Path) -> None:
    lines = [
        "# APY DALY/QALY Source Audit Notes",
        "",
        "This audit separates morbidity and mortality sources for DALYs and QALYs.",
        "",
        "DALYs use active TB YLD plus YLL from TB deaths prevented. YLD uses the active TB disability weight of 0.333 over 0.5 years. YLL uses the configurable TB case fatality risk and YLL per TB death.",
        "",
        "Primary QALYs are Dale-compatible and mortality-adjusted. Morbidity uses the Dale/Bauer utility decrement from healthy/LTBI utility 0.8733 to active TB utility 0.8182 over the active TB duration. Mortality QALYs use TB deaths prevented multiplied by quality-adjusted life expectancy per TB death.",
        "",
        "When qualityAdjustedLifeExpectancyPerTBDeath is not supplied, it is calculated as yllPerTBDeath multiplied by healthyUtility.",
        "",
        "GBD-aligned QALYs are provided as a morbidity sensitivity. They use the active TB disability weight for morbidity and the same mortality QALY logic.",
        "",
        "TB case fatality risk, YLL per TB death, and whether to prefer Dale/Bauer or GBD morbidity weights remain important sensitivity assumptions.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def write_post_tb_notes(
    path: Path,
    post_tb_df: pd.DataFrame,
    assumptions: dict[str, Any],
) -> None:
    prevent_row = select_view(post_tb_df, "prevent", "acute_plus_post_tb_daly_tail")
    prevent_qaly_row = select_view(post_tb_df, "prevent", "acute_plus_post_tb_qaly_tail")
    lines = [
        "# APY Post-TB Sequelae Scenario Notes",
        "",
        "Primary source: `docs/source_material/Post TB Mortality & Standardised Follow up_Byrne June 2026.pptx`.",
        "",
        "Evidence summary from the Byrne June 2026 talk and linked literature:",
        "",
        "- Global post-TB burden was presented as 58 million DALYs due to post-TB sequelae, around 47% of the combined incident-TB plus post-TB burden.",
        "- The talk reports 12.1 DALYs per incident TB case in total, with 6.3 during active TB and 5.8 after TB treatment.",
        "- Byrne highlighted increased long-term mortality after TB treatment, including a pooled SMR of 2.91 from Romanowski et al. and an Australian cause-mortality ratio of 1.16 with most deaths within 10 years.",
        "- PTLD manifestations listed in the talk include fibrosis, bronchiectasis, emphysema/COPD, chronic respiratory failure, pulmonary hypertension, lung cancer, and chronic fungal disease.",
        "- Disability domains highlighted in the talk include respiratory, mental health, renal, visual, hearing, and musculoskeletal impairment.",
        "",
        "APY interpretation:",
        "",
        "- The acute APY model remains unchanged. Post-TB morbidity and mortality are layered on as a scenario tail attached to prevented incident TB cases.",
        "- Base post-TB assumptions remain uncertain for APY. PTLD prevalence is left unresolved by default and the workbook keeps that assumption editable.",
        "- The default post-TB annual care cost is zero until APY-specific utilisation and costing inputs are agreed.",
        "",
        f"Default scenario settings: postTBDalysPerCase={assumptions['postTB']['postTBDalysPerCase']}, postTBQalysLostPerCase={assumptions['postTB']['postTBQalysLostPerCase']}, postTBDurationYears={assumptions['postTB']['postTBDurationYears']}, postTBExcessMortalityMultiplier={assumptions['postTB']['postTBExcessMortalityMultiplier']}.",
        "",
        "Illustrative prevention-targeted strategy under the default tail:",
        "",
        f"- Acute-only DALYs averted: {format_number(select_view(post_tb_df, 'prevent', 'acute_only').get('totalDALYsAvertedIncludingPostTB'))}.",
        f"- Acute plus post-TB DALY tail: {format_number(prevent_row.get('totalDALYsAvertedIncludingPostTB'))}.",
        f"- Acute-only QALYs gained: {format_number(select_view(post_tb_df, 'prevent', 'acute_only').get('totalQALYsGainedIncludingPostTB'))}.",
        f"- Acute plus post-TB QALY tail: {format_number(prevent_qaly_row.get('totalQALYsGainedIncludingPostTB'))}.",
        "",
        "Caveats:",
        "",
        "- These post-TB values are scenario extensions, not APY-specific measured outcomes.",
        "- Excess post-TB mortality is recorded as an evidence input but is not used to alter the APY epidemiological simulation.",
        "- Lifetime post-TB duration remains a sensitivity concept until a numeric APY-compatible duration assumption is agreed.",
        "- The APY model should support planning and sequencing, not denying care.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def build_post_tb_workbook(
    source_path: Path,
    output_path: Path,
    strategy_df: pd.DataFrame,
    post_tb_df: pd.DataFrame,
    assumptions: dict[str, Any],
) -> None:
    wb = load_workbook(source_path)
    remove_sheet_if_exists(wb, "Post_TB_Assumptions")
    remove_sheet_if_exists(wb, "Post_TB_Scenarios")
    remove_sheet_if_exists(wb, "Post_TB_Chart_Data")
    remove_sheet_if_exists(wb, "Post_TB_Audit")

    create_post_tb_assumptions_sheet(wb, assumptions)
    create_post_tb_scenarios_sheet(wb, strategy_df)
    create_post_tb_chart_sheet(wb)
    create_post_tb_audit_sheet(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def remove_sheet_if_exists(wb: Any, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def create_post_tb_assumptions_sheet(wb: Any, assumptions: dict[str, Any]) -> None:
    ws = wb.create_sheet("Post_TB_Assumptions")
    headers = ["Input", "Value", "Units / type", "Status", "Notes"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = Font(bold=True)
    rows = [
        (
            "Primary source",
            "Byrne June 2026 post-TB talk",
            "text",
            "fixed",
            "Primary evidence source: docs/source_material/Post TB Mortality & Standardised Follow up_Byrne June 2026.pptx",
        ),
        (
            "postTBDalysPerCase",
            assumptions["postTB"]["postTBDalysPerCase"],
            "DALYs per prevented TB case",
            "editable",
            "Talk slides report 5.8 post-TB DALYs per incident TB case.",
        ),
        (
            "postTBQalysLostPerCase",
            assumptions["postTB"]["postTBQalysLostPerCase"],
            "QALYs per prevented TB case",
            "editable",
            "User-specified post-TB QALY tail for sensitivity analysis.",
        ),
        (
            "totalTBQalysLostPerCase",
            assumptions["postTB"]["totalTBQalysLostPerCase"],
            "QALYs per incident TB case",
            "reference",
            "US analysis summary: 1.93 QALYs lost per TB case overall.",
        ),
        (
            "pPTLD",
            assumptions["postTB"]["pPTLD"],
            "probability",
            "UNRESOLVED",
            "Leave blank until APY-specific PTLD prevalence assumptions are agreed.",
        ),
        (
            "pPTLD_low",
            assumptions["postTB"]["pPTLD_low"],
            "probability",
            "reference",
            "Low evidence bound from requested scenario range.",
        ),
        (
            "pPTLD_high",
            assumptions["postTB"]["pPTLD_high"],
            "probability",
            "reference",
            "High evidence bound from requested scenario range.",
        ),
        (
            "postTBAnnualCareCost",
            assumptions["postTB"]["postTBAnnualCareCost"],
            "AUD per person-year",
            "editable",
            "Defaults to zero until APY follow-up care utilisation is specified.",
        ),
        (
            "postTBDurationYears",
            assumptions["postTB"]["postTBDurationYears"],
            "years",
            "editable",
            "Use 10 or 20 now; lifetime remains a sensitivity concept requiring explicit numeric translation.",
        ),
        (
            "Duration sensitivity options",
            "10, 20, lifetime",
            "text",
            "reference",
            "Requested scenario set for post-TB duration.",
        ),
        (
            "postTBExcessMortalityMultiplier",
            assumptions["postTB"]["postTBExcessMortalityMultiplier"],
            "multiplier",
            "editable",
            "Australian cohort cause-mortality ratio evidence anchor.",
        ),
        (
            "Excess mortality scenario",
            "Australia CMR 1.16; pooled SMR 2.91",
            "text",
            "editable",
            "Track evidence scenario choice here even though the APY simulation is unchanged.",
        ),
    ]
    yellow = PatternFill(fill_type="solid", fgColor="FFF59D")
    for row_idx, values in enumerate(rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        if values[3] == "UNRESOLVED":
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).fill = yellow
    set_simple_widths(ws, [34, 22, 24, 14, 96])


def create_post_tb_scenarios_sheet(wb: Any, strategy_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Post_TB_Scenarios")
    headers = [
        "Strategy",
        "TB cases prevented",
        "Acute DALYs averted",
        "Post-TB DALYs averted",
        "Total DALYs incl post-TB",
        "Acute QALYs gained",
        "Post-TB QALYs gained",
        "Total QALYs incl post-TB",
        "Acute net cost",
        "Post-TB costs averted",
        "Net cost incl post-TB",
        "Acute cost per DALY",
        "Cost per DALY incl post-TB",
        "Acute cost per QALY",
        "Cost per QALY incl post-TB",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value).font = Font(bold=True)
    for excel_row, (_, row) in enumerate(strategy_df.iterrows(), start=2):
        econ_row = excel_row
        ws.cell(row=excel_row, column=1, value=f"=Strategy_Summary!A{econ_row}")
        ws.cell(row=excel_row, column=2, value=f"=Strategy_Summary!B{econ_row}")
        ws.cell(row=excel_row, column=3, value=f"=Strategy_Summary!D{econ_row}")
        ws.cell(
            row=excel_row,
            column=4,
            value=f"=B{excel_row}*Post_TB_Assumptions!$B$3",
        )
        ws.cell(row=excel_row, column=5, value=f"=C{excel_row}+D{excel_row}")
        ws.cell(row=excel_row, column=6, value=f"=Strategy_Summary!E{econ_row}")
        ws.cell(
            row=excel_row,
            column=7,
            value=f"=B{excel_row}*Post_TB_Assumptions!$B$4",
        )
        ws.cell(row=excel_row, column=8, value=f"=F{excel_row}+G{excel_row}")
        ws.cell(row=excel_row, column=9, value=f"=Strategy_Summary!C{econ_row}")
        ws.cell(
            row=excel_row,
            column=10,
            value=f"=B{excel_row}*Post_TB_Assumptions!$B$9*Post_TB_Assumptions!$B$10",
        )
        ws.cell(row=excel_row, column=11, value=f"=I{excel_row}-J{excel_row}")
        ws.cell(row=excel_row, column=12, value=f"=Strategy_Summary!F{econ_row}")
        ws.cell(row=excel_row, column=13, value=f"=IFERROR(K{excel_row}/E{excel_row},\"\")")
        ws.cell(row=excel_row, column=14, value=f"=Strategy_Summary!G{econ_row}")
        ws.cell(row=excel_row, column=15, value=f"=IFERROR(K{excel_row}/H{excel_row},\"\")")
    set_simple_widths(ws, [22, 18, 18, 18, 22, 18, 18, 22, 16, 18, 18, 18, 21, 18, 21])


def create_post_tb_chart_sheet(wb: Any) -> None:
    ws = wb.create_sheet("Post_TB_Chart_Data")
    headers = ["Strategy", "View", "Health gain", "Net cost", "ICER"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value).font = Font(bold=True)
    views = [
        ("Acute only DALYs", "D"),
        ("Acute + post-TB DALYs", "E"),
        ("Acute only QALYs", "F"),
        ("Acute + post-TB QALYs", "H"),
    ]
    row_idx = 2
    for source_col_idx in range(2, 6):
        strategy_row = source_col_idx
        ws.cell(row=row_idx, column=1, value=f"=Post_TB_Scenarios!A{strategy_row}")
        ws.cell(row=row_idx, column=2, value=views[0][0])
        ws.cell(row=row_idx, column=3, value=f"=Post_TB_Scenarios!C{strategy_row}")
        ws.cell(row=row_idx, column=4, value=f"=Post_TB_Scenarios!I{strategy_row}")
        ws.cell(row=row_idx, column=5, value=f"=Post_TB_Scenarios!L{strategy_row}")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=f"=Post_TB_Scenarios!A{strategy_row}")
        ws.cell(row=row_idx, column=2, value=views[1][0])
        ws.cell(row=row_idx, column=3, value=f"=Post_TB_Scenarios!E{strategy_row}")
        ws.cell(row=row_idx, column=4, value=f"=Post_TB_Scenarios!K{strategy_row}")
        ws.cell(row=row_idx, column=5, value=f"=Post_TB_Scenarios!M{strategy_row}")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=f"=Post_TB_Scenarios!A{strategy_row}")
        ws.cell(row=row_idx, column=2, value=views[2][0])
        ws.cell(row=row_idx, column=3, value=f"=Post_TB_Scenarios!F{strategy_row}")
        ws.cell(row=row_idx, column=4, value=f"=Post_TB_Scenarios!I{strategy_row}")
        ws.cell(row=row_idx, column=5, value=f"=Post_TB_Scenarios!N{strategy_row}")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=f"=Post_TB_Scenarios!A{strategy_row}")
        ws.cell(row=row_idx, column=2, value=views[3][0])
        ws.cell(row=row_idx, column=3, value=f"=Post_TB_Scenarios!H{strategy_row}")
        ws.cell(row=row_idx, column=4, value=f"=Post_TB_Scenarios!K{strategy_row}")
        ws.cell(row=row_idx, column=5, value=f"=Post_TB_Scenarios!O{strategy_row}")
        row_idx += 1
    set_simple_widths(ws, [22, 24, 16, 16, 16])


def create_post_tb_audit_sheet(wb: Any) -> None:
    ws = wb.create_sheet("Post_TB_Audit")
    headers = ["Audit item", "Check / source", "Notes"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=value).font = Font(bold=True)
    rows = [
        (
            "Scenario formulas present",
            "=ISFORMULA(Post_TB_Scenarios!D2)",
            "Should be TRUE.",
        ),
        (
            "Chart sheet formulas present",
            "=ISFORMULA(Post_TB_Chart_Data!C2)",
            "Should be TRUE.",
        ),
        (
            "Primary source recorded",
            "Post_TB_Assumptions!B2",
            "Byrne June 2026 talk should be recorded here.",
        ),
        (
            "APY-specific PTLD remains unresolved",
            "=Post_TB_Assumptions!D6=\"UNRESOLVED\"",
            "Should remain TRUE until APY-specific data are available.",
        ),
    ]
    for row_idx, values in enumerate(rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    set_simple_widths(ws, [34, 24, 72])


def set_simple_widths(ws: Any, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def build_post_tb_report(
    output_path: Path,
    post_tb_df: pd.DataFrame,
    assumptions: dict[str, Any],
) -> None:
    source_report = find_latest_word_report()
    xml = read_docx_document_xml(source_report)
    section_xml = build_post_tb_report_section(post_tb_df, assumptions)
    updated_xml = xml.replace("</w:body>", f"{section_xml}</w:body>", 1)
    write_docx_with_updated_document_xml(source_report, output_path, updated_xml)


def find_latest_word_report() -> Path:
    candidates = sorted(
        REPO_ROOT.glob("APY_LTBI_health_economics_report*.docx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No APY Word report source file was found.")
    for candidate in candidates:
        if candidate.name != POST_TB_REPORT_PATH.name:
            return candidate
    return candidates[0]


def read_docx_document_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read("word/document.xml").decode("utf-8")


def write_docx_with_updated_document_xml(
    source_path: Path,
    output_path: Path,
    document_xml: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(source_path) as src, zipfile.ZipFile(output_path, "w") as dst:
        for item in src.infolist():
            data = document_xml.encode("utf-8") if item.filename == "word/document.xml" else src.read(item.filename)
            dst.writestr(item, data)


def build_post_tb_report_section(
    post_tb_df: pd.DataFrame,
    assumptions: dict[str, Any],
) -> str:
    acute_prevent = select_view(post_tb_df, "prevent", "acute_only")
    daly_prevent = select_view(post_tb_df, "prevent", "acute_plus_post_tb_daly_tail")
    qaly_prevent = select_view(post_tb_df, "prevent", "acute_plus_post_tb_qaly_tail")
    paragraphs = [
        xml_paragraph("Post-TB morbidity and mortality sensitivity analysis", style="Heading1"),
        xml_paragraph(
            "Primary source: docs/source_material/Post TB Mortality & Standardised Follow up_Byrne June 2026.pptx.",
        ),
        xml_paragraph("Evidence summary", style="Heading2"),
        xml_paragraph(
            "The Byrne June 2026 talk emphasises that successful microbiological treatment does not mark the end of TB burden. The slide deck reports 58 million DALYs due to post-TB sequelae, representing about 47% of the combined incident-TB plus post-TB burden, and 12.1 DALYs per incident TB case overall with 5.8 DALYs occurring after TB treatment."
        ),
        xml_paragraph(
            "The same talk summarises substantial long-term mortality after TB treatment, including Romanowski et al. with pooled SMR 2.91 and an Australian cohort analysis with overall cause-mortality ratio 1.16 and most deaths within 10 years of diagnosis."
        ),
        xml_paragraph(
            "Clinical sequelae highlighted in the talk include fibrosis, bronchiectasis, COPD/emphysema, pulmonary hypertension, chronic respiratory failure, lung cancer, and broader mental health, renal, visual, hearing, and musculoskeletal impairment."
        ),
        xml_paragraph("Why the acute-only model may underestimate burden", style="Heading2"),
        xml_paragraph(
            "The current APY health-economic model counts direct acute TB morbidity, TB mortality, and treatment-related losses, but it does not yet include a post-TB tail for survivors. That means acute-only DALYs and QALYs are likely conservative when post-treatment disability and excess mortality persist beyond microbiological cure."
        ),
        xml_paragraph("Post-TB scenario results", style="Heading2"),
        xml_paragraph(
            f"For the prevention-targeted strategy, acute-only DALYs averted are {format_number(acute_prevent.get('totalDALYsAvertedIncludingPostTB'))}. Adding the post-TB DALY tail raises this to {format_number(daly_prevent.get('totalDALYsAvertedIncludingPostTB'))}. Acute-only QALYs gained are {format_number(acute_prevent.get('totalQALYsGainedIncludingPostTB'))}, and adding the post-TB QALY tail raises this to {format_number(qaly_prevent.get('totalQALYsGainedIncludingPostTB'))}."
        ),
        xml_paragraph(
            f"Under the default sensitivity settings, the model uses postTBDalysPerCase={assumptions['postTB']['postTBDalysPerCase']}, postTBQalysLostPerCase={assumptions['postTB']['postTBQalysLostPerCase']}, postTBDurationYears={assumptions['postTB']['postTBDurationYears']}, and postTBExcessMortalityMultiplier={assumptions['postTB']['postTBExcessMortalityMultiplier']}."
        ),
        xml_paragraph("Caveats", style="Heading2"),
        xml_paragraph(
            "These post-TB results are sensitivity scenarios layered onto prevented incident TB cases. They do not alter APY epidemiological equations or rerun the underlying simulation. Excess mortality is recorded as an evidence anchor, but no APY-specific survivor hazard model has yet been added."
        ),
        xml_paragraph(
            "APY-specific PTLD prevalence, duration, annual care cost, and longer-horizon survivor outcomes are not yet available. These unresolved assumptions remain highlighted below and in the workbook."
        ),
        xml_paragraph("Unresolved APY-specific assumptions", style="Heading2"),
        xml_paragraph_runs(
            [
                ("PTLD prevalence in APY TB survivors", True),
                ("; ", False),
                ("post-TB annual care cost", True),
                ("; ", False),
                ("numeric lifetime duration assumption", True),
                ("; ", False),
                ("how to map excess mortality into APY-specific survivor follow-up", True),
            ]
        ),
        xml_paragraph(
            "Statement: APY-specific PTLD data are not yet available, so these results should be interpreted as scenario bounds for planning rather than local estimates."
        ),
    ]
    return "".join(paragraphs)


def xml_paragraph(text: str, style: str | None = None) -> str:
    return xml_paragraph_runs([(text, False)], style=style)


def xml_paragraph_runs(
    runs: list[tuple[str, bool]],
    style: str | None = None,
) -> str:
    ppr = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ""
    return f"<w:p>{ppr}{''.join(xml_run(text, highlight) for text, highlight in runs)}</w:p>"


def xml_run(text: str, highlight: bool = False) -> str:
    rpr = "<w:rPr><w:highlight w:val=\"yellow\"/></w:rPr>" if highlight else ""
    needs_space = text.startswith(" ") or text.endswith(" ")
    xml_space = ' xml:space="preserve"' if needs_space else ""
    return f"<w:r>{rpr}<w:t{xml_space}>{escape(text)}</w:t></w:r>"


def select_view(post_tb_df: pd.DataFrame, strategy: str, view: str) -> dict[str, Any]:
    matched = post_tb_df[
        (post_tb_df["screeningStrategy"] == strategy) & (post_tb_df["view"] == view)
    ]
    if matched.empty:
        return {}
    return matched.iloc[0].to_dict()


def first_not_none(*values: Any) -> Any:
    for value in values:
        if value is not None and not pd.isna(value):
            return value
    return None


def format_number(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "NA"
    return f"{float(value):,.3f}".rstrip("0").rstrip(".")


def best_strategy(df: pd.DataFrame, metric_name: str) -> str:
    if metric_name not in df.columns:
        return "not calculable"
    values = df.dropna(subset=[metric_name])
    if values.empty:
        return "not calculable"
    return str(values.loc[values[metric_name].idxmin(), "screeningStrategy"])


if __name__ == "__main__":
    raise SystemExit(main())
