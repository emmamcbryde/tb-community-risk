from __future__ import annotations

import math
import sys
from itertools import product
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from engine.apy.age_distribution import broad_age_group_from_years
from engine.apy.calibration import (
    calibrate_from_config,
    disease_multiplier_from_flags,
)
from engine.apy.config import build_default_config
from engine.apy.cohort import infection_probability, preventable_active_risk
from engine.apy.runner import run_scenario_with_do_nothing
from engine.apy.simulation import simulate_one_cohort_from_config


OUTPUT_DIR = REPO_ROOT / "outputs" / "apy_cho_epi_outputs"
WORKBOOK_PATH = OUTPUT_DIR / "APY_CHO_epidemiology_outputs.xlsx"
WORKBOOK_V2_PATH = OUTPUT_DIR / "APY_CHO_epidemiology_outputs_v2.xlsx"
CURVE_CSV = OUTPUT_DIR / "apy_cho_epi_curve_data.csv"
TARGETING_CSV = OUTPUT_DIR / "apy_cho_targeting_order.csv"
STOPPING_CSV = OUTPUT_DIR / "apy_cho_stopping_rule_candidates.csv"
NOTES_PATH = OUTPUT_DIR / "apy_cho_epi_notes.md"
PREVALENCE_UPDATE_NOTES_PATH = OUTPUT_DIR / "apy_prevalence_update_notes.md"
PREVALENCE_GRID_CSV = OUTPUT_DIR / "apy_prevalence_scenario_grid.csv"
EXPECTED_EARLY_YIELD_CSV = OUTPUT_DIR / "apy_expected_early_yield.csv"

N = 1500
N_REPS = 2000
SEED = 1
SCREEN_WINDOW = 2
FOLLOW_HORIZON = 20
TEST_TYPE = "IGRA"
REGIMEN = "3HP"

MAIN_STRATEGIES = ["random", "ltbi", "prevent"]
CURE_AUDIT_STRATEGIES = ["ltbi", "cure"]
STRATEGY_LABELS = {
    "random": "Random screening",
    "ltbi": "LTBI-targeted screening",
    "prevent": "Active TB prevention-targeted screening",
    "cure": "Cure-targeted screening",
}
SCREEN_COUNT_GRID = [
    0,
    50,
    100,
    150,
    200,
    250,
    300,
    350,
    400,
    450,
    500,
    600,
    700,
    800,
    900,
    1000,
    1100,
    1200,
    1300,
    1400,
    1500,
]
PREVALENCE_SCENARIOS = {
    "base_default_LTBI": None,
    "low_LTBI_1pct": 0.01,
}
DEFAULT_LTBI_PREVALENCE = 47 / 624
DEFAULT_ACTIVE_TB_2Y_PREVALENCE = 10 / 770
CHECKPOINTS = [50, 100, 150, 200, 300]
STOPPING_THRESHOLDS = [0, 1, 2, 3, 4, 5, 10]
PREVALENCE_UPDATE_SCREEN_COUNTS = [0, 50, 100, 150, 200, 300, 450, 600, 1000, 1500]
PREVALENCE_UPDATE_CHECKPOINTS = [50, 100, 150, 200, 300]
PREVALENCE_UPDATE_SCENARIOS = [
    ("LTBI_0_5pct", 0.005, "0.5% LTBI"),
    ("LTBI_1pct", 0.01, "1% LTBI"),
    ("LTBI_2pct", 0.02, "2% LTBI"),
    ("LTBI_3pct", 0.03, "3% LTBI"),
    ("LTBI_5pct", 0.05, "5% LTBI"),
    ("LTBI_7_53pct", DEFAULT_LTBI_PREVALENCE, "7.53% LTBI"),
    ("LTBI_10pct", 0.10, "10% LTBI"),
]


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if CURVE_CSV.exists() and TARGETING_CSV.exists() and STOPPING_CSV.exists():
        curve_df = pd.read_csv(CURVE_CSV)
        targeting_profiles = pd.read_csv(TARGETING_CSV)
        stopping_df = pd.read_csv(STOPPING_CSV)
    else:
        curve_df, raw_by_scenario = run_curve_scenarios()
        targeting_profiles = build_targeting_order_profiles()
        stopping_df = build_stopping_rule_candidates(curve_df, raw_by_scenario)
        curve_df.to_csv(CURVE_CSV, index=False)
        targeting_profiles.to_csv(TARGETING_CSV, index=False)
        stopping_df.to_csv(STOPPING_CSV, index=False)
    cure_audit_df = run_cure_audit()
    example_cohort = build_example_cohort_ordering()
    decile_summary = build_rank_decile_summary(example_cohort)
    prevalence_grid_df = run_prevalence_update_grid()
    expected_early_yield_df = build_expected_early_yield(prevalence_grid_df)
    prevalence_grid_df.to_csv(PREVALENCE_GRID_CSV, index=False)
    expected_early_yield_df.to_csv(EXPECTED_EARLY_YIELD_CSV, index=False)
    try:
        write_workbook(
            curve_df,
            cure_audit_df,
            targeting_profiles,
            example_cohort,
            decile_summary,
            stopping_df,
        )
    except PermissionError as exc:
        print(f"Skipped rewriting locked v1 workbook: {exc}")
    write_notes(curve_df, cure_audit_df)
    write_workbook_v2(
        curve_df,
        cure_audit_df,
        targeting_profiles,
        example_cohort,
        decile_summary,
        stopping_df,
        prevalence_grid_df,
        expected_early_yield_df,
    )
    write_prevalence_update_notes()

    print(f"Wrote {CURVE_CSV} rows={len(curve_df)}")
    print(f"Wrote {TARGETING_CSV} rows={len(targeting_profiles)}")
    print(f"Wrote {STOPPING_CSV} rows={len(stopping_df)}")
    print(f"Wrote {WORKBOOK_PATH}")
    print(f"Wrote {NOTES_PATH}")
    print(f"Wrote {WORKBOOK_V2_PATH}")
    print(f"Wrote {PREVALENCE_UPDATE_NOTES_PATH}")
    print(f"Prevalence update grid rows={len(prevalence_grid_df)}")
    print(cure_audit_df.to_string(index=False))
    return 0


def run_curve_scenarios() -> tuple[pd.DataFrame, dict[tuple[str, str, int], pd.DataFrame]]:
    rows: list[dict[str, Any]] = []
    raw_by_scenario: dict[tuple[str, str, int], pd.DataFrame] = {}
    for prevalence_name, ltbi_prevalence in PREVALENCE_SCENARIOS.items():
        for strategy in MAIN_STRATEGIES:
            for screen_count in SCREEN_COUNT_GRID:
                config = scenario_config(strategy, screen_count, ltbi_prevalence)
                print(
                    f"Running {prevalence_name} {strategy} screened={screen_count}",
                    flush=True,
                )
                model_out = run_scenario_with_do_nothing(config)
                raw_by_scenario[(prevalence_name, strategy, screen_count)] = model_out[
                    "results"
                ]["raw"]
                rows.append(
                    build_curve_row(
                        prevalence_name,
                        strategy,
                        screen_count,
                        config,
                        model_out,
                    )
                )
    return pd.DataFrame(rows), raw_by_scenario


def run_cure_audit() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for strategy in CURE_AUDIT_STRATEGIES:
        for screen_count in SCREEN_COUNT_GRID:
            model_out = run_scenario_with_do_nothing(
                scenario_config(strategy, screen_count, None)
            )
            rows.append(
                {
                    "screeningStrategy": strategy,
                    "screen_count_requested": screen_count,
                    "nScreened_median": summary_value(model_out, "nScreened"),
                    "nTestPositiveNonActive_median": summary_value(
                        model_out, "nTestPositiveNonActive"
                    ),
                    "nCuredInfection_median": summary_value(
                        model_out, "nCuredInfection"
                    ),
                    "nPreventedActiveTB_median": summary_value(
                        model_out, "nPreventedActiveTB"
                    ),
                    "nFalsePositiveTests_median": summary_value(
                        model_out, "nFalsePositiveTests"
                    ),
                }
            )
    audit = pd.DataFrame(rows)
    wide = audit.pivot(index="screen_count_requested", columns="screeningStrategy")
    diffs = []
    for metric in [
        "nTestPositiveNonActive_median",
        "nCuredInfection_median",
        "nPreventedActiveTB_median",
        "nFalsePositiveTests_median",
    ]:
        diff = (wide[(metric, "cure")] - wide[(metric, "ltbi")]).abs()
        diffs.append(
            {
                "metric": metric,
                "max_abs_difference": float(diff.max()),
                "mean_abs_difference": float(diff.mean()),
            }
        )
    return pd.DataFrame(diffs)


def run_prevalence_update_grid() -> pd.DataFrame:
    if PREVALENCE_GRID_CSV.exists():
        return pd.read_csv(PREVALENCE_GRID_CSV)
    rows: list[dict[str, Any]] = []
    for scenario_id, ltbi_prevalence, prevalence_label in PREVALENCE_UPDATE_SCENARIOS:
        for strategy in MAIN_STRATEGIES:
            for screen_count in PREVALENCE_UPDATE_SCREEN_COUNTS:
                print(
                    f"Running prevalence update {prevalence_label} {strategy} screened={screen_count}",
                    flush=True,
                )
                config = scenario_config(strategy, screen_count, ltbi_prevalence)
                model_out = run_scenario_with_do_nothing(config)
                row = build_curve_row(scenario_id, strategy, screen_count, config, model_out)
                rows.append(
                    {
                        "prevalence_scenario": scenario_id,
                        "ltbi_prevalence": ltbi_prevalence,
                        "prevalence_label": prevalence_label,
                        "screeningStrategy": strategy,
                        "strategy_label": STRATEGY_LABELS[strategy],
                        "screen_count_requested": screen_count,
                        "screenCoverage": screen_count / N,
                        "expected_test_positive_nonactive": row[
                            "nTestPositiveNonActive_median"
                        ],
                        "expected_true_positive_tests": row[
                            "nTruePositiveTests_median"
                        ],
                        "expected_false_positive_tests": row[
                            "nFalsePositiveTests_median"
                        ],
                        "expected_PPV": row["PPV_trueAmongTestPositive"],
                        "expected_false_positive_share": row[
                            "falsePositiveShareAmongTestPositive"
                        ],
                        "expected_active_TB_prevented": row[
                            "nPreventedActiveTB_median"
                        ],
                        "expected_NNS_prevent_TB": row[
                            "NNS_prevent_one_active_TB"
                        ],
                    }
                )
                pd.DataFrame(rows).to_csv(PREVALENCE_GRID_CSV, index=False)
    return pd.DataFrame(rows)


def build_expected_early_yield(prevalence_grid_df: pd.DataFrame) -> pd.DataFrame:
    return prevalence_grid_df[
        prevalence_grid_df["screen_count_requested"].isin(PREVALENCE_UPDATE_CHECKPOINTS)
    ].copy()


def scenario_config(
    strategy: str,
    screen_count: int,
    ltbi_prevalence: float | None,
) -> dict[str, Any]:
    config = build_default_config()
    config.update(
        {
            "scenarioLabel": f"CHO epidemiology - {strategy} - {screen_count}",
            "N": N,
            "nReps": N_REPS,
            "seed": SEED,
            "screenWindow": SCREEN_WINDOW,
            "followHorizon": FOLLOW_HORIZON,
            "screenCoverage": screen_count / N,
            "screeningStrategy": strategy,
            "testType": TEST_TYPE,
            "regimen": REGIMEN,
            "pStartTPT": 0.85,
            "regimenPComplete": 0.80,
            "regimenADRstop": 0.05,
            "regimenEffFull": 0.85,
        }
    )
    if ltbi_prevalence is not None:
        config["ltbiPrevalence"] = float(ltbi_prevalence)
        # The default active-TB target is infeasible when LTBI prevalence is
        # forced to 1%. Scale the active target proportionally so this remains
        # a coherent sensitivity rather than a new APY base case.
        config["activeTBPrevalence"] = (
            DEFAULT_ACTIVE_TB_2Y_PREVALENCE
            * float(ltbi_prevalence)
            / DEFAULT_LTBI_PREVALENCE
        )
    return config


def build_curve_row(
    prevalence_name: str,
    strategy: str,
    screen_count: int,
    config: dict[str, Any],
    model_out: dict[str, Any],
) -> dict[str, Any]:
    dynamic = model_out["bundle"]["technical"]["dynamicComparison"]
    med = lambda name: summary_value(model_out, name)
    low = lambda name: summary_value(model_out, name, "Low95")
    high = lambda name: summary_value(model_out, name, "High95")

    n_test_positive = med("nTestPositiveNonActive")
    n_false_positive = med("nFalsePositiveTests")
    n_started = med("nTotalCoursesStarted")
    n_false_positive_treated = med("nFalsePositiveTreated")
    n_true_positive = n_test_positive - n_false_positive
    n_true_positive_treated = n_started - n_false_positive_treated
    n_prevented = med("nPreventedActiveTB")

    return {
        "prevalence_scenario": prevalence_name,
        "screeningStrategy": strategy,
        "strategy_label": STRATEGY_LABELS[strategy],
        "screen_count_requested": screen_count,
        "screenCoverage": config["screenCoverage"],
        "nScreened_median": med("nScreened"),
        "nScreened_low95": low("nScreened"),
        "nScreened_high95": high("nScreened"),
        "nTestPositiveNonActive_median": n_test_positive,
        "nFalsePositiveTests_median": n_false_positive,
        "nFalsePositiveTreated_median": n_false_positive_treated,
        "nTotalCoursesStarted_median": n_started,
        "nTotalCoursesCompleted_median": med("nTotalCoursesCompleted"),
        "nCuredInfection_median": med("nCuredInfection"),
        "nPreventedActiveTB_median": n_prevented,
        "nActiveBy20y_DoNothing_median": dynamic.get(
            "cumulative_baseline_active_tb_cases"
        ),
        "nActiveBy20y_AfterStrategy_median": dynamic.get(
            "cumulative_intervention_active_tb_cases"
        ),
        "relativeReduction20y_median": dynamic.get(
            "relative_reduction_cumulative_active_tb_cases"
        ),
        "nIGRApos_median": med("nIGRApos"),
        "nTestPositive_median": med("nTestPositive"),
        "nScreenedBCG_median": med("nScreenedBCG"),
        "nFalsePositiveTestsBCG_median": med("nFalsePositiveTestsBCG"),
        "nFalsePositiveTestsNoBCG_median": med("nFalsePositiveTestsNoBCG"),
        "nTruePositiveTests_median": n_true_positive,
        "nTruePositiveTreated_median": n_true_positive_treated,
        "falsePositiveShareAmongTestPositive": safe_div(
            n_false_positive, n_false_positive + n_true_positive
        ),
        "PPV_trueAmongTestPositive": safe_div(
            n_true_positive, n_false_positive + n_true_positive
        ),
        "falsePositiveTreatedShareAmongCoursesStarted": safe_div(
            n_false_positive_treated, n_started
        ),
        "falsePositiveTestsPerTruePositiveTest": safe_div(
            n_false_positive, n_true_positive
        ),
        "falsePositiveTreatedPerTruePositiveTreated": safe_div(
            n_false_positive_treated, n_true_positive_treated
        ),
        "NNS_prevent_one_active_TB": safe_div(med("nScreened"), n_prevented),
        "NNT_started_prevent_one_active_TB": safe_div(n_started, n_prevented),
        "NNT_completed_prevent_one_active_TB": safe_div(
            med("nTotalCoursesCompleted"), n_prevented
        ),
        "notes": (
            "Low LTBI scenario uses ltbiPrevalence=0.01 sensitivity."
            if prevalence_name == "low_LTBI_1pct"
            else "Default APY LTBI calibration."
        ),
    }


def summary_value(
    model_out: dict[str, Any],
    metric: str,
    column: str = "Median",
) -> float:
    summary = model_out["results"]["summary"]
    matched = summary.loc[summary["Metric"] == metric, column]
    return math.nan if matched.empty else float(matched.iloc[0])


def safe_div(numerator: float, denominator: float) -> float:
    if denominator is None or pd.isna(denominator) or float(denominator) == 0:
        return math.nan
    return float(numerator) / float(denominator)


def build_targeting_order_profiles() -> pd.DataFrame:
    config = scenario_config("prevent", 0, None)
    calibration = calibrate_from_config(config)
    pars = calibration["parameters"]
    rows: list[dict[str, Any]] = []
    for age in pars["exactAgeValues"]:
        age_band = age_band_from_age(age)
        for mj, contact, renal, diabetes, smoking, cld, alcohol in product(
            [False, True], repeat=7
        ):
            p_ltbi = float(
                infection_probability(
                    age,
                    pars,
                    calibration["ageInfLogLambda"],
                    calibration["ageInfGamma"],
                    mj,
                    contact,
                    renal,
                )
            )
            disease_mult = disease_multiplier_from_flags(
                pars, mj, contact, renal, diabetes, smoking, cld, alcohol
            )
            prevention_score = p_ltbi * float(
                preventable_active_risk(
                    disease_mult,
                    calibration["lambdaEarly"],
                    calibration["lambdaLate"],
                    SCREEN_WINDOW,
                    FOLLOW_HORIZON,
                )
            )
            rows.append(
                {
                    "age": age,
                    "age_band": age_band,
                    "marijuana_use": mj,
                    "close_contact": contact,
                    "renal_disease": renal,
                    "diabetes": diabetes,
                    "smoking": smoking,
                    "chronic_lung_disease": cld,
                    "alcohol_drugs": alcohol,
                    "pLTBI": p_ltbi,
                    "LTBI risk score": p_ltbi,
                    "disease multiplier": disease_mult,
                    "prevention score": prevention_score,
                }
            )
    profiles = pd.DataFrame(rows)
    profiles["LTBI targeting priority"] = profiles["LTBI risk score"].rank(
        method="first", ascending=False
    )
    profiles["prevention targeting priority"] = profiles["prevention score"].rank(
        method="first", ascending=False
    )
    ltbi_top = profiles.nsmallest(100, "LTBI targeting priority").copy()
    ltbi_top["targeting_list"] = "top_100_LTBI_targeted"
    prevent_top = profiles.nsmallest(100, "prevention targeting priority").copy()
    prevent_top["targeting_list"] = "top_100_prevention_targeted"
    out = pd.concat([ltbi_top, prevent_top], ignore_index=True)
    out.insert(0, "rank", out.groupby("targeting_list").cumcount() + 1)
    out["notes"] = (
        "Model-derived risk profile ordering. This is not a list of real people."
    )
    return out


def build_example_cohort_ordering() -> pd.DataFrame:
    config = scenario_config("prevent", N, None)
    cohort = simulate_one_cohort_from_config(config, n=N, seed=SEED)["cohort"].copy()
    cohort = cohort.rename(columns={"id": "example_id", "pInfection": "pLTBI"})
    cohort["age band"] = cohort["ageYears"].map(age_band_from_age)
    cohort["LTBI rank"] = cohort["ltbiRiskScore"].rank(method="first", ascending=False)
    cohort["prevention rank"] = cohort["preventTargetScore"].rank(
        method="first", ascending=False
    )
    return cohort[
        [
            "example_id",
            "ageYears",
            "age band",
            "MJ",
            "contact",
            "renal",
            "diabetes",
            "smoking",
            "chronicLungDisease",
            "alcoholDrugs",
            "pLTBI",
            "ltbiRiskScore",
            "diseaseMultiplier",
            "preventTargetScore",
            "LTBI rank",
            "prevention rank",
        ]
    ].sort_values("prevention rank")


def build_rank_decile_summary(example: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    mappings = {
        "LTBI-targeted screening": ("LTBI rank", "ltbiRiskScore"),
        "Active TB prevention-targeted screening": (
            "prevention rank",
            "preventTargetScore",
        ),
    }
    risk_cols = [
        "contact",
        "MJ",
        "renal",
        "diabetes",
        "smoking",
        "chronicLungDisease",
        "alcoholDrugs",
    ]
    for strategy_label, (rank_col, score_col) in mappings.items():
        ordered = example.sort_values(rank_col).reset_index(drop=True)
        ordered["decile"] = pd.qcut(
            ordered.index + 1, 10, labels=False, duplicates="drop"
        ) + 1
        for decile, sub in ordered.groupby("decile"):
            dominant = [
                col for col in risk_cols if float(sub[col].astype(float).mean()) >= 0.2
            ]
            rows.append(
                {
                    "strategy": strategy_label,
                    "decile": int(decile),
                    "rank_start": int(sub[rank_col].min()),
                    "rank_end": int(sub[rank_col].max()),
                    "median age": float(sub["ageYears"].median()),
                    "dominant risk factors": ", ".join(dominant) or "none >=20%",
                    "expected pLTBI": float(sub["pLTBI"].mean()),
                    "expected disease multiplier": float(
                        sub["diseaseMultiplier"].mean()
                    ),
                    "expected prevention score": float(sub[score_col].mean()),
                    "share with close contact": float(sub["contact"].mean()),
                    "share with marijuana use": float(sub["MJ"].mean()),
                    "share with renal disease": float(sub["renal"].mean()),
                    "share with diabetes": float(sub["diabetes"].mean()),
                    "share with smoking": float(sub["smoking"].mean()),
                    "share with chronic lung disease": float(
                        sub["chronicLungDisease"].mean()
                    ),
                    "share with alcohol/drugs": float(sub["alcoholDrugs"].mean()),
                }
            )
    return pd.DataFrame(rows)


def build_stopping_rule_candidates(
    curve_df: pd.DataFrame,
    raw_by_scenario: dict[tuple[str, str, int], pd.DataFrame],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for prevalence_name in PREVALENCE_SCENARIOS:
        for strategy in MAIN_STRATEGIES:
            for checkpoint in CHECKPOINTS:
                current = select_curve(curve_df, prevalence_name, strategy, checkpoint)
                to_450 = select_curve(curve_df, prevalence_name, strategy, 450)
                to_1500 = select_curve(curve_df, prevalence_name, strategy, 1500)
                rows.append(
                    {
                        "row_type": "checkpoint_summary",
                        "prevalence_scenario": prevalence_name,
                        "strategy": strategy,
                        "checkpoint_screened": checkpoint,
                        "min_test_positive_nonactive_to_continue": None,
                        "median_test_positive_nonactive": current[
                            "nTestPositiveNonActive_median"
                        ],
                        "median_true_positive_tests": current[
                            "nTruePositiveTests_median"
                        ],
                        "median_false_positive_tests": current[
                            "nFalsePositiveTests_median"
                        ],
                        "PPV": current["PPV_trueAmongTestPositive"],
                        "false_positive_share": current[
                            "falsePositiveShareAmongTestPositive"
                        ],
                        "projected_TB_cases_prevented_if_continue_to_450": to_450[
                            "nPreventedActiveTB_median"
                        ],
                        "projected_TB_cases_prevented_if_continue_to_1500": to_1500[
                            "nPreventedActiveTB_median"
                        ],
                        "NNS_prevent_TB_at_checkpoint": current[
                            "NNS_prevent_one_active_TB"
                        ],
                        "NNS_prevent_TB_if_continue_to_450": to_450[
                            "NNS_prevent_one_active_TB"
                        ],
                        "NNS_prevent_TB_if_continue_to_1500": to_1500[
                            "NNS_prevent_one_active_TB"
                        ],
                        "probability_yield_below_threshold": None,
                        "median_yield_above_threshold": None,
                        "expected_consequences_of_stopping": (
                            "Stopping here would forgo the projected additional "
                            "direct person-level TB prevention shown in the 450 and "
                            "1500 columns."
                        ),
                        "notes": (
                            "Stopping rules should be review triggers, not automatic "
                            "stop decisions."
                        ),
                    }
                )
            raw = raw_by_scenario[(prevalence_name, strategy, 100)]
            current_100 = select_curve(curve_df, prevalence_name, strategy, 100)
            to_450 = select_curve(curve_df, prevalence_name, strategy, 450)
            to_1500 = select_curve(curve_df, prevalence_name, strategy, 1500)
            for threshold in STOPPING_THRESHOLDS:
                rows.append(
                    {
                        "row_type": "threshold_at_100_screened",
                        "prevalence_scenario": prevalence_name,
                        "strategy": strategy,
                        "checkpoint_screened": 100,
                        "min_test_positive_nonactive_to_continue": threshold,
                        "median_test_positive_nonactive": current_100[
                            "nTestPositiveNonActive_median"
                        ],
                        "median_true_positive_tests": current_100[
                            "nTruePositiveTests_median"
                        ],
                        "median_false_positive_tests": current_100[
                            "nFalsePositiveTests_median"
                        ],
                        "PPV": current_100["PPV_trueAmongTestPositive"],
                        "false_positive_share": current_100[
                            "falsePositiveShareAmongTestPositive"
                        ],
                        "projected_TB_cases_prevented_if_continue_to_450": to_450[
                            "nPreventedActiveTB_median"
                        ],
                        "projected_TB_cases_prevented_if_continue_to_1500": to_1500[
                            "nPreventedActiveTB_median"
                        ],
                        "NNS_prevent_TB_at_checkpoint": current_100[
                            "NNS_prevent_one_active_TB"
                        ],
                        "NNS_prevent_TB_if_continue_to_450": to_450[
                            "NNS_prevent_one_active_TB"
                        ],
                        "NNS_prevent_TB_if_continue_to_1500": to_1500[
                            "NNS_prevent_one_active_TB"
                        ],
                        "probability_yield_below_threshold": float(
                            (raw["nTestPositiveNonActive"] < threshold).mean()
                        ),
                        "median_yield_above_threshold": bool(
                            current_100["nTestPositiveNonActive_median"] >= threshold
                        ),
                        "expected_consequences_of_stopping": (
                            "If stopping at 100 screened, median direct TB cases "
                            f"prevented would remain {current_100['nPreventedActiveTB_median']}; "
                            f"continuing to 450 gives {to_450['nPreventedActiveTB_median']} "
                            f"and continuing to 1500 gives {to_1500['nPreventedActiveTB_median']}."
                        ),
                        "notes": (
                            "Probability uses replicate-level model output for the "
                            "100-screened scenario; prospective validation still "
                            "requires operational data."
                        ),
                    }
                )
    return pd.DataFrame(rows)


def select_curve(
    curve_df: pd.DataFrame,
    prevalence_name: str,
    strategy: str,
    screen_count: int,
) -> dict[str, Any]:
    row = curve_df[
        (curve_df["prevalence_scenario"] == prevalence_name)
        & (curve_df["screeningStrategy"] == strategy)
        & (curve_df["screen_count_requested"] == screen_count)
    ]
    if row.empty:
        raise ValueError((prevalence_name, strategy, screen_count))
    return row.iloc[0].to_dict()


def write_workbook(
    curve_df: pd.DataFrame,
    cure_audit_df: pd.DataFrame,
    targeting_profiles: pd.DataFrame,
    example_cohort: pd.DataFrame,
    decile_summary: pd.DataFrame,
    stopping_df: pd.DataFrame,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_readme(wb)
    write_run_settings(wb, cure_audit_df)
    write_df(wb, "Curve_Data_All", curve_df)
    write_chart_data(wb, "Chart_Data_Base_LTBI", curve_df, "base_default_LTBI")
    write_chart_data(wb, "Chart_Data_Low_LTBI_1pct", curve_df, "low_LTBI_1pct")
    write_df(
        wb,
        "False_Positive_Data",
        curve_df[
            [
                "prevalence_scenario",
                "screeningStrategy",
                "strategy_label",
                "screen_count_requested",
                "nTestPositiveNonActive_median",
                "nFalsePositiveTests_median",
                "nTruePositiveTests_median",
                "falsePositiveShareAmongTestPositive",
                "PPV_trueAmongTestPositive",
                "nFalsePositiveTreated_median",
                "falsePositiveTreatedShareAmongCoursesStarted",
                "nScreenedBCG_median",
                "nFalsePositiveTestsBCG_median",
                "nFalsePositiveTestsNoBCG_median",
            ]
        ],
    )
    write_df(wb, "Targeting_Order_Profile", targeting_profiles)
    write_df(wb, "Targeting_Order_Example_Cohort", example_cohort)
    write_df(wb, "Rank_Decile_Summary", decile_summary)
    write_df(wb, "Stopping_Rule_Candidates", stopping_df)
    write_definitions(wb)
    wb.save(WORKBOOK_PATH)


def write_workbook_v2(
    curve_df: pd.DataFrame,
    cure_audit_df: pd.DataFrame,
    targeting_profiles: pd.DataFrame,
    example_cohort: pd.DataFrame,
    decile_summary: pd.DataFrame,
    stopping_df: pd.DataFrame,
    prevalence_grid_df: pd.DataFrame,
    expected_early_yield_df: pd.DataFrame,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_readme_v2(wb)
    write_run_settings(wb, cure_audit_df)
    write_df(wb, "Curve_Data_All", curve_df)
    write_chart_data(wb, "Chart_Data_Base_LTBI", curve_df, "base_default_LTBI")
    write_chart_data(wb, "Chart_Data_Low_LTBI_1pct", curve_df, "low_LTBI_1pct")
    write_df(
        wb,
        "False_Positive_Data",
        curve_df[
            [
                "prevalence_scenario",
                "screeningStrategy",
                "strategy_label",
                "screen_count_requested",
                "nTestPositiveNonActive_median",
                "nFalsePositiveTests_median",
                "nTruePositiveTests_median",
                "falsePositiveShareAmongTestPositive",
                "PPV_trueAmongTestPositive",
                "nFalsePositiveTreated_median",
                "falsePositiveTreatedShareAmongCoursesStarted",
                "nScreenedBCG_median",
                "nFalsePositiveTestsBCG_median",
                "nFalsePositiveTestsNoBCG_median",
            ]
        ],
    )
    write_df(wb, "Targeting_Order_Profile", targeting_profiles)
    write_df(wb, "Targeting_Order_Example_Cohort", example_cohort)
    write_df(wb, "Rank_Decile_Summary", decile_summary)
    write_df(wb, "Stopping_Rule_Candidates_Legacy", stopping_df)
    write_df(wb, "Prevalence_Scenario_Grid", prevalence_grid_df)
    write_df(wb, "Expected_Early_Yield", expected_early_yield_df)
    write_prevalence_update_sheet(wb)
    write_updated_future_benefit_sheet(wb)
    write_definitions_v2(wb)
    wb.save(WORKBOOK_V2_PATH)


def write_readme_v2(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ("Purpose", "CHO-facing epidemiological output workbook with sequential prevalence/yield updating."),
        ("Scope", "Epidemiological outputs only; no costs or health-economic formulas."),
        ("Planning axis", "Actual number screened, from 0 to 1500."),
        ("Main strategies", "Random screening; LTBI-targeted screening; Active TB prevention-targeted screening."),
        ("Sequential update", "Observed early non-active test positives update posterior weights over LTBI prevalence/yield scenarios."),
        ("Review language", "Outputs are pause-and-review triggers, not automatic stop rules."),
        ("Cure strategy", "Omitted from main outputs; audit comparison is in Run_Settings."),
        ("Transmission", "Direct person-level TB prevention only; no downstream transmission benefits."),
        ("Use guardrail", "Supports planning and sequencing, not denying screening or treatment."),
    ]
    write_key_value_rows(ws, rows)


def write_prevalence_update_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Prevalence_Update")
    ws["A1"] = "Editable inputs"
    ws["A1"].font = Font(bold=True)
    inputs = [
        ("strategy used", "ltbi"),
        ("checkpoint_screened", 100),
        ("observed_test_positive_nonactive", 2),
        ("posterior probability threshold", 0.80),
        ("NNS threshold", 100),
        ("false-positive share threshold", 0.50),
        ("minimum additional TB cases prevented to 450", 1),
    ]
    for row_idx, (label, value) in enumerate(inputs, start=2):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
    ws["D1"] = "Use strategy values: random, ltbi, prevent."
    ws["D2"] = "Observed count should be practical non-active test positives."

    start = 12
    headers = [
        "prevalence_scenario",
        "ltbi_prevalence",
        "prevalence_label",
        "prior_weight",
        "expected_positive_rate",
        "likelihood_of_observed_yield",
        "unnormalized_posterior",
        "posterior_weight",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    first = start + 1
    last = first + len(PREVALENCE_UPDATE_SCENARIOS) - 1
    for row_idx, (scenario_id, prevalence, label) in enumerate(
        PREVALENCE_UPDATE_SCENARIOS, start=first
    ):
        ws.cell(row_idx, 1, scenario_id)
        ws.cell(row_idx, 2, prevalence)
        ws.cell(row_idx, 3, label)
        ws.cell(row_idx, 4, 1 / len(PREVALENCE_UPDATE_SCENARIOS))
        ws.cell(
            row_idx,
            5,
            (
                f'=MAX(0.000001,MIN(0.999999,'
                f'SUMIFS(Prevalence_Scenario_Grid!$H:$H,'
                f'Prevalence_Scenario_Grid!$A:$A,$A{row_idx},'
                f'Prevalence_Scenario_Grid!$D:$D,$B$2,'
                f'Prevalence_Scenario_Grid!$F:$F,$B$3)/$B$3))'
            ),
        )
        ws.cell(
            row_idx,
            6,
            (
                f'=IF($B$4>$B$3,0,EXP(GAMMALN($B$3+1)-GAMMALN($B$4+1)'
                f'-GAMMALN($B$3-$B$4+1)+$B$4*LN(E{row_idx})'
                f'+($B$3-$B$4)*LN(1-E{row_idx})))'
            ),
        )
        ws.cell(row_idx, 7, f"=D{row_idx}*F{row_idx}")
        ws.cell(row_idx, 8, f"=IFERROR(G{row_idx}/SUM($G${first}:$G${last}),0)")

    summary_start = last + 3
    summary_rows = [
        ("posterior_mean_LTBI_prevalence", f"=SUMPRODUCT($B${first}:$B${last},$H${first}:$H${last})"),
        ("posterior_probability_LTBI_le_1pct", f'=SUMIFS($H${first}:$H${last},$B${first}:$B${last},"<=0.01")'),
        ("posterior_probability_LTBI_le_2pct", f'=SUMIFS($H${first}:$H${last},$B${first}:$B${last},"<=0.02")'),
        ("posterior_probability_LTBI_le_3pct", f'=SUMIFS($H${first}:$H${last},$B${first}:$B${last},"<=0.03")'),
        ("most_likely_prevalence_scenario", f"=INDEX($C${first}:$C${last},MATCH(MAX($H${first}:$H${last}),$H${first}:$H${last},0))"),
        ("review_if_posterior_prob_LTBI_le_1pct_gt_80pct", f'=IF(B{summary_start+1}>$B$5,"pause and review","continue with monitoring")'),
        ("review_if_posterior_prob_LTBI_le_2pct_gt_80pct", f'=IF(B{summary_start+2}>$B$5,"pause and review","continue with monitoring")'),
        ("review_if_updated_NNS_gt_threshold", '=IF(Updated_Future_Benefit!H2>$B$6,"pause and review","continue with monitoring")'),
        ("review_if_false_positive_share_gt_threshold", '=IF(Updated_Future_Benefit!F2>$B$7,"pause and review","continue with monitoring")'),
        ("review_if_additional_TB_prevented_to_450_lt_threshold", '=IF(Updated_Future_Benefit!I2<$B$8,"pause and review","continue with monitoring")'),
    ]
    ws.cell(summary_start - 1, 1, "Posterior outputs and review triggers").font = Font(bold=True)
    for row_idx, (label, formula) in enumerate(summary_rows, start=summary_start):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, formula)
    set_widths(ws)


def write_updated_future_benefit_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Updated_Future_Benefit")
    headers = [
        "continue_to_screened",
        "updated_expected_test_positive_nonactive",
        "updated_expected_false_positive_tests",
        "updated_expected_true_positive_tests",
        "updated_expected_PPV",
        "updated_expected_false_positive_share",
        "updated_expected_active_TB_prevented",
        "updated_NNS_prevent_TB",
        "updated_additional_TB_cases_prevented_from_checkpoint_to_450",
        "updated_additional_TB_cases_prevented_from_checkpoint_to_1500",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    checkpoint_active = formula_body(weighted_grid_formula("M", "Prevalence_Update!$B$3"))
    active_450 = formula_body(weighted_grid_formula("M", 450))
    active_1500 = formula_body(weighted_grid_formula("M", 1500))
    for row_idx, target in enumerate([450, 1500], start=2):
        ws.cell(row_idx, 1, target)
        ws.cell(row_idx, 2, weighted_grid_formula("H", target))
        ws.cell(row_idx, 3, weighted_grid_formula("J", target))
        ws.cell(row_idx, 4, f"=B{row_idx}-C{row_idx}")
        ws.cell(row_idx, 5, f'=IFERROR(D{row_idx}/B{row_idx},"")')
        ws.cell(row_idx, 6, f'=IFERROR(C{row_idx}/B{row_idx},"")')
        ws.cell(row_idx, 7, weighted_grid_formula("M", target))
        ws.cell(row_idx, 8, f'=IFERROR(A{row_idx}/G{row_idx},"")')
        ws.cell(row_idx, 9, f"=({active_450})-({checkpoint_active})")
        ws.cell(row_idx, 10, f"=({active_1500})-({checkpoint_active})")
    set_widths(ws)


def formula_body(formula: str) -> str:
    return formula[1:] if formula.startswith("=") else formula


def weighted_grid_formula(metric_column: str, screen_count: int | str) -> str:
    terms = []
    for idx in range(len(PREVALENCE_UPDATE_SCENARIOS)):
        update_row = 13 + idx
        screen_expr = screen_count if isinstance(screen_count, str) else str(screen_count)
        terms.append(
            f"Prevalence_Update!$H${update_row}*"
            f"SUMIFS(Prevalence_Scenario_Grid!${metric_column}:${metric_column},"
            f"Prevalence_Scenario_Grid!$A:$A,Prevalence_Update!$A${update_row},"
            f"Prevalence_Scenario_Grid!$D:$D,Prevalence_Update!$B$2,"
            f"Prevalence_Scenario_Grid!$F:$F,{screen_expr})"
        )
    return "=SUM(" + ",".join(terms) + ")"


def write_definitions_v2(wb: Workbook) -> None:
    write_definitions(wb)
    ws = wb["Definitions"]
    extra = [
        ("sequential prevalence update", "Uses observed early non-active test positives to update weights over candidate LTBI prevalence/yield scenarios."),
        ("prior weight", "Editable starting weight for each prevalence scenario before observing early programme yield."),
        ("likelihood", "Binomial probability of observed test-positive non-active count given the scenario-specific expected positive rate."),
        ("posterior weight", "Normalized prior times likelihood; weights sum to 1."),
        ("pause and review", "A governance trigger to review assumptions and operations, not an automatic stop rule."),
    ]
    start = ws.max_row + 1
    for row_idx, (term, definition) in enumerate(extra, start=start):
        ws.cell(row_idx, 1, term)
        ws.cell(row_idx, 2, definition)
    set_widths(ws)


def write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ("Purpose", "CHO-facing epidemiological output workbook for APY LTBI screening strategy planning."),
        ("Scope", "Epidemiological outputs only; no costs or health-economic formulas."),
        ("Planning axis", "Actual number screened, from 0 to 1500."),
        ("Main strategies", "Random screening; LTBI-targeted screening; Active TB prevention-targeted screening."),
        ("Cure strategy", "Omitted from main outputs; audit comparison is in Run_Settings."),
        ("Low LTBI scenario", "Sensitivity only: ltbiPrevalence=0.01, not an accepted APY base case unless locally reviewed."),
        ("Low LTBI calibration note", "The 2-year active TB prevalence target is scaled proportionally because the default active TB target is infeasible with 1% LTBI prevalence."),
        ("Transmission", "Direct person-level TB prevention only; no downstream transmission benefits."),
        ("Use guardrail", "Supports planning and sequencing, not denying screening or treatment."),
    ]
    write_key_value_rows(ws, rows)


def write_run_settings(wb: Workbook, cure_audit_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Run_Settings")
    rows = [
        ("N", N),
        ("nReps", N_REPS),
        ("seed", SEED),
        ("screenWindow", SCREEN_WINDOW),
        ("followHorizon", FOLLOW_HORIZON),
        ("testType", TEST_TYPE),
        ("regimen", REGIMEN),
        ("screen_count_grid", ", ".join(str(x) for x in SCREEN_COUNT_GRID)),
        ("prevalence_scenarios", ", ".join(PREVALENCE_SCENARIOS.keys())),
    ]
    write_key_value_rows(ws, rows)
    start = len(rows) + 4
    ws.cell(start, 1, "Cure vs LTBI audit").font = Font(bold=True)
    write_df_to_sheet(ws, cure_audit_df, start + 1)


def write_chart_data(
    wb: Workbook,
    title: str,
    curve_df: pd.DataFrame,
    prevalence_name: str,
) -> None:
    ws = wb.create_sheet(title)
    blocks = [
        ("NNS to prevent one active TB case", "NNS_prevent_one_active_TB"),
        ("Number of active TB cases prevented", "nPreventedActiveTB_median"),
        ("False-positive tests", "nFalsePositiveTests_median"),
        ("True-positive tests", "nTruePositiveTests_median"),
        ("False-positive share among test positives", "falsePositiveShareAmongTestPositive"),
        ("False-positive treated", "nFalsePositiveTreated_median"),
        ("False-positive treated share among TPT starts", "falsePositiveTreatedShareAmongCoursesStarted"),
        ("TPT starts", "nTotalCoursesStarted_median"),
        ("TPT completions", "nTotalCoursesCompleted_median"),
    ]
    current_row = 1
    for label, metric in blocks:
        ws.cell(current_row, 1, label).font = Font(bold=True)
        headers = ["number_screened"] + [STRATEGY_LABELS[s] for s in MAIN_STRATEGIES]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(current_row + 1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for offset, screen_count in enumerate(SCREEN_COUNT_GRID, start=1):
            ws.cell(current_row + 1 + offset, 1, screen_count)
            for col, strategy in enumerate(MAIN_STRATEGIES, start=2):
                row = select_curve(curve_df, prevalence_name, strategy, screen_count)
                ws.cell(current_row + 1 + offset, col, clean_cell(row[metric]))
        current_row += len(SCREEN_COUNT_GRID) + 4
    set_widths(ws)


def write_definitions(wb: Workbook) -> None:
    rows = [
        ("number screened", "Actual count selected for screening in the model scenario."),
        ("test positive non-active", "Positive test among people not active TB at screening; includes true and false positives."),
        ("false-positive test", "Uninfected simulated person with a positive test."),
        ("false-positive treated", "False-positive tested person who starts preventive treatment."),
        ("true-positive test", "nTestPositiveNonActive minus nFalsePositiveTests."),
        ("true-positive treated", "nTotalCoursesStarted minus nFalsePositiveTreated."),
        ("PPV", "True-positive tests divided by all non-active test positives."),
        ("FP/(FP+TP)", "False-positive tests divided by false-positive plus true-positive tests."),
        ("NNS to prevent one active TB case", "nScreened divided by nPreventedActiveTB."),
        ("NNT started to prevent one active TB case", "nTotalCoursesStarted divided by nPreventedActiveTB."),
        ("NNT completed to prevent one active TB case", "nTotalCoursesCompleted divided by nPreventedActiveTB."),
        ("LTBI-targeted strategy", "Ranks people by modelled LTBI risk score."),
        ("Active TB prevention-targeted strategy", "Ranks people by modelled preventable active TB risk."),
        ("why cure-targeting is omitted", "Excluded from main CHO workbook because it is operationally similar to LTBI-targeting in the current IGRA + 3HP pathway; audit values are retained in Run_Settings."),
        ("direct person-level active TB cases prevented", "Cases prevented among screened/treated people only."),
        ("downstream transmission", "Not included in these APY outputs."),
    ]
    ws = wb.create_sheet("Definitions")
    write_key_value_rows(ws, rows, headers=("Term", "Definition"))


def write_df(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title)
    write_df_to_sheet(ws, df, 1)


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int) -> None:
    for col, header in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, clean_cell(value))
    ws.freeze_panes = f"A{start_row + 1}"
    set_widths(ws)


def write_key_value_rows(
    ws,
    rows: list[tuple[Any, Any]],
    headers: tuple[str, str] = ("Item", "Value"),
) -> None:
    ws.cell(1, 1, headers[0]).font = Font(bold=True)
    ws.cell(1, 2, headers[1]).font = Font(bold=True)
    for row_idx, (key, value) in enumerate(rows, start=2):
        ws.cell(row_idx, 1, key)
        ws.cell(row_idx, 2, value)
    set_widths(ws)


def write_notes(curve_df: pd.DataFrame, cure_audit_df: pd.DataFrame) -> None:
    base = curve_df[curve_df["prevalence_scenario"] == "base_default_LTBI"]
    low = curve_df[curve_df["prevalence_scenario"] == "low_LTBI_1pct"]
    lines = [
        "# APY CHO Epidemiology Outputs",
        "",
        "Purpose: provide CHO-facing epidemiological outputs for APY LTBI screening strategy planning, with number screened as the x-axis and no cost calculations.",
        "",
        f"Settings: N={N}, nReps={N_REPS}, seed={SEED}, screenWindow={SCREEN_WINDOW}, followHorizon={FOLLOW_HORIZON}, testType={TEST_TYPE}, regimen={REGIMEN}.",
        "",
        "Cure-targeting is omitted from the main outputs because it is operationally similar to LTBI-targeting in the current pathway. The audit comparison is retained in the workbook Run_Settings sheet.",
        "",
        "Chart-data sheets use number screened as the first column in every block. The strategy columns are Random screening, LTBI-targeted screening, and Active TB prevention-targeted screening.",
        "",
        "False positives are simulated uninfected people with positive tests. True positives are calculated as test-positive non-active minus false-positive tests. Model positives are not clinical diagnoses.",
        "",
        "The low LTBI sensitivity sets ltbiPrevalence=0.01. The active TB prevalence target is scaled proportionally because the unscaled default active TB target is infeasible with 1% LTBI prevalence. It is a sensitivity scenario, not an APY calibrated base case unless accepted through local epidemiological review.",
        "",
        "Targeting-order sheets are model-derived illustrative orderings and risk profiles, not lists of real people.",
        "",
        "Stopping rules should be used as review triggers, not automatic stop decisions. The 100-screened threshold rows include replicate-level probabilities from the model, but prospective validation would require operational data.",
        "",
        "Limitations: direct person-level prevention only; no downstream transmission benefits; stochastic medians; local prevalence and risk-factor assumptions still require review.",
        "",
        "Key base-case patterns:",
    ]
    for strategy in MAIN_STRATEGIES:
        sub = base[(base["screeningStrategy"] == strategy) & (base["screen_count_requested"] == 100)]
        full = base[(base["screeningStrategy"] == strategy) & (base["screen_count_requested"] == 1500)]
        if not sub.empty and not full.empty:
            lines.append(
                f"- {STRATEGY_LABELS[strategy]}: at 100 screened, PPV={sub.iloc[0]['PPV_trueAmongTestPositive']:.2f}; at 1500 screened, false-positive tests={full.iloc[0]['nFalsePositiveTests_median']:.0f}, active TB cases prevented={full.iloc[0]['nPreventedActiveTB_median']:.0f}."
            )
    lines.append("")
    lines.append("Low LTBI sensitivity impact at 100 screened:")
    for strategy in MAIN_STRATEGIES:
        sub = low[(low["screeningStrategy"] == strategy) & (low["screen_count_requested"] == 100)]
        if not sub.empty:
            lines.append(
                f"- {STRATEGY_LABELS[strategy]}: PPV={sub.iloc[0]['PPV_trueAmongTestPositive']:.2f}, false-positive share={sub.iloc[0]['falsePositiveShareAmongTestPositive']:.2f}."
            )
    lines.append("")
    lines.append("Cure vs LTBI audit:")
    for row in cure_audit_df.to_dict(orient="records"):
        lines.append(
            f"- {row['metric']}: max absolute difference={row['max_abs_difference']:.3g}, mean absolute difference={row['mean_abs_difference']:.3g}."
        )
    NOTES_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_prevalence_update_notes() -> None:
    lines = [
        "# APY CHO Prevalence/Yield Updating Notes",
        "",
        "The previous stopping-rule table treated early thresholds as fixed rules under the original prevalence assumption. That is not the intended programme interpretation: if early observed yield is much lower than expected, the more appropriate response is to update beliefs about prevalence, targeting performance, and operational fit.",
        "",
        "The v2 workbook adds a sequential prevalence/yield updating framework. The `Prevalence_Scenario_Grid` sheet contains modelled yield scenarios for 0.5%, 1%, 2%, 3%, 5%, 7.53%, and 10% LTBI prevalence. The `Prevalence_Update` sheet lets the user enter the strategy used, checkpoint screened, and observed test-positive non-active count.",
        "",
        "For each candidate prevalence scenario, the workbook calculates the expected positive rate and a binomial likelihood: observed_yield ~ Binomial(checkpoint_screened, expected_positive_rate). Prior weights are multiplied by these likelihoods and normalized to posterior weights that sum to 1.",
        "",
        "For targeted screening, low yield can mean lower community prevalence, poor match between operational targeting and modelled targeting, incorrect risk-factor assumptions, incomplete capture of the intended high-risk group, or implementation issues. It should not be interpreted automatically as evidence that screening is futile.",
        "",
        "Observed test-positive non-active count is the practical programme input. True-positive and false-positive counts are model-estimated quantities; they are not directly observed clinical diagnoses.",
        "",
        "The `Updated_Future_Benefit` sheet uses posterior weights to estimate expected outcomes if screening continues to 450 or 1500 people. Review-trigger outputs deliberately say `pause and review`, not `stop`.",
        "",
        "The low-prevalence scenarios scale the active TB target proportionally to keep calibration feasible. These are sensitivity scenarios, not accepted APY base cases unless reviewed locally.",
    ]
    PREVALENCE_UPDATE_NOTES_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value) or math.isinf(float(value)):
            return None
    except (TypeError, ValueError):
        return value
    return value


def set_widths(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 42)


def age_band_from_age(age: float) -> str:
    age = float(age)
    if age < 5:
        return "0-4"
    if age < 15:
        return "5-14"
    if age < 25:
        return "15-24"
    if age < 35:
        return "25-34"
    if age < 45:
        return "35-44"
    if age < 55:
        return "45-54"
    if age < 65:
        return "55-64"
    if age < 75:
        return "65-74"
    if age < 85:
        return "75-84"
    return "85+"


if __name__ == "__main__":
    raise SystemExit(main())
